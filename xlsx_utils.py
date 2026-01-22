# -*- coding: utf-8 -*-
"""Excel export helpers for unavailability."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl


def build_unavailability_xlsx(
    rows: List[Dict[str, str]],
    template_path: Path,
    out_path: Path,
    sheet_name: str = "Indisponibilita",
) -> Path:
    """Create a well-formatted XLSX following `unavailability_template.xlsx`."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Clear existing rows after header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Sort rows by date then doctor
    def _key(r):
        d = str(r.get("date",""))[:10]
        return (d, r.get("doctor",""))
    for r in sorted(rows, key=_key):
        doctor = r.get("doctor","")
        dstr = str(r.get("date",""))[:10]
        try:
            d = dt.date.fromisoformat(dstr)
            dcell = dt.datetime(d.year, d.month, d.day)
        except Exception:
            dcell = dstr
        ws.append([doctor, dcell, r.get("shift",""), r.get("note","")])

    wb.save(out_path)
    return out_path
