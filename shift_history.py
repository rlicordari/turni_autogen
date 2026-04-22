"""shift_history.py -- Parser per Excel turni definitivi e aggregazione statistiche.

Legge i file Excel finalizzati dal primario, estrae le assegnazioni giornaliere
per ogni colonna di turno e calcola statistiche per medico (totali, festivi,
feriali, sabati, domeniche).
"""

from __future__ import annotations

import json
import math
from datetime import datetime
from typing import Any, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string

# ── Colonne tracciate ──────────────────────────────────────────────────────
TRACKED_COLUMNS: dict[str, str] = {
    "C": "Reperibilità",
    "D": "UTIC mattina",
    "E": "Cardiologia mattina",
    "F": "Supporto 118",
    "G": "Riabilitazione",
    "H": "UTIC pomeriggio",
    "I": "Cardiologia pomeriggio",
    "J": "Notte",
    "K": "Letto",
    "L": "Padiglioni",
    "M": "Emodinamica 1",
    "N": "Emodinamica 2",
    "O": "Emodinamica 3",
    "P": "Emodinamica 4",
    "Q": "ECO base",
    "R": "ECOSTRESS/ETE",
    "S": "Ecosala",
    "T": "Interni",
    "U": "Contr.PM",
    "V": "Sala PM",
    "W": "Ergometria/CPET",
    "Y": "Ambulatori",
    "Z": "Vascolare",
    "AA": "SPOC",
    "AB": "Holter/Brugada/FA",
    "AC": "Scintigrafia",
}

# Pre-calcola indici colonna (1-based)
_COL_INDICES: dict[str, int] = {
    letter: column_index_from_string(letter) for letter in TRACKED_COLUMNS
}

# ── Festività italiane (mese, giorno) ─────────────────────────────────────
_ITALIAN_HOLIDAYS: set[tuple[int, int]] = {
    (1, 1),   # Capodanno
    (1, 6),   # Epifania
    (4, 25),  # Liberazione
    (5, 1),   # Festa del Lavoro
    (6, 2),   # Festa della Repubblica
    (8, 15),  # Ferragosto
    (11, 1),  # Ognissanti
    (12, 8),  # Immacolata
    (12, 25), # Natale
    (12, 26), # Santo Stefano
}

_DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Colonne "attive" (tutte tranne C=Reperibilità) per conteggio domeniche/sabati
_ACTIVE_COLUMNS = {k for k in TRACKED_COLUMNS if k != "C"}

# Colonne D/E/H/I per festivi DEHI
_DEHI_COLUMNS = {"D", "E", "H", "I"}

# Nomi medici canonici (title-case). Usati per normalizzare casing
# inconsistente nei file definitivi modificati a mano dal primario.
_CANONICAL_NAMES: dict[str, str] = {
    "allegra": "Allegra",
    "andò": "Andò",  # nota: "ando" senza accento gestito sotto
    "ando": "Andò",
    "calabrò": "Calabrò",
    "calabro": "Calabrò",
    "carciotto": "Carciotto",
    "cimino": "Cimino",
    "colarusso": "Colarusso",
    "crea": "Crea",
    "cusmà": "Cusmà",
    "cusma": "Cusmà",
    "d'angelo": "D'Angelo",
    "dangelo": "D'Angelo",
    "dattilo": "Dattilo",
    "de gregorio": "De Gregorio",
    "degregorio": "De Gregorio",
    "de luca": "De Luca",
    "deluca": "De Luca",
    "giusti": "Giusti",
    "grimaldi": "Grimaldi",
    "licordari": "Licordari",
    "manganaro": "Manganaro",
    "migliorato": "Migliorato",
    "pugliatti": "Pugliatti",
    "recupero": "Recupero",
    "saporito": "Saporito",
    "trio": "Trio",
    "virga": "Virga",
    "vizzari": "Vizzari",
    "zito": "Zito",
}

# Nomi da escludere completamente dallo storico (non sono medici del reparto
# oppure non fanno parte del pool turni).
_EXCLUDED_NAMES = {"Recupero"}

# Pattern che indicano note/commenti e non nomi medici
import re as _re
_NOTE_PATTERNS = _re.compile(
    r"^(spostati?|spostare|anticipat[io]|posticipat[io]|da spostare|note|n\.b\.|nb:)",
    _re.IGNORECASE,
)
# Pattern per stringhe che non sono nomi (frammenti numerici, date, ecc.)
_NOT_A_NAME = _re.compile(r"^[\d/\)\(\.\-\s]+$")


# ── Helpers ────────────────────────────────────────────────────────────────

def _is_holiday(dt: datetime) -> bool:
    """True se domenica o festività nazionale italiana."""
    return dt.weekday() == 6 or (dt.month, dt.day) in _ITALIAN_HOLIDAYS


def _normalize_name(raw: str) -> str | None:
    """Normalizza un nome medico grezzo.

    - Rimuove parentesi e contenuto tra parentesi (es. "Virga (anticipare il 17/04)" → "Virga")
    - Lookup case-insensitive nella tabella canonica
    - Fallback: title case
    - Ritorna None per note/commenti e nomi esclusi
    """
    s = raw.strip()
    if not s:
        return None
    # Rimuovi contenuto tra parentesi (anche non chiuse)
    s = _re.sub(r"\s*\(.*?\)", "", s).strip()
    s = _re.sub(r"\s*\(.*$", "", s).strip()  # parentesi aperta senza chiusura
    if not s:
        return None
    # Filtra frammenti non-nome (es. "04)", numeri, date)
    if _NOT_A_NAME.match(s):
        return None
    # Filtra note/commenti
    if _NOTE_PATTERNS.match(s):
        return None
    # Lookup canonico
    key = s.lower().strip()
    canonical = _CANONICAL_NAMES.get(key)
    if canonical:
        return canonical if canonical not in _EXCLUDED_NAMES else None
    # Fallback: title case per nomi non in tabella (emodinamisti esterni, etc.)
    result = s.title()
    return result if result not in _EXCLUDED_NAMES else None


def _parse_cell(value: Any) -> list[str]:
    """Estrae i nomi dei medici da una cella Excel.

    Gestisce: None, float NaN, stringhe singole, stringhe multilinea,
    nomi separati da "/" (es. "Grimaldi/Cimino"), casing inconsistente.
    Filtra note, commenti e nomi esclusi.
    """
    if value is None:
        return []
    if isinstance(value, float) and math.isnan(value):
        return []
    text = str(value).strip()
    if not text:
        return []
    # Split su newline
    parts = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    expanded: list[str] = []
    for p in parts:
        p = p.strip()
        # Split su " / " o "/" solo se sembra contenere nomi (non date come "17/04")
        # Euristica: split se almeno una parte contiene lettere
        if "/" in p and not _re.match(r"^\d+/\d+", p):
            sub = [s.strip() for s in p.split("/")]
            if all(any(c.isalpha() for c in s) for s in sub if s):
                expanded.extend(sub)
            else:
                expanded.append(p)
        else:
            expanded.append(p)
    result = []
    for p in expanded:
        name = _normalize_name(p)
        if name and name not in _EXCLUDED_NAMES:
            result.append(name)
    return result


# ── 1. Parser Excel ───────────────────────────────────────────────────────

def parse_finalized_xlsx(xlsx_path: str, sheet_name: str | None = None) -> dict:
    """Legge un file Excel di turni finalizzato e restituisce la struttura dati.

    Parameters
    ----------
    xlsx_path : str
        Percorso al file .xlsx.
    sheet_name : str | None
        Nome del foglio da leggere. Se None, usa il primo foglio
        (saltando 'Riepilogo' se presente).

    Returns
    -------
    dict con chiavi: year, month, month_label, days.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name:
            ws = wb[sheet_name]
        else:
            # Usa il primo foglio che non sia "Riepilogo"
            for name in wb.sheetnames:
                if name.lower() != "riepilogo":
                    ws = wb[name]
                    break
            else:
                ws = wb[wb.sheetnames[0]]

        days: list[dict] = []
        year: int | None = None
        month: int | None = None

        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row_idx, 1).value  # Colonna A
            if date_val is None:
                break
            if not isinstance(date_val, datetime):
                continue

            if year is None:
                year = date_val.year
                month = date_val.month

            is_hol = _is_holiday(date_val)
            dow = _DOW_NAMES[date_val.weekday()]

            assignments: dict[str, list[str]] = {}
            for col_letter, col_idx in _COL_INDICES.items():
                names = _parse_cell(ws.cell(row_idx, col_idx).value)
                if names:
                    assignments[col_letter] = names

            days.append({
                "date": date_val.strftime("%Y-%m-%d"),
                "dow": dow,
                "is_holiday": is_hol,
                "assignments": assignments,
            })

        return {
            "year": year,
            "month": month,
            "month_label": f"{year:04d}-{month:02d}" if year and month else "",
            "days": days,
        }
    finally:
        wb.close()


# ── 2. Statistiche per medico ─────────────────────────────────────────────

def compute_doctor_stats(parsed: dict) -> dict:
    """Calcola statistiche per medico dal risultato di parse_finalized_xlsx.

    Returns
    -------
    dict[str, dict] -- chiave = nome medico, valore = statistiche.
    Per ogni colonna: {total, festivi, feriali} (+ sabati/domeniche per J).
    Campi speciali: _festivi_DE_HI, _domeniche, _sabati.
    """
    stats: dict[str, dict] = {}

    def _ensure(doc: str) -> dict:
        if doc not in stats:
            stats[doc] = {"_festivi_DE_HI": 0, "_domeniche": 0, "_sabati": 0}
        return stats[doc]

    for day in parsed["days"]:
        is_hol = day["is_holiday"]
        is_sun = day["dow"] == "Sun"
        is_sat = day["dow"] == "Sat"
        assignments = day["assignments"]

        # Traccia medici attivi (non-C) per sabati/domeniche
        active_doctors_today: set[str] = set()
        # Traccia medici che hanno lavorato in D/E/H/I su questo giorno festivo
        # (un set per evitare doppio conteggio quando D e E, o H e I, hanno lo
        # stesso medico nel turno unico di festivo)
        dehi_doctors_today: set[str] = set()

        for col, names in assignments.items():
            for name in names:
                ds = _ensure(name)

                # Statistiche per colonna
                if col not in ds:
                    ds[col] = {"total": 0, "festivi": 0, "feriali": 0}
                bucket = ds[col]
                bucket["total"] += 1
                if is_hol:
                    bucket["festivi"] += 1
                else:
                    bucket["feriali"] += 1

                # Sottconteggi sabato/domenica per colonna J (Notte)
                if col == "J":
                    if "sabati" not in bucket:
                        bucket["sabati"] = 0
                    if "domeniche" not in bucket:
                        bucket["domeniche"] = 0
                    if is_sat:
                        bucket["sabati"] += 1
                    if is_sun:
                        bucket["domeniche"] += 1

                # Raccolta medici DEHI festivi (conteggio posticipato per dedup)
                if is_hol and col in _DEHI_COLUMNS:
                    dehi_doctors_today.add(name)

                # Raccolta medici attivi
                if col in _ACTIVE_COLUMNS:
                    active_doctors_today.add(name)

        # Conta festivi DEHI: un solo +1 per medico per giorno
        for name in dehi_doctors_today:
            _ensure(name)["_festivi_DE_HI"] += 1

        # Conta sabati/domeniche per medici attivi
        if is_sun:
            for name in active_doctors_today:
                _ensure(name)["_domeniche"] += 1
        if is_sat:
            for name in active_doctors_today:
                _ensure(name)["_sabati"] += 1

    return stats


# ── 3. Aggregazione multi-mese ────────────────────────────────────────────

def aggregate_multi_month(all_months: dict[str, dict]) -> dict:
    """Aggrega le statistiche di più mesi.

    Parameters
    ----------
    all_months : dict[str, dict]
        Chiave = month_label (es. "2026-04"), valore = output di compute_doctor_stats.

    Returns
    -------
    dict[str, dict] -- statistiche aggregate per medico, con _months_counted.
    """
    agg: dict[str, dict] = {}
    special_keys = {"_festivi_DE_HI", "_domeniche", "_sabati", "_months_counted"}

    for month_label, month_stats in all_months.items():
        for doc, ds in month_stats.items():
            if doc not in agg:
                agg[doc] = {
                    "_festivi_DE_HI": 0,
                    "_domeniche": 0,
                    "_sabati": 0,
                    "_months_counted": 0,
                }
            ad = agg[doc]
            ad["_months_counted"] += 1

            for key, val in ds.items():
                if key in special_keys:
                    # Somma scalari speciali
                    if key == "_months_counted":
                        continue
                    ad[key] = ad.get(key, 0) + val
                elif isinstance(val, dict):
                    # Colonna turno
                    if key not in ad:
                        ad[key] = {}
                    for subkey, subval in val.items():
                        ad[key][subkey] = ad[key].get(subkey, 0) + subval

    return agg


# ── 4. Serializzazione ────────────────────────────────────────────────────

def history_to_json(history: dict) -> str:
    """Serializza il dizionario di storico in stringa JSON."""
    return json.dumps(history, ensure_ascii=False, indent=2)


def history_from_json(text: str) -> dict:
    """Deserializza una stringa JSON in dizionario di storico."""
    return json.loads(text)


# ── 5. GitHub storage ────────────────────────────────────────────────────

HISTORY_PATH = "data/shift_history.json"


def load_history_from_github(
    owner: str,
    repo: str,
    token: str,
    branch: str = "main",
) -> Tuple[dict, Optional[str]]:
    """Carica lo storico turni da GitHub.

    Returns
    -------
    (history_dict, sha) — se il file non esiste restituisce ({}, None).
    """
    from github_utils import get_file

    gf = get_file(owner, repo, HISTORY_PATH, token, branch)
    if gf is None:
        return {}, None
    return history_from_json(gf.text), gf.sha


def save_history_to_github(
    history: dict,
    owner: str,
    repo: str,
    token: str,
    branch: str = "main",
    sha: Optional[str] = None,
) -> dict:
    """Salva lo storico turni su GitHub.

    Returns
    -------
    dict — risposta della GitHub Contents API.
    """
    from github_utils import put_file

    text = history_to_json(history)
    return put_file(
        owner,
        repo,
        HISTORY_PATH,
        token,
        "Aggiornamento memoria storica turni",
        text,
        branch,
        sha,
    )


def upload_month_to_history(
    xlsx_path: str,
    owner: str,
    repo: str,
    token: str,
    branch: str = "main",
    sheet_name: Optional[str] = None,
) -> Tuple[str, dict]:
    """Pipeline completa: parse xlsx → stats → load → upsert → save.

    Returns
    -------
    (month_label, month_stats)
    """
    parsed = parse_finalized_xlsx(xlsx_path, sheet_name)
    month_label: str = parsed["month_label"]
    month_stats = compute_doctor_stats(parsed)

    history, sha = load_history_from_github(owner, repo, token, branch)
    history[month_label] = month_stats
    save_history_to_github(history, owner, repo, token, branch, sha)

    return month_label, month_stats


# ── CLI ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Uso: python shift_history.py <file.xlsx> [sheet_name]")
        sys.exit(1)

    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    parsed = parse_finalized_xlsx(path, sheet)
    print(f"Mese: {parsed['month_label']}, Giorni: {len(parsed['days'])}")

    stats = compute_doctor_stats(parsed)
    for doc in sorted(stats):
        j_info = stats[doc].get("J", {})
        c_info = stats[doc].get("C", {})
        dom = stats[doc].get("_domeniche", 0)
        sat = stats[doc].get("_sabati", 0)
        fdehi = stats[doc].get("_festivi_DE_HI", 0)
        print(
            f"  {doc}: J={j_info.get('total',0)}, "
            f"C={c_info.get('total',0)}, "
            f"dom={dom}, sab={sat}, fest_DEHI={fdehi}"
        )
