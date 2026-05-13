#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Turni Autogenerator – UTIC/Cardiologia
- Legge un template Excel (openpyxl)
- Legge regole YAML
- Legge indisponibilità mensili (Excel/CSV) con colonne: Medico, Data, Fascia (Mattina/Pomeriggio/Notte/Diurno/Tutto il giorno)
- Compila le colonne operative e salva un nuovo file Excel
Solver principale: OR-Tools CP-SAT (pip install ortools)
Fallback: greedy + report conflitti (meno robusto)
Autore: prototype by ChatGPT (GPT-5.2 Thinking)
"""



from __future__ import annotations
__version__ = "2026-01-24-freecols-highlight-1"

import argparse
import dataclasses
import calendar
import datetime as dt
import re
import sys
from collections import defaultdict, Counter
from copy import deepcopy, copy as _copy
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Iterable
import pandas as pd
import yaml
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill
# -------------------------
# Utilities
# -------------------------
DOW_MAP = {
    0: "Mon",
    1: "Tue",
    2: "Wed",
    3: "Thu",
    4: "Fri",
    5: "Sat",
    6: "Sun",
}

# Optional style template used to make auto-generated monthly templates look
# like the official hospital model (header, column widths, weekend/holiday
# shading, etc.).
#
# Place a file named `Style_Template.xlsx` in the repo root (same folder as
# this script) to enable styling.
STYLE_TEMPLATE_FILENAME = "Style_Template.xlsx"
STYLE_TEMPLATE_FALLBACK = "Modello_Febbraio_2026.xlsx"  # backward-compat if you keep the old name


def _find_style_template() -> Optional[Path]:
    base = Path(__file__).resolve().parent
    p1 = base / STYLE_TEMPLATE_FILENAME
    if p1.exists():
        return p1
    p2 = base / STYLE_TEMPLATE_FALLBACK
    if p2.exists():
        return p2
    return None


def _load_style_ws() -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Load the first worksheet from the style template, if present."""
    p = _find_style_template()
    if not p:
        return None
    try:
        wb = openpyxl.load_workbook(p)
        return wb[wb.sheetnames[0]]
    except Exception:
        return None


def _easter_date_gregorian(year: int) -> dt.date:
    """Compute Easter Sunday (Gregorian calendar) using the Meeus/Jones/Butcher algorithm."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return dt.date(year, month, day)


def italy_public_holidays(year: int) -> Set[dt.date]:
    """Italian national public holidays (incl. Easter Monday)."""
    easter = _easter_date_gregorian(year)
    easter_monday = easter + dt.timedelta(days=1)
    fixed = {
        dt.date(year, 1, 1),   # Capodanno
        dt.date(year, 1, 6),   # Epifania
        dt.date(year, 4, 25),  # Liberazione
        dt.date(year, 5, 1),   # Lavoro
        dt.date(year, 6, 2),   # Repubblica
        dt.date(year, 8, 15),  # Ferragosto
        dt.date(year, 11, 1),  # Ognissanti
        dt.date(year, 12, 8),  # Immacolata
        dt.date(year, 12, 25), # Natale
        dt.date(year, 12, 26), # Santo Stefano
    }
    return fixed | {easter_monday}


def _is_grey_solid(cell) -> bool:
    """Return True if cell has the same grey solid fill used by the model."""
    try:
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            return False
        fg = getattr(fill.fgColor, "rgb", None)
        return fg == "FFC0C0C0"
    except Exception:
        return False


def _copy_cell_style(src, dst) -> None:
    """Copy style elements from src cell to dst cell (without copying value)."""
    try:
        # IMPORTANT: openpyxl style objects do not play well with deepcopy
        # (it can recurse). Use shallow copies or direct assignment.
        dst.font = _copy(src.font)
        dst.fill = _copy(src.fill)
        dst.border = _copy(src.border)
        dst.alignment = _copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = _copy(src.protection)
    except Exception:
        # best-effort: ignore style copy failures
        pass


def _apply_model_style_to_template(ws, cfg: dict, year: int, month: int, last_day: int) -> None:
    """Apply header/column widths and weekend/holiday shading using Style_Template.xlsx if available."""
    style_ws = _load_style_ws()
    if style_ws is None:
        return

    # Identify representative rows in the style worksheet
    # - sunday_row: first row where column B contains 'domenica' (Italian)
    # - weekday_row: first data row that is not sunday_row
    sunday_row = None
    weekday_row = None
    for r in range(2, min(style_ws.max_row, 20) + 1):
        b = style_ws.cell(r, 2).value
        if isinstance(b, str) and b.strip().lower().startswith("domen"):
            sunday_row = r
            break
    # fallback: assume row 2 is sunday in the provided model
    if sunday_row is None:
        sunday_row = 2
    # weekday row
    for r in range(2, min(style_ws.max_row, 20) + 1):
        if r == sunday_row:
            continue
        a = style_ws.cell(r, 1).value
        if isinstance(a, (dt.date, dt.datetime)):
            weekday_row = r
            break
    if weekday_row is None:
        weekday_row = min(3, style_ws.max_row)

    # Column widths + header styles/labels
    style_max_col = style_ws.max_column

    # The auto-template can declare columns beyond the style model (e.g. adding
    # extra "Medici liberi" columns AF/AG). In that case we still want the new
    # columns to look consistent: we extend styles by cloning from the last
    # available style column.
    needed_max_col = style_max_col
    try:
        cols_map = cfg.get("columns") or {}
        keep_empty = cfg.get("keep_empty_columns") or []
        cand_letters = []
        if isinstance(cols_map, dict):
            cand_letters.extend(list(cols_map.keys()))
        if isinstance(keep_empty, list):
            cand_letters.extend(list(keep_empty))
        for _cl in cand_letters:
            try:
                idx = column_index_from_string(str(_cl).strip().upper())
                if idx > needed_max_col:
                    needed_max_col = idx
            except Exception:
                pass
    except Exception:
        pass

    max_col = max(style_max_col, needed_max_col)
    # Ensure we have at least up to max_col in row 1
    _ = ws.cell(row=1, column=max_col)

    # Row heights
    if style_ws.row_dimensions[1].height:
        ws.row_dimensions[1].height = style_ws.row_dimensions[1].height
    if style_ws.row_dimensions[weekday_row].height:
        data_h = style_ws.row_dimensions[weekday_row].height
        for r in range(2, last_day + 2):
            ws.row_dimensions[r].height = data_h

    # Copy column widths and header style/value
    # For columns beyond the style model, clone from the last style column.
    ref_c = style_max_col
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        src_idx = c if c <= style_max_col else ref_c
        src_letter = get_column_letter(src_idx)
        w = style_ws.column_dimensions[src_letter].width
        if w:
            ws.column_dimensions[letter].width = w

        src_h = style_ws.cell(1, src_idx)
        dst_h = ws.cell(1, c)
        _copy_cell_style(src_h, dst_h)
        # Fill missing header labels from model (important for empty spacer columns)
        if (dst_h.value is None or str(dst_h.value).strip() == "") and (src_h.value is not None):
            dst_h.value = src_h.value

    # Determine which columns are shaded in the model on Sundays/holidays.
    # Extra columns (beyond the model) inherit the last-model-column shading.
    grey_cols = {c for c in range(1, style_max_col + 1) if _is_grey_solid(style_ws.cell(sunday_row, c))}
    try:
        if max_col > style_max_col and _is_grey_solid(style_ws.cell(sunday_row, ref_c)):
            grey_cols |= set(range(style_max_col + 1, max_col + 1))
    except Exception:
        pass

    # Holiday set for styling
    extra = set()
    for x in cfg.get("festivi_extra", []) or []:
        try:
            extra.add(parse_date(x))
        except Exception:
            pass
    holidays = italy_public_holidays(int(year)) | extra

    # Apply per-cell styles for the month (lightweight: <= 31 rows * ~31 cols)
    for r in range(2, last_day + 2):
        d = ws.cell(r, 1).value
        if isinstance(d, dt.datetime):
            d = d.date()
        if not isinstance(d, dt.date):
            continue
        is_holiday = (d.weekday() == 6) or (d in holidays)
        for c in range(1, max_col + 1):
            dst = ws.cell(r, c)
            src_idx = c if c <= style_max_col else ref_c
            # Choose style row based on holiday shading columns
            if is_holiday and (c in grey_cols):
                src = style_ws.cell(sunday_row, src_idx)
            else:
                src = style_ws.cell(weekday_row, src_idx)
            _copy_cell_style(src, dst)

SHIFT_NORMALIZE = {
    "mattina": "Mattina",
    "pom": "Pomeriggio",
    "pomeriggio": "Pomeriggio",
    "notte": "Notte",
}
def parse_date(x) -> dt.date:
    """Parse a date from Excel/str/datetime."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        raise ValueError("Empty date")
    if isinstance(x, dt.date) and not isinstance(x, dt.datetime):
        return x
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, (int, float)):
        # excel serial date: pandas handles it better; fallback not used here
        raise ValueError(f"Numeric date not supported directly: {x}")
    s = str(x).strip()
    # Accept dd/mm/yyyy or yyyy-mm-dd
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unrecognized date format: {x}")
def norm_name(x: str) -> str:
    return re.sub(r"\s+", " ", str(x).strip())
def norm_shift(x: str) -> str:
    s = str(x).strip().lower()
    if s in SHIFT_NORMALIZE:
        return SHIFT_NORMALIZE[s]
    # allow first letter
    if s.startswith("m"):
        return "Mattina"
    if s.startswith("p"):
        return "Pomeriggio"
    if s.startswith("n"):
        return "Notte"
    raise ValueError(f"Unrecognized shift: {x}")
def shifts_from_fascia(x: str) -> Set[str]:
    """Map unavailability 'Fascia' value to one or more internal shifts.
    Accepted:
      - Mattina / Pomeriggio / Notte  -> that single shift
      - Diurno (or Giorno)            -> {'Mattina','Pomeriggio'}
      - Tutto il giorno / All day     -> {'Any'} (treated as full-day)
      - Ferie / Vacanza / Holiday / Leave -> {'Any'} (treated as full-day)
    """
    s = str(x).strip().lower()
    # Full-day first (because it contains 'giorno')
    if any(k in s for k in ["tutto", "intera", "completa", "allday", "all day", "full day", "24h", "24 h", "ferie", "vacan", "holiday", "leave"]):
        return {"Any"}
    # Daytime (morning + afternoon, but still allows night)
    if s in ["diurno", "giorno", "daytime", "day"] or s.startswith("diurn"):
        return {"Mattina", "Pomeriggio"}
    # Default single-shift
    return {norm_shift(x)}
def dayspec_contains(dow: str, spec) -> bool:
    """
    spec can be:
    - string: 'Mon-Sat', 'Mon-Fri', 'Wed'
    - list of strings: ['Mon','Tue']
    - None: means all days
    """
    if spec is None:
        return True
    if isinstance(spec, str):
        spec = spec.strip()
        if "-" in spec:
            a, b = spec.split("-", 1)
            a, b = a.strip(), b.strip()
            order = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
            ia, ib = order.index(a), order.index(b)
            if ia <= ib:
                return order.index(dow) >= ia and order.index(dow) <= ib
            # wrap (rare)
            return order.index(dow) >= ia or order.index(dow) <= ib
        return dow == spec
    if isinstance(spec, list):
        return dow in spec
    raise TypeError(f"Invalid days spec: {spec}")
# -------------------------
# Data structures
# -------------------------
@dataclasses.dataclass(frozen=True)
class DayRow:
    date: dt.date
    dow: str  # 'Mon'...'Sun'
    row_idx: int
@dataclasses.dataclass
class Slot:
    """One assignable decision for a day."""
    day: DayRow
    slot_id: str                 # unique id
    columns: List[str]           # Excel column letters to fill with same doctor
    allowed: List[str]           # allowed doctor names (domain)
    required: bool = True        # if False, can be blank
    blank_penalty: int = 0      # penalty if left blank (only if required=False)
    shift: str = "Mattina"       # Mattina / Pomeriggio / Notte / Any
    rule_tag: str = ""           # for reporting
    empty_domain: bool = False    # True if allowed domain becomes empty after applying unavailability

    force_same_doctor: bool = False  # if True, this slot must use same doctor as its paired slot (e.g., D=F fallback)
    emergency_doctors: Optional[List[str]] = None  # medici di emergenza (non nel pool primario) — usati con penalità
# -------------------------
# Reperibilità (C) – final assignment layer
# -------------------------

def assign_reperibilita_C(cfg: dict, days: List[DayRow], slots: List[Slot],
                         assignment: Dict[str, Optional[str]]) -> Tuple[Dict[str, Optional[str]], Dict]:
    """Assign/Reassign Reperibilità (column C) as a final layer (STRICT).

    HARD rules (no automatic relaxation):
      - Exactly one doctor per day on column C (if C_reperibilita is configured).
      - C CAN overlap with other tasks on weekdays.
      - C is NOT allowed if:
          1) doctor has Night (J) on the same day
          2) doctor has Night (J) the previous day
          3) on Sundays/holidays, doctor is already working any other task that same day
      - Per doctor per month: min_per_doctor and max_per_doctor (typ. 2..3)
      - Minimum spacing between two C for the same doctor: spacing_min_days (typ. 3)

    If constraints are infeasible, this function raises ValueError with an actionable message.
    """
    rules = cfg.get("rules", {}) or {}
    rC = rules.get("C_reperibilita", {}) if isinstance(rules.get("C_reperibilita", {}), dict) else {}
    if not rC:
        return assignment, {"C_reperibilita_diag": {"status": "SKIPPED", "reason": "C_reperibilita not configured"}}

    constraints = set(rC.get("constraints") or [])
    excluded = {norm_name(x) for x in (rC.get("excluded") or [])}
    spacing_min = int(rC.get("spacing_min_days", 0) or 0)

    # Festivi pool: doctors preferred for C on festivo days (excluding those already
    # excluded from C, e.g. De Gregorio, Manganaro).
    # Doctors doing Notte that day are already blocked by the not_night_same_day constraint.
    rFest = rules.get("Festivi") if isinstance(rules.get("Festivi"), dict) else {}
    festivi_pool_norm = {norm_name(d) for d in (rFest.get("pool") or [])} - excluded
    min_per = int(rC.get("min_per_doctor", rC.get("target_per_doctor", 0) or 0) or 0)
    max_per = int(rC.get("max_per_doctor", 0) or 0)
    target = int(rC.get("target_per_doctor", 0) or 0)

    night_col = "J"  # fixed: Notte is column J

    # Map slots by day
    slots_by_day: Dict[dt.date, List[Slot]] = defaultdict(list)
    for s in slots:
        slots_by_day[s.day.date].append(s)

    # Identify C slots
    cslot_by_date: Dict[dt.date, Slot] = {}
    c_dates: List[dt.date] = []
    for d in days:
        cslot = next((s for s in slots_by_day.get(d.date, []) if s.columns == ["C"]), None)
        if cslot is not None:
            cslot_by_date[d.date] = cslot
            c_dates.append(d.date)

    if not c_dates:
        return assignment, {"C_reperibilita_diag": {"status": "SKIPPED", "reason": "No C slots in template"}}

    # Helper: does doctor work in any non-C slot the same day?
    def doctor_works_same_day(date_: dt.date, doc: str) -> bool:
        for s in slots_by_day.get(date_, []):
            if s.columns == ["C"]:
                continue
            if assignment.get(s.slot_id) == doc:
                return True
        return False

    # Day lookup
    dayrow_by_date = {d.date: d for d in days}

    # Allowed candidates per day
    allowed_norm_by_date: Dict[dt.date, List[str]] = {}
    pool_set: Set[str] = set()
    for d in c_dates:
        cand = []
        for raw in (cslot_by_date[d].allowed or []):
            dn = norm_name(raw)
            if not dn or dn in excluded or dn == "Recupero":
                continue
            cand.append(dn)
        # unique, stable
        seen=set(); cand2=[]
        for x in cand:
            if x not in seen:
                seen.add(x); cand2.append(x)
        allowed_norm_by_date[d] = cand2
        pool_set |= set(cand2)

    pool = sorted(pool_set, key=lambda s: s.lower())
    if not pool:
        raise ValueError("C_reperibilita: pool vuoto (tutti esclusi o non presenti nei pools YAML).")

    # Candidate filter applying hard constraints
    def ok_candidate(date_: dt.date, doc: str) -> bool:
        if doc not in set(allowed_norm_by_date.get(date_, [])):
            return False
        if "not_night_same_day" in constraints:
            if assignment.get(f"{date_}-{night_col}") == doc:
                return False
        if "not_night_prev_day" in constraints:
            prev = date_ - dt.timedelta(days=1)
            if assignment.get(f"{prev}-{night_col}") == doc:
                return False
        if "not_night_prev_2_days" in constraints:
            for delta in [1, 2]:
                prev = date_ - dt.timedelta(days=delta)
                if assignment.get(f"{prev}-{night_col}") == doc:
                    return False
        if "not_night_next_2_days" in constraints:
            for delta in [1, 2]:
                nxt = date_ + dt.timedelta(days=delta)
                if assignment.get(f"{nxt}-{night_col}") == doc:
                    return False
        if "not_working_same_day_on_sundays_and_holidays" in constraints:
            drow = dayrow_by_date.get(date_)
            if drow is not None and is_festivo(drow, cfg):
                if doctor_works_same_day(date_, doc):
                    return False
        return True

    # ── Funzione candidati con proximity J configurabile ──────────────────────
    def build_candidates_with_prox(night_prox: int) -> Dict[dt.date, List[str]]:
        """Costruisce candidates_by_date con finestra J±night_prox (0 = solo same-day)."""
        result: Dict[dt.date, List[str]] = {}
        for d in c_dates:
            cands = []
            for doc in pool:
                if doc not in set(allowed_norm_by_date.get(d, [])):
                    continue
                # not_night_same_day sempre attivo
                if assignment.get(f"{d}-{night_col}") == doc:
                    continue
                # Proximity J rilassabile
                blocked = False
                for delta in range(1, night_prox + 1):
                    if assignment.get(f"{(d - dt.timedelta(days=delta))}-{night_col}") == doc:
                        blocked = True; break
                    if assignment.get(f"{(d + dt.timedelta(days=delta))}-{night_col}") == doc:
                        blocked = True; break
                if blocked:
                    continue
                # Festivi: non assegnato nello stesso giorno
                if "not_working_same_day_on_sundays_and_holidays" in constraints:
                    drow = dayrow_by_date.get(d)
                    if drow is not None and is_festivo(drow, cfg):
                        if doctor_works_same_day(d, doc):
                            continue
                cands.append(doc)
            result[d] = cands
        return result

    # Costruisci candidati con proximity progressivamente rilassata
    _c_relaxation_warnings: List[str] = []
    candidates_by_date: Dict[dt.date, List[str]] = {}
    for _night_prox in [2, 1, 0]:
        candidates_by_date = build_candidates_with_prox(_night_prox)
        _empty_days = [d for d in c_dates if not candidates_by_date[d]]
        if not _empty_days:
            if _night_prox < 2:
                _c_relaxation_warnings.append(
                    f"C reperibilità: vincolo J proximity rilassato a ±{_night_prox} giorni "
                    f"per trovare candidati."
                )
            break
        if _night_prox == 0:
            # Ultimo resort: ignora del tutto la disponibilità per i giorni problematici
            for d in _empty_days:
                fallback = list(allowed_norm_by_date.get(d, [])) or pool[:]
                candidates_by_date[d] = fallback
                _c_relaxation_warnings.append(
                    f"C reperibilità: {d} senza candidati strict → usato pool completo ({len(fallback)} dottori)."
                )

    total_days = len(c_dates)
    n_docs = len(pool)

    # Feasibility checks for min/max (con auto-aumento max_per se necessario)
    if max_per <= 0:
        raise ValueError("C_reperibilita: max_per_doctor deve essere > 0.")
    if min_per < 0:
        min_per = 0

    if total_days > n_docs * max_per:
        # Auto-rilassa max_per invece di fallire
        max_per = (total_days + n_docs - 1) // n_docs  # ceil(total/n)
        _c_relaxation_warnings.append(
            f"C reperibilità: max_per_doctor aumentato a {max_per} per coprire {total_days} giorni con {n_docs} medici."
        )
    if min_per > 0 and total_days < n_docs * min_per:
        min_per = 0
        _c_relaxation_warnings.append(
            f"C reperibilità: min_per_doctor azzerato — troppi medici nel pool per il mese."
        )

    # Build desired counts: start at min_per, distribute remaining +1 up to max_per
    desired = {doc: (min_per if min_per > 0 else 0) for doc in pool}
    cur = sum(desired.values())
    remaining = total_days - cur
    order_docs = pool[:]
    while remaining > 0:
        progress = False
        for doc in order_docs:
            if desired[doc] < max_per and remaining > 0:
                desired[doc] += 1
                remaining -= 1
                progress = True
        if not progress:
            break
    if sum(desired.values()) != total_days:
        # Distribuisci il residuo ignorando max_per
        for doc in order_docs:
            if remaining <= 0:
                break
            desired[doc] += 1
            remaining -= 1

    # Backtracking DFS con retry su spacing rilassato
    assigned: Dict[dt.date, str] = {}

    def _run_dfs_attempt(sp: int) -> bool:
        _cnt = {doc: 0 for doc in pool}
        _dates_by_doc: Dict[str, List[dt.date]] = {doc: [] for doc in pool}
        _c_dates_sorted = sorted(c_dates, key=lambda d: (len(candidates_by_date[d]), d))

        def _spacing_ok(doc: str, d: dt.date) -> bool:
            if sp > 1:
                for prev in _dates_by_doc[doc]:
                    if abs((d - prev).days) < sp:
                        return False
            return True

        def _pick(d: dt.date) -> List[str]:
            cands = candidates_by_date[d]
            drow = dayrow_by_date.get(d)
            day_is_festivo = drow is not None and is_festivo(drow, cfg)
            def key(doc):
                festivo_pref = 1 if (day_is_festivo and doc in festivi_pool_norm) else 0
                return (festivo_pref, desired[doc] - _cnt[doc], -_cnt[doc], doc.lower())
            return sorted(cands, key=key, reverse=True)

        def _dfs(i: int) -> bool:
            if i == len(_c_dates_sorted):
                return True
            d = _c_dates_sorted[i]
            for doc in _pick(d):
                if _cnt[doc] >= desired[doc]:
                    continue
                if not _spacing_ok(doc, d):
                    continue
                assigned[d] = doc
                _cnt[doc] += 1
                _dates_by_doc[doc].append(d)
                if _dfs(i + 1):
                    return True
                _dates_by_doc[doc].pop()
                _cnt[doc] -= 1
                assigned.pop(d, None)
            return False

        assigned.clear()
        return _dfs(0)

    solved = False
    _spacing_tried = spacing_min
    for _sp in ([spacing_min] + list(range(spacing_min - 1, -1, -1))):
        solved = _run_dfs_attempt(_sp)
        if solved:
            if _sp < spacing_min:
                _c_relaxation_warnings.append(
                    f"C reperibilità: spacing rilassato a {_sp} giorni (originale {spacing_min})."
                )
            break

    if not solved:
        raise ValueError(
            "C_reperibilita: impossibile assegnare la reperibilità anche con vincoli rilassati. "
            f"Pool: {pool}, pool_size={n_docs}, giorni={total_days}. "
            "Verifica che il pool C non sia vuoto e che max_per_doctor sia sufficiente."
        )

    # Write back into assignment
    for d in c_dates:
        assignment[cslot_by_date[d].slot_id] = assigned.get(d)

    # Diagnostics (compute counts from assigned dict)
    _assigned_cnt: Dict[str, int] = {}
    for d in c_dates:
        doc = assigned.get(d)
        if doc:
            _assigned_cnt[doc] = _assigned_cnt.get(doc, 0) + 1

    status_str = "OK_RELAXED" if _c_relaxation_warnings else "OK_STRICT"
    diag: Dict = {
        "status": status_str,
        "pool_size": n_docs,
        "total_days": total_days,
        "spacing_min_days": spacing_min,
    }
    diag["counts"] = {k: v for k, v in sorted(_assigned_cnt.items(), key=lambda kv: (-kv[1], kv[0].lower())) if v}
    if _c_relaxation_warnings:
        diag["relaxation_warnings"] = _c_relaxation_warnings
    # Overlap stats
    overlap_total = 0
    overlap_weekdays = 0
    for d in c_dates:
        doc = assigned.get(d)
        if not doc:
            continue
        works = doctor_works_same_day(d, doc)
        if works:
            overlap_total += 1
            drow = dayrow_by_date.get(d)
            if drow is not None and not is_festivo(drow, cfg):
                overlap_weekdays += 1
    diag["overlap_days_total"] = overlap_total
    diag["overlap_days_weekdays"] = overlap_weekdays

    return assignment, {"C_reperibilita_diag": diag}


# -------------------------
# Load config / template
# -------------------------
def load_rules(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    if not isinstance(cfg, dict):
        raise ValueError("Rules YAML must be a mapping.")
    return cfg


def apply_pool_config(cfg_yaml: dict, pool_cfg: Optional[dict]) -> dict:
    """Sovrascrive pool/quote/flag nel cfg YAML con i valori del pool_config JSON.

    Se pool_cfg è None o vuoto ritorna una deep copy di cfg_yaml invariata.
    La funzione è idempotente: applicata più volte produce lo stesso risultato.
    """
    import copy as _copy

    cfg = _copy.deepcopy(cfg_yaml)
    if not pool_cfg or not pool_cfg.get("doctors"):
        return cfg

    doctors: dict = pool_cfg.get("doctors", {})
    col_settings: dict = pool_cfg.get("column_settings", {})
    rules = cfg.setdefault("rules", {})
    gc = cfg.setdefault("global_constraints", {})

    # 1. active=false → absolute_exclusions
    abs_excl: list = list(cfg.get("absolute_exclusions") or [])
    abs_excl_set = {norm_name(x) for x in abs_excl}
    for doc, dcfg in doctors.items():
        if not dcfg.get("active", True) and norm_name(doc) not in abs_excl_set:
            abs_excl.append(doc)
            abs_excl_set.add(norm_name(doc))
    cfg["absolute_exclusions"] = abs_excl

    # Medici attivi (normalizzati)
    active_docs = [doc for doc, dcfg in doctors.items() if dcfg.get("active", True)]
    active_set = {norm_name(d) for d in active_docs}

    # 2. Mappa colonna → pool key nel YAML (per sostituire i pool)
    _COL_RULE: dict[str, list[tuple[str, str]]] = {
        # (rule_key, pool_field)
        "C":  [("C_reperibilita", None)],         # C si gestisce via excluded, non pool
        "D":  [("D_F", "allowed")],
        "F":  [("D_F", "allowed")],
        "E":  [("E_G", "allowed")],
        "G":  [("E_G", "allowed")],
        "H":  [("H", "pool_mon_fri"), ("H", "distribution_pool")],
        "I":  [("I", "distribution_pool")],
        "J":  [("J", "pool_other")],
        "K":  [("K", "pool")],
        "L":  [("L", "pool_other")],
        "Q":  [("Q", "pool")],
        "R":  [("R", "pool")],
        "S":  [("S", "pool")],
        "T":  [("T", "pool")],
        "U":  [("U", "pool")],
        "V":  [("V", "pool")],
        "W":  [("W", "other_days_pool")],
        "Y":  [("Y", "other_pool")],
        "Z":  [("Z", "pool")],
        "AB": [("AB", "fallback_pool")],
    }

    for col, rule_targets in _COL_RULE.items():
        if col == "C":
            continue  # gestito separatamente al punto 5
        new_pool = [
            doc for doc, dcfg in doctors.items()
            if dcfg.get("active", True) and col in (dcfg.get("columns") or [])
        ]
        if not new_pool:
            continue  # safety: se il pool è vuoto, lascia il YAML invariato
        for rule_key, pool_field in rule_targets:
            rules.setdefault(rule_key, {})[pool_field] = new_pool

    # 3. Pool festivi diurni (D/E/H/I nei giorni festivi)
    fest_incl = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True) and dcfg.get("festivi_diurni", True)
    ]
    fest_excl = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True) and not dcfg.get("festivi_diurni", True)
    ]
    if fest_incl:
        rules.setdefault("Festivi", {})["pool"] = fest_incl
    if fest_excl:
        rules.setdefault("Festivi", {})["excluded"] = fest_excl

    # 4. Pool festivi notti (J nei giorni festivi)
    festivi_notti_excl = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True)
        and "J" in (dcfg.get("columns") or [])
        and not dcfg.get("festivi_notti", True)
    ]
    cfg["pool_festivi_notti_excluded"] = {norm_name(d) for d in festivi_notti_excl}

    # 5. Reperibilità C: sostituisce excluded list
    c_excl = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True) and dcfg.get("excluded_from_reperibilita", False)
    ]
    # Aggiungi anche i non-attivi (non assegnabili comunque ma meglio espliciti)
    for doc, dcfg in doctors.items():
        if not dcfg.get("active", True) and doc not in c_excl:
            c_excl.append(doc)
    rules.setdefault("C_reperibilita", {})["excluded"] = c_excl

    # 6. Weekend J excluded — non richiede J in columns (può entrare via monthly_quotas)
    j_weekend_excl = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True)
        and not (dcfg.get("column_overrides") or {}).get("J", {}).get("weekend_nights", True)
    ]
    rules.setdefault("J", {})["weekend_excluded_doctors"] = j_weekend_excl

    # 7. Quote J fixed → J.monthly_quotas (compatibile con solver esistente)
    j_mq: dict = dict(rules.get("J", {}).get("monthly_quotas") or {})
    quota_overrides: dict = {}  # (doc_norm, col) → {type, value}

    for doc, dcfg in doctors.items():
        dn = norm_name(doc)
        overrides: dict = dcfg.get("column_overrides") or {}
        for col, ov in overrides.items():
            if not isinstance(ov, dict):
                continue
            mq = ov.get("monthly_quota")
            qt = ov.get("quota_type", "fixed")
            if mq is None:
                continue
            if col == "J" and qt == "fixed":
                j_mq[doc] = mq  # usa chiave originale per compatibilità
            else:
                quota_overrides[(dn, col)] = {"type": qt, "value": int(mq)}

    rules.setdefault("J", {})["monthly_quotas"] = j_mq
    cfg["pool_quota_overrides"] = quota_overrides

    # 8. counts_as per colonna
    counts_as_map: dict[str, int] = {}
    for col, cs in col_settings.items():
        if isinstance(cs, dict) and "counts_as" in cs:
            counts_as_map[col] = int(cs["counts_as"])
    if counts_as_map:
        cfg["pool_counts_as"] = counts_as_map

    # 9. service_combinations
    combos = pool_cfg.get("service_combinations")
    if combos is not None:
        cfg["pool_service_combinations"] = combos

    # 10. critical_services
    critical = pool_cfg.get("critical_services")
    if critical is not None:
        cfg["pool_critical_services"] = critical

    # 11. Spacing J: scrive direttamente nei global_constraints già letti dal solver
    j_cs = col_settings.get("J", {})
    if isinstance(j_cs, dict):
        if "spacing_min_days" in j_cs:
            gc["night_spacing_days_min"] = int(j_cs["spacing_min_days"])
        if "spacing_preferred_days" in j_cs:
            gc["night_spacing_days_preferred"] = int(j_cs["spacing_preferred_days"])

    # 11b. service_combinations → mappa alle chiavi relief_valves esistenti
    combos = pool_cfg.get("service_combinations") or []
    relief = gc.setdefault("relief_valves", {})
    for combo in combos:
        cols = tuple(sorted(combo.get("columns") or []))
        mode = combo.get("mode", "fallback")
        if cols == ("K", "T"):
            if mode in ("fallback", "always"):
                relief["enable_kt_share"] = True
            elif mode == "preferred":
                relief["enable_kt_share"] = False  # preferred = soft, solver gestisce a bassa penalità
        # Q+R: già gestito da allow_blank_columns.R — nessuna modifica necessaria

    # 12. university_doctors — aggiorna da pool_config
    gc_uni: dict = dict(gc.get("university_doctors") or {})
    for doc, dcfg in doctors.items():
        uni = dcfg.get("university_doctor")
        if uni and isinstance(uni, dict):
            ratio = float(uni.get("ratio", gc.get("university_ratio", 0.6)))
            night_double = "J" in (dcfg.get("columns") or [])
            gc_uni[doc] = {
                "type": "university",
                "night_counts_double": night_double,
            }
            gc["university_ratio"] = ratio
        elif doc in gc_uni and uni is None:
            # Rimosso da pool_config → rimuovi anche da gc_uni
            del gc_uni[doc]
    gc["university_doctors"] = gc_uni

    # 13. Soft balance per colonne senza balance nel YAML
    # U (Contr.PM) non ha balance:true nel YAML — lo aggiungiamo quando pool_config
    # definisce il pool, così il solver bilancia automaticamente il carico.
    _u_pool = [
        doc for doc, dcfg in doctors.items()
        if dcfg.get("active", True) and "U" in (dcfg.get("columns") or [])
    ]
    if _u_pool:
        rules.setdefault("U", {}).setdefault("balance", True)
        rules["U"].setdefault("balance_weight", 200)

    return cfg


def _strip_festivi_unavailability(
    unav_map: Dict[str, Dict[dt.date, Set[str]]],
    tf_fixed: List[dict],
) -> None:
    """Remove unavailability entries that would block pre-assigned festivo shifts.

    If a doctor declares unavailability on a day/shift where they've been assigned
    via sorteggio (turni_festivi.yml), the sorteggio takes precedence and the
    conflicting unavailability entry is silently ignored.
    Modifies unav_map in place.
    """
    # Shifts blocked by each column assignment
    COL_TO_SHIFTS: Dict[str, Set[str]] = {
        "D": {"Mattina", "Diurno", "Tutto il giorno", "Any"},
        "H": {"Pomeriggio", "Diurno", "Tutto il giorno", "Any"},
        "J": {"Notte", "Tutto il giorno", "Any"},
    }
    for fa in tf_fixed:
        doc = norm_name(str(fa.get("doctor", "")).strip())
        col = str(fa.get("column", "")).strip().upper()
        date_str = str(fa.get("date", "")).strip()
        try:
            date = dt.date.fromisoformat(date_str)
        except Exception:
            continue
        blocked_shifts = COL_TO_SHIFTS.get(col, set())
        if not blocked_shifts:
            continue
        doc_unav = unav_map.get(doc)
        if doc_unav and date in doc_unav:
            doc_unav[date] -= blocked_shifts
            if not doc_unav[date]:
                del doc_unav[date]


# Mapping shift name → Excel column letter for fixed-assignment injection
_FESTIVI_SHIFT_TO_COL = {
    "mattina": "D",    # festivo → slot DE (columns D+E); regular → slot D
    "pomeriggio": "H", # festivo → slot HI (columns H+I); regular → slot H
    "notte": "J",      # always → slot J
}


def load_turni_festivi(base_dir: Optional[Path] = None) -> dict:
    """Load pre-assigned holiday shifts from data/turni_festivi.yml.

    Returns a dict:
      'festivi_extra'     : list of ISO date strings to merge into cfg['festivi_extra']
      'fixed_assignments' : list of {doctor, date, column} ready for the solver
    """
    if base_dir is None:
        base_dir = Path(__file__).resolve().parent
    path = base_dir / "data" / "turni_festivi.yml"
    if not path.exists():
        return {"festivi_extra": [], "fixed_assignments": []}

    try:
        with path.open("r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except Exception:
        return {"festivi_extra": [], "fixed_assignments": []}

    if not isinstance(data, dict):
        return {"festivi_extra": [], "fixed_assignments": []}

    festivi_extra = [str(x).strip() for x in (data.get("festivi_extra") or []) if x]

    fixed: List[dict] = []
    for entry in (data.get("entries") or []):
        if not isinstance(entry, dict):
            continue
        try:
            date_str = str(entry.get("date", "")).strip()
            doctor = str(entry.get("doctor", "")).strip()
            shift_raw = str(entry.get("shift", "")).strip().lower()
            # explicit 'column' field overrides shift-to-column mapping
            col = entry.get("column") or _FESTIVI_SHIFT_TO_COL.get(shift_raw)
            if not date_str or not doctor or not col:
                continue
            dt.date.fromisoformat(date_str)  # validate format
            fixed.append({
                "doctor": doctor,
                "date": date_str,
                "column": str(col).strip().upper(),
            })
        except Exception:
            continue

    return {"festivi_extra": festivi_extra, "fixed_assignments": fixed}


def create_month_template_xlsx(
    rules_yml: "Path | str",
    year: int,
    month: int,
    out_path: "Path | str",
    sheet_name: Optional[str] = None,
) -> Path:
    """Create a minimal Excel template for the given month.

    The generated template is compatible with this generator:
    - Column A (from row 2): dates
    - Optional headers in row 1
    - Creates the worksheet specified by `sheet_name` (or a default)
    - Writes headers for columns declared in the YAML `columns:` mapping
    - Ensures `keep_empty_columns:` exist (as blank headers) for layout

    Parameters
    ----------
    rules_yml: YAML rules file path
    year, month: target year/month
    out_path: output .xlsx path
    sheet_name: worksheet name to create

    Returns
    -------
    Path to the created template.
    """
    rules_path = Path(rules_yml)
    outp = Path(out_path)
    cfg = load_rules(rules_path)

    # Determine columns from YAML
    cols_map = cfg.get("columns") or {}
    if not isinstance(cols_map, dict):
        cols_map = {}
    keep_empty = cfg.get("keep_empty_columns") or []
    if not isinstance(keep_empty, list):
        keep_empty = []

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    if sheet_name:
        ws.title = str(sheet_name)
    else:
        ws.title = f"GUARDIE_{year}_{month:02d}"

    # Header cells (kept blank like the official model)
    _ = ws["A1"]
    _ = ws["B1"]

    # Column headers from YAML mapping (letters -> labels)
    for col_letter, label in cols_map.items():
        col_letter = str(col_letter).strip().upper()
        if not col_letter:
            continue
        ws[f"{col_letter}1"] = str(label) if label is not None else ""

    # Keep empty spacer columns
    for col_letter in keep_empty:
        col_letter = str(col_letter).strip().upper()
        if not col_letter:
            continue
        # Ensure the cell exists (leave blank)
        _ = ws[f"{col_letter}1"]

    # Fill dates
    last_day = calendar.monthrange(int(year), int(month))[1]
    r = 2
    for day in range(1, last_day + 1):
        d = dt.date(int(year), int(month), int(day))
        ws.cell(row=r, column=1).value = d
        ws.cell(row=r, column=1).number_format = "dd/mm/yyyy"
        # Italian day labels in the Excel view (logic uses DOW_MAP internally)
        ws.cell(row=r, column=2).value = ["lunedi","martedi","mercoledi","giovedi","venerdi","sabato","domenica"][d.weekday()]
        r += 1

    # Nice-to-have formatting
    try:
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 6
    except Exception:
        pass

    # Apply model styling (if Style_Template.xlsx is present)
    _apply_model_style_to_template(ws, cfg, int(year), int(month), int(last_day))

    outp.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outp)
    return outp

def load_template_days(xlsx_path: Path, sheet_name: Optional[str]=None) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet, List[DayRow]]:
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise KeyError(f"Worksheet '{sheet_name}' does not exist. Available: {available}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    # Find day rows: column A contains dates
    days: List[DayRow] = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (dt.datetime, dt.date)):
            d = v.date() if isinstance(v, dt.datetime) else v
            dow = DOW_MAP[d.weekday()]
            days.append(DayRow(date=d, dow=dow, row_idx=r))
    if not days:
        raise ValueError("No date rows found in column A (starting from row 2).")
    return wb, ws, days
def load_unavailability(unav_path: Optional[Path]) -> Dict[str, Dict[dt.date, Set[str]]]:
    """
    Returns: unav[doctor][date] = {'Mattina','Pomeriggio','Notte'} (oppure 'Any' per full-day)
    """
    unav: Dict[str, Dict[dt.date, Set[str]]] = defaultdict(lambda: defaultdict(set))
    if unav_path is None:
        return unav
    if not unav_path.exists():
        raise FileNotFoundError(unav_path)
    if unav_path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(unav_path)
    elif unav_path.suffix.lower() in [".csv", ".tsv"]:
        sep = "\t" if unav_path.suffix.lower() == ".tsv" else ","
        df = pd.read_csv(unav_path, sep=sep)
    else:
        raise ValueError("Unavailability file must be .xlsx/.xls/.csv/.tsv")
    # Flexible column names
    cols = {c.lower().strip(): c for c in df.columns}
    def pick(*names):
        for n in names:
            if n in cols:
                return cols[n]
        return None
    c_med = pick("medico", "doctor", "name")
    c_dat = pick("data", "date", "giorno")
    c_fas = pick("fascia", "shift", "turno")
    if not (c_med and c_dat and c_fas):
        raise ValueError("Unavailability file must contain columns: Medico, Data, Fascia")
    for _, row in df.iterrows():
        med = row.get(c_med)
        dat = row.get(c_dat)
        fas = row.get(c_fas)
        if pd.isna(med) or pd.isna(dat) or pd.isna(fas):
            continue
        doctor = norm_name(med)
        date = parse_date(dat)
        for shift in shifts_from_fascia(fas):
            unav[doctor][date].add(shift)
    return unav
def collect_doctors(cfg: dict) -> List[str]:
    """
    Union of all pools/allowed in YAML, minus absolute_exclusions.
    Keeps special placeholder 'Recupero' if present.
    """
    doctors: Set[str] = set()
    # From unavailability section too
    for d in (cfg.get("unavailability") or {}).keys():
        doctors.add(norm_name(d))
    rules = cfg.get("rules", {})
    if isinstance(rules, dict):
        for _, rule in rules.items():
            if not isinstance(rule, dict):
                continue
            for k in ["allowed", "pool", "pool_other", "pool_mon_fri", "other_pool", "fallback_pool", "distribution_pool"]:
                if k in rule and isinstance(rule[k], list):
                    doctors |= {norm_name(x) for x in rule[k]}
            for k in ["fixed", "tuesday_fixed", "prefer"]:
                if k in rule and rule[k]:
                    doctors.add(norm_name(rule[k]))
    # Remove absolute exclusions
    abs_excl = {norm_name(x) for x in (cfg.get("absolute_exclusions") or [])}
    doctors = {d for d in doctors if d not in abs_excl}
    # Stable sorting: keep 'Recupero' last-ish
    doctors_list = sorted(doctors, key=lambda s: (s == "Recupero", s.lower()))
    return doctors_list
# -------------------------
# Build slots from rules
# -------------------------
def is_festivo(day: DayRow, cfg: dict) -> bool:
    extra = set()
    for x in cfg.get("festivi_extra", []) or []:
        try:
            extra.add(parse_date(x))
        except Exception:
            pass
    # Treat Sunday and Italian national holidays as "festivo".
    # Extra holidays can be provided via cfg.festivi_extra.
    hol = italy_public_holidays(int(day.date.year))
    return day.dow == "Sun" or day.date in hol or day.date in extra
def apply_unavailability(allowed: List[str], day: DayRow, shift: str, unav: Dict[str, Dict[dt.date, Set[str]]]) -> List[str]:
    out = []
    for doc in allowed:
        d_unav = unav.get(doc, {}).get(day.date, set())
        # If any shift marked 'Any' treat as full-day
        if "Any" in d_unav:
            continue
        if shift == "Any":
            if d_unav:
                continue
        elif shift in d_unav:
            continue
        out.append(doc)
    return out
def slots_for_month(cfg: dict, days: List[DayRow], unav: Dict[str, Dict[dt.date, Set[str]]], fixed_assignments: Optional[List[dict]] = None, v_double_overrides: Optional[List[str]] = None, j_blank_week_overrides: Optional[Dict[str, Optional[str]]] = None) -> List[Slot]:
    """
    Converts YAML column rules into per-day slots.
    Handles exception days (festivi) by merging D+E and H+I, and merging E+G always.
    """
    rules = cfg.get("rules", {})
    if not isinstance(rules, dict):
        raise ValueError("cfg.rules must be a mapping.")
    doctors_all = collect_doctors(cfg)
    doctors_set = set(doctors_all)
    # Relief valves (optional): allow specific columns to be left blank with penalties (used only if needed).
    gc = cfg.get("global_constraints") or {}
    relief = gc.get("relief_valves") or {}
    blank_penalties: Dict[str, int] = {}
    if isinstance(relief.get("allow_blank_columns"), dict):
        for _k, _v in (relief.get("allow_blank_columns") or {}).items():
            try:
                blank_penalties[str(_k).strip().upper()] = int(_v)
            except Exception:
                pass
    def req_and_blank(col_letter: str) -> Tuple[bool, int]:
        col_letter = str(col_letter).strip().upper()
        if col_letter in blank_penalties:
            return False, blank_penalties[col_letter]
        return True, 0
    # YAML also has inline date-only unavailability (full-day)
    for doc, dates in (cfg.get("unavailability") or {}).items():
        for ds in dates or []:
            try:
                unav[norm_name(doc)][parse_date(ds)].add("Any")
            except Exception:
                pass
    # FASE 0 — PRE-PROCESSA i fixed_assignments PRIMA della costruzione degli slot.
    # I fixed_assignment in J rendono quel medico di fatto "non disponibile" per D/F
    # lo stesso giorno (night_off same_day). Li trattiamo come indisponibilità temporanee
    # per la costruzione dell'allowed D/F.
    forced_j_by_date: Dict[dt.date, Set[str]] = {}
    forced_de_by_date: Dict[dt.date, str] = {}  # sorteggio DE (Mattina) festivi
    forced_hi_by_date: Dict[dt.date, str] = {}  # sorteggio HI (Pomeriggio) festivi
    for fa in (fixed_assignments or []):
        fa_col = str(fa.get("column","")).strip().upper()
        try:
            fa_date = dt.date.fromisoformat(str(fa.get("date","")).strip())
            fa_doc = norm_name(str(fa.get("doctor","")).strip())
            if fa_col == "J":
                forced_j_by_date.setdefault(fa_date, set()).add(fa_doc)
            elif fa_col == "D" and fa_doc in doctors_set:
                # Il medico sorteggiato potrebbe non essere nel pool Festivi (es. Grimaldi, Calabrò)
                # → lo registriamo qui per usarlo come unico allowed nel slot DE,
                #   evitando il conflitto sum(allowed)==0 == 1 nel CP-SAT.
                forced_de_by_date[fa_date] = fa_doc
            elif fa_col == "H" and fa_doc in doctors_set:
                forced_hi_by_date[fa_date] = fa_doc
        except Exception:
            pass

    # PRE-PROCESSA j_blank_week_overrides: per settimana, qual è il giorno in cui J è vuota
    # Formato chiave: "YYYY-WNN" (es. "2026-W16"), valore: data ISO o None (= nessun vuoto)
    _j_week_ov: Dict[tuple, Optional[dt.date]] = {}
    for _wk_str, _bd_str in (j_blank_week_overrides or {}).items():
        try:
            _parts = str(_wk_str).split("-W")
            _iso_key = (int(_parts[0]), int(_parts[1]))
            _blank_d = dt.date.fromisoformat(str(_bd_str).strip()) if _bd_str else None
            _j_week_ov[_iso_key] = _blank_d
        except Exception:
            pass

    # PRE-PROCESSA v_double_overrides: date esatte in cui V è in doppio (al posto del venerdì)
    # Se quella settimana ha un override, il venerdì di quella settimana diventa turno singolo.
    # Sentinel "NODOUBLE:{year}:{week}" → quella settimana non ha nessun turno doppio.
    _v_override_dates: Set[dt.date] = set()
    _v_no_double_weeks: Set[tuple] = set()
    if v_double_overrides:
        for _ds in v_double_overrides:
            _ds = str(_ds).strip()
            if _ds.startswith("NODOUBLE:"):
                try:
                    _, _yr_wk = _ds.split(":", 1)
                    _yr_s, _wk_s = _yr_wk.split(":")
                    _v_no_double_weeks.add((int(_yr_s), int(_wk_s)))
                except Exception:
                    pass
            else:
                try:
                    _v_override_dates.add(dt.date.fromisoformat(_ds))
                except Exception:
                    pass
    # ISO week keys delle settimane che hanno un override (doppio spostato O nessun doppio)
    _v_override_weeks: Set[tuple] = {d.isocalendar()[:2] for d in _v_override_dates} | _v_no_double_weeks

    slots: List[Slot] = []
    def mk_allowed(pool: List[str]) -> List[str]:
        # keep only known doctors + keep 'Recupero' if used
        out = [norm_name(x) for x in pool if norm_name(x) in doctors_set]
        return out
    for day in days:
        festivo = is_festivo(day, cfg)
        # Optional: on Saturdays, assign the SAME doctor to K and T (single combined slot K+T)
        gc = cfg.get("global_constraints", {}) or {}
        merge_KT_sat = bool(gc.get("saturday_K_equals_T", False)) and (day.dow == "Sat") and (not festivo)             and ("K" in rules) and ("T" in rules) and dayspec_contains(day.dow, (rules.get("T") or {}).get("days"))
        # ---- C: Reperibilità (daily)
        if "C_reperibilita" in rules:
            r = rules["C_reperibilita"]
            excluded = {norm_name(x) for x in (r.get("excluded") or [])}
            pool = [d for d in doctors_all if d not in excluded and d != "Recupero"]  # usually a real doctor
            pool = apply_unavailability(pool, day, "Any", unav)
            slots.append(Slot(day, f"{day.date}-C", ["C"], pool, required=True, shift="Any", rule_tag="C_reperibilita"))
        # ---- Morning D/F and E/G and Afternoon H/I depend on festivo
        if festivo:
            rFest = rules.get("Festivi", {}) if isinstance(rules.get("Festivi", {}), dict) else {}
            fest_excl = {norm_name(x) for x in (rFest.get("excluded") or [])}

            # DE unified (D+E) – required
            # Se c'è un medico sorteggiato (anche escluso dal pool Festivi, es. Grimaldi/Calabrò),
            # lo mettiamo come UNICO allowed per evitare sum(allowed)==0==1 nel CP-SAT.
            _forced_de = forced_de_by_date.get(day.date)
            if _forced_de:
                allowed_de = [_forced_de]
            else:
                fest_pool_m = rFest.get("pool_mattina") or rFest.get("pool") or []
                allowed_de = mk_allowed(fest_pool_m)
                if not allowed_de:
                    allowed_de = [d for d in doctors_all if d != "Recupero" and d not in fest_excl]
                else:
                    allowed_de = [d for d in allowed_de if d not in fest_excl]
                allowed_de = apply_unavailability(allowed_de, day, "Mattina", unav)
            slots.append(Slot(day, f"{day.date}-DE", ["D","E"], allowed_de, required=True, shift="Mattina", rule_tag="Festivo_DE"))

            # HI unified (H+I) – required
            _forced_hi = forced_hi_by_date.get(day.date)
            if _forced_hi:
                allowed_hi = [_forced_hi]
            else:
                fest_pool_p = rFest.get("pool_pomeriggio") or rFest.get("pool") or []
                allowed_hi = mk_allowed(fest_pool_p)
                if not allowed_hi:
                    allowed_hi = [d for d in doctors_all if d != "Recupero" and d not in fest_excl]
                else:
                    allowed_hi = [d for d in allowed_hi if d not in fest_excl]
                allowed_hi = apply_unavailability(allowed_hi, day, "Pomeriggio", unav)
            slots.append(Slot(day, f"{day.date}-HI", ["H","I"], allowed_hi, required=True, shift="Pomeriggio", rule_tag="Festivo_HI"))
        else:
            # D / F (Mon-Sat)
            if "D_F" in rules and dayspec_contains(day.dow, rules["D_F"].get("days")):
                r = rules["D_F"]
                pair_docs = mk_allowed(r.get("allowed") or [])
                pair_avail = apply_unavailability(pair_docs, day, "Mattina", unav)
                # Rimuovi i medici forzati in J quel giorno (night_off same_day li esclude da D/F)
                forced_j_today = forced_j_by_date.get(day.date, set())
                pair_avail = [d for d in pair_avail if d not in forced_j_today]

                # Fallback source = H.pool_mon_fri (as requested)
                h_rule = rules.get("H", {}) if isinstance(rules.get("H", {}), dict) else {}
                h_pool = mk_allowed(h_rule.get("pool_mon_fri") or [])
                h_avail = apply_unavailability(h_pool, day, "Mattina", unav)

                # Ultimate fallback: any doctor available that morning (after unavailability filter)
                any_pool = apply_unavailability(sorted(doctors_set), day, "Mattina", unav)

                # Build a robust shared domain (never empty for required slots)
                allowed_base = sorted({*pair_avail, *h_avail, *any_pool})
                if not allowed_base:
                    allowed_base = sorted(doctors_set)

                # Se solo uno del pair disponibile → share obbligatorio (lui fa D e F)
                # Se nessuno del pair → pool completo (H pool + qualsiasi medico libero al mattino)
                # Se entrambi → solo il pair, no share
                if len(pair_avail) == 1:
                    allowed_df = pair_avail
                    prefer_share = True
                elif len(pair_avail) == 0:
                    # Fallback a cascata:
                    # 1. Pool H (preferito)
                    # 2. Qualsiasi medico disponibile al mattino (allowed_base = H + any_pool)
                    # Questo evita che un solo medico (es. Migliorato) sia l'unico per D/F
                    # bloccandosi poi per H pomeriggio.
                    allowed_df = allowed_base if allowed_base else sorted(doctors_set)
                    prefer_share = True
                else:
                    allowed_df = pair_avail
                    prefer_share = False

                slots.append(Slot(day, f"{day.date}-D", ["D"], allowed_df, required=True, shift="Mattina", rule_tag="D_F.D", force_same_doctor=prefer_share))
                slots.append(Slot(day, f"{day.date}-F", ["F"], allowed_df, required=True, shift="Mattina", rule_tag="D_F.F", force_same_doctor=prefer_share))
# EG paired (Mon-Sat)
            if "E_G" in rules and dayspec_contains(day.dow, rules["E_G"].get("days")):
                r = rules["E_G"]
                allowed = mk_allowed(r.get("allowed") or [])
                allowed = apply_unavailability(allowed, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-EG", ["E","G"], allowed, required=True, shift="Mattina", rule_tag="E_G"))
            # H (Mon-Sat) + I (Mon-Sat)
            # MODIFICA 1: Grimaldi e Calabrò NON devono MAI comparire in H
            if "H" in rules:
                rH = rules["H"]
                # Saturdays must be covered; for Mon-Fri, use pool_mon_fri only
                if day.dow == "Sat" or dayspec_contains(day.dow, "Mon-Fri"):
                    allowed = mk_allowed(rH.get("pool_mon_fri") or [])
                    # Grimaldi e Calabrò esclusi esplicitamente da H
                    _h_excl = {norm_name("Grimaldi"), norm_name("Calabrò")}
                    allowed = [d for d in allowed if norm_name(d) not in _h_excl]
                    allowed = apply_unavailability(allowed, day, "Pomeriggio", unav)
                    slots.append(Slot(day, f"{day.date}-H", ["H"], allowed, required=True, shift="Pomeriggio", rule_tag="H"))
            if "I" in rules and dayspec_contains(day.dow, "Mon-Sat"):
                rI = rules["I"]
                allowed = mk_allowed(rI.get("distribution_pool") or [])
                allowed = apply_unavailability(allowed, day, "Pomeriggio", unav)
                # In practice I is an afternoon activity; required Mon-Sat
                slots.append(Slot(day, f"{day.date}-I", ["I"], allowed, required=True, shift="Pomeriggio", rule_tag="I"))
        # ---- Night J (daily except Thu if configured; MODIFICA 7: override data specifica)
        if "J" in rules:
            rJ = rules["J"]
            # Data di override (es. "2026-03-04" per mercoledì 4 marzo al posto di giovedì 5)
            _week_key_j = day.date.isocalendar()[:2]
            if _week_key_j in _j_week_ov:
                # Override UI: questa settimana ha un'impostazione specifica
                _ui_blank_d = _j_week_ov[_week_key_j]
                _skip_j = (_ui_blank_d is not None and day.date == _ui_blank_d)
            else:
                # Comportamento default: giovedì vuoto + eventuale override YAML
                _j_override_raw = (cfg.get("global_constraints") or {}).get("j_blank_override_date")
                _j_override_date = None
                if _j_override_raw:
                    try:
                        _j_override_date = dt.date.fromisoformat(str(_j_override_raw))
                    except Exception:
                        pass
                _is_normal_thu_blank = rJ.get("thursday_blank", False) and day.dow == "Thu"
                _is_override_blank = (_j_override_date is not None and day.date == _j_override_date)
                _thu_suppressed_by_override = False
                if _j_override_date is not None and day.dow == "Thu":
                    import datetime as _dt3
                    _thu_week_mon = day.date - _dt3.timedelta(days=day.date.weekday())
                    _ov_week_mon = _j_override_date - _dt3.timedelta(days=_j_override_date.weekday())
                    if _thu_week_mon == _ov_week_mon:
                        _thu_suppressed_by_override = True
                _skip_j = (_is_normal_thu_blank and not _thu_suppressed_by_override) or _is_override_blank
            if not _skip_j:
                # Se esiste un fixed_assignment in J per questo giorno,
                # lo slot usa SOLO quel medico — l'indisponibilità viene ignorata
                # perché l'admin ha deciso esplicitamente.
                forced_j_today = forced_j_by_date.get(day.date, set())
                if forced_j_today:
                    allowed = [d for d in forced_j_today if d in doctors_set]
                else:
                    allowed = mk_allowed(rJ.get("pool_other") or [])
                    # add quota doctors even if not in pool_other
                    for special in (rJ.get("monthly_quotas") or {}).keys():
                        if special in doctors_set and special not in allowed:
                            allowed.append(special)
                    allowed = [d for d in allowed if d != "Recupero"]
                    # Esclusioni permanenti da J (mai in notte, qualunque giorno)
                    j_never = {norm_name(d) for d in (rJ.get("never_in_J") or ["De Gregorio", "Manganaro"])}
                    allowed = [d for d in allowed if norm_name(d) not in j_never]
                    # Weekend exclusions (e.g., Calabrò not allowed on Sat/Sun nights)
                    wex = [norm_name(x) for x in (rJ.get('weekend_excluded_doctors') or [])]
                    if day.dow in ['Sat','Sun'] and wex:
                        allowed = [d for d in allowed if norm_name(d) not in set(wex)]
                    # festivi_notti filter: rimuove medici esclusi da J nei giorni festivi
                    if is_festivo(day, cfg):
                        _fest_notti_excl = cfg.get("pool_festivi_notti_excluded") or set()
                        if _fest_notti_excl:
                            allowed = [d for d in allowed if norm_name(d) not in _fest_notti_excl]
                    allowed = apply_unavailability(allowed, day, "Notte", unav)
                slots.append(Slot(day, f"{day.date}-J", ["J"], allowed, required=True, shift="Notte", rule_tag="J"))
        # ---- K Letto (daily, but blank on Sundays/festivi)
        # If merge_KT_sat is enabled, we create a single combined slot K+T on Saturday.
        if merge_KT_sat:
            rK = rules.get("K", {}) if isinstance(rules.get("K", {}), dict) else {}
            rT = rules.get("T", {}) if isinstance(rules.get("T", {}), dict) else {}
            allowed_k = mk_allowed(rK.get("pool") or [])
            allowed_t = mk_allowed(rT.get("pool") or [])
            allowed_k = apply_unavailability(allowed_k, day, "Mattina", unav)
            allowed_t = apply_unavailability(allowed_t, day, "Mattina", unav)
            inter = [d for d in allowed_k if d in set(allowed_t)]
            allowed = inter if inter else sorted({*allowed_k, *allowed_t})
            slots.append(Slot(day, f"{day.date}-KT", ["K","T"], allowed, required=True, shift="Mattina", rule_tag="K_T_SAT"))
        elif "K" in rules and not festivo:
            rK = rules["K"]
            allowed = mk_allowed(rK.get("pool") or [])
            allowed = apply_unavailability(allowed, day, "Mattina", unav)
            slots.append(Slot(day, f"{day.date}-K", ["K"], allowed, required=True, shift="Mattina", rule_tag="K"))
        # ---- L Padiglioni (Mon-Wed)
        if "L" in rules and not festivo:
            rL = rules["L"]
            if dayspec_contains(day.dow, rL.get("days")):
                pool = mk_allowed(rL.get("pool_other") or [])
                # allow Recupero as placeholder
                if "Recupero" in doctors_set and "Recupero" not in pool:
                    pool.append("Recupero")
                pool = apply_unavailability(pool, day, "Mattina", unav)
                # L usa sempre il relief valve (20K) — priorità inferiore a H (5M obbligatorio)
                req_relief, bp = req_and_blank("L")
                slots.append(Slot(day, f"{day.date}-L", ["L"], pool, required=req_relief, blank_penalty=bp, shift="Mattina", rule_tag="L"))
        # ---- Q Eco base (Mon-Sat)
        if "Q" in rules and not festivo:
            r = rules["Q"]
            if dayspec_contains(day.dow, r.get("days")):
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-Q", ["Q"], pool, required=True, shift="Mattina", rule_tag="Q"))
        # ---- R (Mon-Fri)
        if "R" in rules and not festivo:
            r = rules["R"]
            if dayspec_contains(day.dow, r.get("days")):
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                # R usa sempre il relief valve (40K) — priorità inferiore agli slot obbligatori
                req_relief, bp = req_and_blank("R")
                slots.append(Slot(day, f"{day.date}-R", ["R"], pool, required=req_relief, blank_penalty=bp, shift="Mattina", rule_tag="R"))
        # ---- S (Wed, optional if can be absorbed in R)
        if "S" in rules and not festivo:
            r = rules["S"]
            if dayspec_contains(day.dow, r.get("days") or r.get("day")):
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                required = not bool(r.get("if_not_dedicated_put_in_R", False))
                slots.append(Slot(day, f"{day.date}-S", ["S"], pool, required=required, shift="Mattina", rule_tag="S"))
        # ---- T Interni (Mon-Sat)
        if "T" in rules and not merge_KT_sat and not festivo:
            r = rules["T"]
            if dayspec_contains(day.dow, r.get("days")):
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-T", ["T"], pool, required=True, shift="Mattina", rule_tag="T"))
        # ---- U Contr.PM (Mon-Tue)
        if "U" in rules and not festivo:
            r = rules["U"]
            if dayspec_contains(day.dow, r.get("days")):
                # Lunedì Contr.PM è pomeridiano; martedì è mattutino
                u_shift = "Pomeriggio" if day.dow == "Mon" else "Mattina"
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, u_shift, unav)
                # MODIFICA 3: se lunedì e V ha solo Allegra disponibile (o è assegnato ad Allegra),
                # U deve essere SOLO Crea o Dattilo.
                if day.dow == "Mon" and r.get("v_allegra_monday_constraint", False):
                    rV = rules.get("V", {})
                    v_pool_avail = apply_unavailability(mk_allowed(rV.get("pool") or []), day, "Pomeriggio", unav)
                    _allegra = norm_name("Allegra")
                    _crea = norm_name("Crea")
                    _dattilo = norm_name("Dattilo")
                    # Se Allegra è l'unico disponibile in V (o i soli disponibili sono Allegra),
                    # oppure Crea e Dattilo non sono nel pool V → forza U = {Crea, Dattilo}
                    only_allegra_in_v = all(norm_name(d) == _allegra for d in v_pool_avail) if v_pool_avail else False
                    if only_allegra_in_v:
                        restricted = [d for d in pool if norm_name(d) in {_crea, _dattilo}]
                        if restricted:  # applica solo se almeno uno è disponibile
                            pool = restricted
                slots.append(Slot(day, f"{day.date}-U", ["U"], pool, required=True, shift=u_shift, rule_tag="U"))
        # ---- V Sala PM (Mon,Wed,Fri)
        # Default: venerdì = turno doppio (Crea + Dattilo|Allegra), lun/mer = singolo.
        # Override admin: se un lunedì o mercoledì è in _v_override_dates, quella settimana
        # il doppio è su quel giorno e il venerdì diventa singolo.
        if "V" in rules and not festivo:
            r = rules["V"]
            if dayspec_contains(day.dow, r.get("days")):
                # Lunedì la Sala PM è di pomeriggio; mercoledì e venerdì è di mattina
                v_shift = "Pomeriggio" if day.dow == "Mon" else "Mattina"
                pool_base = mk_allowed(r.get("pool") or [])
                pool_base = apply_unavailability(pool_base, day, v_shift, unav)
                # Determina se questo giorno è il "turno doppio" della settimana
                _week_key = day.date.isocalendar()[:2]
                _is_override_double = day.date in _v_override_dates
                _is_default_double = (day.dow == "Fri") and (_week_key not in _v_override_weeks)
                _is_double_day = _is_override_double or _is_default_double
                if _is_double_day:
                    crea = norm_name(r.get("friday_required_doctor") or "Crea")
                    pool_crea = [crea] if crea in pool_base else []
                    other_allowed = {norm_name("Dattilo"), norm_name("Allegra")}
                    pool_other = [d for d in pool_base if norm_name(d) in other_allowed and norm_name(d) != crea]
                    # Turno doppio: solo se entrambi i pool sono non vuoti
                    if (not pool_crea) or (not pool_other):
                        pool_crea = []
                        pool_other = []
                    slots.append(Slot(day, f"{day.date}-V1", ["V"], pool_crea, required=True, shift=v_shift, rule_tag="V"))
                    slots.append(Slot(day, f"{day.date}-V2", ["V"], pool_other, required=True, shift=v_shift, rule_tag="V"))
                else:
                    slots.append(Slot(day, f"{day.date}-V", ["V"], pool_base, required=True, shift=v_shift, rule_tag="V"))
        # ---- Z Vascolare (Wed,Fri)
        if "Z" in rules and not festivo:
            r = rules["Z"]
            if dayspec_contains(day.dow, r.get("days")):
                pool = mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                # Z usa il relief valve (30K) — si sacrifica prima di R (40K)
                req_relief_z, bp_z = req_and_blank("Z")
                slots.append(Slot(day, f"{day.date}-Z", ["Z"], pool, required=req_relief_z, blank_penalty=bp_z, shift="Mattina", rule_tag="Z"))
        # ---- W Ergometria/CPET (Mon-Fri; Tue fixed) — escluso nei festivi
        if "W" in rules and not festivo:
            r = rules["W"]
            if day.dow in ["Mon","Tue","Wed","Thu","Fri"]:
                if day.dow == "Tue" and r.get("tuesday_fixed"):
                    fixed = norm_name(r["tuesday_fixed"])
                    pool = [fixed] if fixed in doctors_set else []
                else:
                    pool = mk_allowed(r.get("other_days_pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                # W è sempre richiesto (lun-ven); se pool vuoto per indisponibilità
                # lo slot diventa opzionale con blank_penalty alta → cella gialla
                if pool:
                    required = True
                    blank_pen = 0
                else:
                    required = False
                    blank_pen = 5000
                slots.append(Slot(day, f"{day.date}-W", ["W"], pool, required=required,
                                  blank_penalty=blank_pen, shift="Mattina", rule_tag="W"))
        # ---- Y Amb specialistici (Mon only)
        # Requirement:
        #  - Every Monday: 1 doctor among other_pool
        #  - PLUS: on exactly 2 Mondays: also 'Recupero' (appended in the same cell)
        if "Y" in rules and not festivo:
            r = rules["Y"]
            if dayspec_contains(day.dow, r.get("days") or r.get("day")):
                # Main doctor (always required)
                pool_main = [d for d in mk_allowed(r.get("other_pool") or []) if d != "Recupero"]
                pool_main = apply_unavailability(pool_main, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-Y", ["Y"], pool_main, required=True, shift="Mattina", rule_tag="Y_MAIN"))
        # ---- AB Holter/Brugada/FA (Thu)
        if "AB" in rules and not festivo:
            r = rules["AB"]
            if r.get("weekly", False) and dayspec_contains(day.dow, r.get("fixed_day")):
                # MODIFICA 5: nessuna preferenza per Crea il giovedì — pool bilanciato
                prefer = norm_name(r.get("prefer") or "")
                pool = []
                if prefer and prefer in doctors_set:
                    pool.append(prefer)
                pool += mk_allowed(r.get("fallback_pool") or [])
                # unique list preserving order
                seen=set(); pool=[x for x in pool if not (x in seen or seen.add(x))]
                pool = apply_unavailability(pool, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-AB", ["AB"], pool, required=True, shift="Mattina", rule_tag="AB"))
            # Slot aggiuntivo al Sabato (2 al mese) SOLO con CREA (quota gestita nel solver)
            sat_n = int(r.get("saturday_per_month", 0) or 0)
            if sat_n > 0 and day.dow == "Sat":
                doc_sat = norm_name(r.get("saturday_only_doctor") or "Crea")
                pool_sat = [doc_sat] if doc_sat in doctors_set else []
                pool_sat = apply_unavailability(pool_sat, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-AB_SAT", ["AB"], pool_sat, required=False, shift="Mattina", rule_tag="AB_SAT"))
        # ---- AC Scintigrafia (Tue/Wed fixed)
        if "AC" in rules and not festivo:
            r = rules["AC"]
            if dayspec_contains(day.dow, r.get("days")):
                fixed = norm_name(r.get("fixed") or "")
                pool = [fixed] if fixed in doctors_set else mk_allowed(r.get("pool") or [])
                pool = apply_unavailability(pool, day, "Mattina", unav)
                slots.append(Slot(day, f"{day.date}-AC", ["AC"], pool, required=True, shift="Mattina", rule_tag="AC"))
    # ── Espansione pool per colonne indispensabili (critical_services) ──────────
    # Per ogni colonna marcata come "indispensabile" in pool_critical_services:
    # se il pool primario è vuoto o molto ridotto, si espande a qualsiasi medico
    # disponibile per quel turno (con penalità nel solver per scoraggiarne l'uso).
    _abs_excl_norm = {norm_name(x) for x in (cfg.get("absolute_exclusions") or [])}
    _emerg_any_pool = [d for d in doctors_all
                       if norm_name(d) not in _abs_excl_norm and norm_name(d) != "recupero"]
    _critical_svc = cfg.get("pool_critical_services") or {}

    for s in slots:
        if not s.required:
            continue
        _match_col = next((c for c in (s.columns or []) if c in _critical_svc), None)
        if not _match_col:
            continue
        _spec = _critical_svc.get(_match_col, {})
        _fb = _spec.get("fallback", "")
        if _fb == "any":
            _fb_base = _emerg_any_pool
        elif isinstance(_fb, list) and _fb:
            _fb_base = [norm_name(d) for d in _fb if norm_name(d) in doctors_set]
        else:
            continue
        _fb_avail = apply_unavailability(_fb_base, s.day, s.shift, unav)
        _primary_set = set(s.allowed)
        _new_emerg = [d for d in _fb_avail if d not in _primary_set]
        if _new_emerg:
            s.allowed = s.allowed + _new_emerg
            s.emergency_doctors = _new_emerg

    # ── Validate domains: slot required con pool ancora vuoto ────────────────
    # Fallback generico: se un required slot ha il pool completamente esaurito
    # (non è in critical_services), prova l'emergency pool globale prima di
    # renderlo opzionale.
    for s in slots:
        if not s.allowed:
            if s.required:
                _emerg = apply_unavailability(_emerg_any_pool, s.day, s.shift, unav)
                if _emerg:
                    s.allowed = _emerg
                    s.emergency_doctors = _emerg
                    continue
            s.empty_domain = True
            if s.required:
                s.required = False
                if getattr(s, "blank_penalty", 0) == 0:
                    s.blank_penalty = 1
    return slots
# -------------------------
# Solver (OR-Tools CP-SAT)
# -------------------------
def _max_bipartite_matching(slots_day: List[Slot]) -> Tuple[int, Dict[str, str]]:
    """
    Simple DFS-based bipartite matching: Slot -> Doctor.
    Returns:
      matched_count, slot_to_doc (by slot_id)
    """
    # Ensure deterministic iteration
    slots_day = list(slots_day)
    doc_to_slot: Dict[str, Slot] = {}
    def try_assign(slot: Slot, seen: Set[str]) -> bool:
        for d in dict.fromkeys(slot.allowed):
            if d in seen:
                continue
            seen.add(d)
            if d not in doc_to_slot or try_assign(doc_to_slot[d], seen):
                doc_to_slot[d] = slot
                return True
        return False
    for slot in sorted(slots_day, key=lambda s: (len(s.allowed), s.slot_id)):
        try_assign(slot, set())
    matched_slot_ids = set(s.slot_id for s in doc_to_slot.values())
    slot_to_doc: Dict[str, str] = {}
    for d, s in doc_to_slot.items():
        slot_to_doc[s.slot_id] = d
    return len(matched_slot_ids), slot_to_doc
def diagnose_day_level(days: List[DayRow], slots: List[Slot]) -> List[Dict]:
    """
    Day-level feasibility diagnostics (ignores cross-day constraints such as night spacing).
    Useful to pinpoint single-day bottlenecks created by unavailability.
    """
    slots_by_day: Dict[dt.date, List[Slot]] = defaultdict(list)
    for s in slots:
        slots_by_day[s.day.date].append(s)
    report: List[Dict] = []
    for day in days:
        day_slots = slots_by_day.get(day.date, [])
        uniq_slots = [s for s in day_slots if not _slot_is_exempt_daily(s)]
        # consider only required slots (and also "penalized optional" slots) as "should be filled"
        must_fill = [s for s in day_slots if s.required or (getattr(s, "blank_penalty", 0) and int(getattr(s, "blank_penalty", 0)) > 0)]
        if not must_fill:
            continue
        matched, slot_to_doc = _max_bipartite_matching(must_fill)
        ok = (matched == len(must_fill))
        union_docs = sorted({d for s in must_fill for d in s.allowed})
        tight = sorted([(s.slot_id, s.columns, len(s.allowed)) for s in must_fill], key=lambda x: x[2])[:6]
        if not ok:
            # identify some unmatched slots
            matched_ids = set(slot_to_doc.keys())
            unmatched = []
            for s in sorted(must_fill, key=lambda s: (len(s.allowed), s.slot_id)):
                if s.slot_id not in matched_ids:
                    unmatched.append({"slot_id": s.slot_id, "columns": s.columns, "allowed_n": len(s.allowed), "allowed": s.allowed[:15]})
            report.append({
                "date": day.date.isoformat(),
                "dow": day.dow,
                "required_slots": len(must_fill),
                "union_doctors": len(union_docs),
                "tightest_slots": tight,
                "unmatched_slots": unmatched[:6],
            })
    return report
def build_relief_log(days: List[DayRow], slots: List[Slot], assignment: Dict[str, Optional[str]]) -> Dict:
    """
    Summarize where relief valves were used (K=T same doctor, blanks on penalized optional columns).
    """
    slots_by_day: Dict[dt.date, List[Slot]] = defaultdict(list)
    for s in slots:
        slots_by_day[s.day.date].append(s)
    kt_share_days: List[str] = []
    df_share_days: List[str] = []
    df_forced_same_days: List[str] = []
    blank_cols: Dict[str, List[str]] = defaultdict(list)
    for day in days:
        day_slots = slots_by_day.get(day.date, [])
        # blanks
        for s in day_slots:
            if (not s.required) and int(getattr(s, "blank_penalty", 0)) > 0:
                if assignment.get(s.slot_id) is None:
                    for c in s.columns:
                        blank_cols[str(c)].append(day.date.isoformat())
        # K/T share (only when K and T are separate slots)
        sk = next((s for s in day_slots if s.columns == ["K"]), None)
        st = next((s for s in day_slots if s.columns == ["T"]), None)
        if sk and st:
            dk = assignment.get(sk.slot_id)
            dt_ = assignment.get(st.slot_id)
            if dk is not None and dk == dt_:
                kt_share_days.append(day.date.isoformat())
        # D/F share (when fallback forces/needs D=F)
        sD = next((s for s in day_slots if s.columns == ["D"]), None)
        sF = next((s for s in day_slots if s.columns == ["F"]), None)
        if sD and sF:
            dD = assignment.get(sD.slot_id)
            dF = assignment.get(sF.slot_id)
            if dD is not None and dD == dF:
                df_share_days.append(day.date.isoformat())
                if bool(getattr(sD, "force_same_doctor", False)) and bool(getattr(sF, "force_same_doctor", False)):
                    df_forced_same_days.append(day.date.isoformat())
    return {
        "kt_share_days": kt_share_days,
        "df_share_days": df_share_days,
        "df_forced_same_days": df_forced_same_days,
        "blank_columns": dict(blank_cols),
    }

def build_daily_diagnostic(
    days: List[DayRow],
    slots: List[Slot],
    assignment: Dict[str, Optional[str]],
    cfg: dict,
) -> List[Dict]:
    """Build a per-day diagnostic: blanks, relief valves, forced blanks."""
    slots_by_day: Dict[dt.date, List[Slot]] = defaultdict(list)
    for s in slots:
        slots_by_day[s.day.date].append(s)

    DOW_ITA = {"Mon": "Lun", "Tue": "Mar", "Wed": "Mer", "Thu": "Gio",
               "Fri": "Ven", "Sat": "Sab", "Sun": "Dom"}
    diagnostics: List[Dict] = []

    for day in days:
        day_slots = slots_by_day.get(day.date, [])
        issues: List[str] = []

        for s in day_slots:
            doc = assignment.get(s.slot_id)
            cols_str = "+".join(s.columns)

            if doc is None:
                if getattr(s, "empty_domain", False):
                    issues.append(f"col {cols_str} vuota (nessun candidato eleggibile)")
                elif s.required and int(getattr(s, "blank_penalty", 0)) >= 5_000_000:
                    issues.append(f"col {cols_str} NON COPERTA (obbligatoria)")
                elif int(getattr(s, "blank_penalty", 0)) > 0:
                    bp = int(getattr(s, "blank_penalty", 0))
                    issues.append(f"col {cols_str} blank (relief valve, penalty={bp})")
                elif not s.required:
                    pass  # optional slot left blank is normal
                else:
                    issues.append(f"col {cols_str} blank (non assegnata)")

        # K/T same doctor
        sk = next((s for s in day_slots if s.columns == ["K"]), None)
        st_ = next((s for s in day_slots if s.columns == ["T"]), None)
        if sk and st_:
            dk = assignment.get(sk.slot_id)
            dt_ = assignment.get(st_.slot_id)
            if dk is not None and dk == dt_:
                issues.append(f"K=T stesso medico ({dk})")

        # D/F same doctor
        sD = next((s for s in day_slots if s.columns == ["D"]), None)
        sF = next((s for s in day_slots if s.columns == ["F"]), None)
        if sD and sF:
            dD = assignment.get(sD.slot_id)
            dF = assignment.get(sF.slot_id)
            if dD is not None and dD == dF:
                forced = getattr(sD, "force_same_doctor", False) and getattr(sF, "force_same_doctor", False)
                tag = " (forzato)" if forced else ""
                issues.append(f"D=F stesso medico ({dD}){tag}")

        if issues:
            festivo = is_festivo(day, cfg)
            diagnostics.append({
                "date": day.date.isoformat(),
                "dow": DOW_ITA.get(day.dow, day.dow),
                "festivo": festivo,
                "issues": issues,
            })

    return diagnostics

def solve_with_ortools(
    cfg: dict,
    days: List[DayRow],
    slots: List[Slot],
    fixed_assignments: Optional[List[dict]] = None,
    availability_preferences: Optional[List[dict]] = None,
    unav_map: Optional[Dict[str, Dict[dt.date, Set[str]]]] = None,
    historical_stats: Optional[dict] = None,
) -> Tuple[Dict[str, Optional[str]], Dict]:
    """
    Returns:
      assignment: slot_id -> doctor (or None for optional left blank)
      stats: dict with diagnostic info

    fixed_assignments: [{"doctor": str, "date": "YYYY-MM-DD", "column": str}, ...]
      Vincolo HARD: quel medico DEVE comparire in quella colonna quel giorno.
    availability_preferences: [{"doctor": str, "date": "YYYY-MM-DD", "shift": str}, ...]
      Vincolo SOFT: il solver prova a far comparire il medico in qualsiasi slot
      della fascia indicata in quel giorno.
    unav_map: mappa indisponibilità, usata per calcolare il target universitari corretto.
    """
    try:
        from ortools.sat.python import cp_model
    except Exception as e:
        raise RuntimeError("OR-Tools not installed. Install with: pip install ortools") from e
    model = cp_model.CpModel()
    # Collect extra objective terms built during constraint setup
    extra_obj = []
    # Warnings raccolti durante la costruzione del modello (non bloccanti)
    pre_solve_warnings: List[str] = []

    # Diagnostics for the "Recupero su T in 2 lunedì" rule.
    # NOTE: this MUST be defined in the OR-Tools path too; otherwise the code may
    # crash after solving and trigger the greedy fallback.
    MonT_need_rec = 0
    MonT_target_dates: List[dt.date] = []
    doctors = collect_doctors(cfg)
    doctors = [d for d in doctors if d != "Recupero"] + (["Recupero"] if "Recupero" in doctors else [])
    doc_to_idx = {d:i for i,d in enumerate(doctors)}
    # Decision vars: x[(slot_id, doc)] in {0,1}
    x = {}
    for s in slots:
        for d in s.allowed:
            if d not in doc_to_idx:
                continue
            x[(s.slot_id, d)] = model.NewBoolVar(f"x_{hash(s.slot_id)%10**8}_{hash(d)%10**8}")
    # Slot assignment constraints.
    # Required slots use a "soft-required" approach: sum(vars) + b_blank == 1.
    # b_blank == 1 means the slot is left empty (penalized at 5_000_000).
    # This makes the model ALWAYS feasible and lets the solver identify
    # which slots genuinely cannot be filled (they show up blank in the output).
    # Gerarchia sacrifici (crescente = si sacrifica per primo):
    #   K=T share: 5K  →  L blank: 20K  →  R blank: 40K  →  altri required: 5M  →  H/I/J: 50M (MAI)
    BLANK_REQUIRED_PENALTY = 5_000_000
    BLANK_CRITICAL_PENALTY = 50_000_000  # H, I, J: praticamente mai vuoti
    _CRITICAL_TAGS = {"H", "I", "J", "Festivo_HI"}
    blank_required_vars: Dict[str, object] = {}  # slot_id -> b_blank var (for diagnostics)
    for s in slots:
        vars_ = [x[(s.slot_id, d)] for d in s.allowed if (s.slot_id, d) in x]
        if not vars_:
            # No eligible doctors at all → slot will be blank (no variable to add penalty to)
            continue
        if s.required:
            b_blank = model.NewBoolVar(f"blank_req_{hash(s.slot_id)%10**8}")
            model.Add(sum(vars_) + b_blank == 1)
            penalty = BLANK_CRITICAL_PENALTY if s.rule_tag in _CRITICAL_TAGS else BLANK_REQUIRED_PENALTY
            extra_obj.append(penalty * b_blank)
            blank_required_vars[s.slot_id] = b_blank
        else:
            # Optional slot: can be left blank. If blank_penalty>0, we penalize blanks so it is used only as a last resort.
            if getattr(s, "blank_penalty", 0) and int(getattr(s, "blank_penalty", 0)) > 0:
                # Optional slot may be left blank, but only as a last resort: we add a large penalty
                # term to the objective. This must go into `extra_obj` because `objective_terms`
                # is defined later, when all high-level soft constraints are assembled.
                b = model.NewBoolVar(f"blank_{hash(s.slot_id)%10**8}")
                model.Add(sum(vars_) + b == 1)
                extra_obj.append(b * int(getattr(s, "blank_penalty", 0)))
            else:
                model.Add(sum(vars_) <= 1)
    # Penalità per medici di emergenza (non nel pool primario della colonna)
    # Il solver li usa solo se non c'è alternativa migliore (penalità < blank penalty).
    EMERGENCY_FILL_PENALTY = 2_000_000  # < BLANK_REQUIRED_PENALTY (5M) → meglio di blank
    for s in slots:
        for _emerg_doc in (s.emergency_doctors or []):
            _ev = x.get((s.slot_id, norm_name(_emerg_doc)))
            if _ev is not None:
                extra_obj.append(EMERGENCY_FILL_PENALTY * _ev)

    # Helper: slots per day
    slots_by_day: Dict[dt.date, List[Slot]] = defaultdict(list)
    for s in slots:
        slots_by_day[s.day.date].append(s)

    # ---------------------------------------------------------------
    # ASSEGNAZIONI FISSE (hard): l'admin ha fissato un medico in una
    # colonna specifica in un giorno specifico.
    # ---------------------------------------------------------------
    for fa in (fixed_assignments or []):
        try:
            fa_date = dt.date.fromisoformat(str(fa.get("date","")).strip())
            fa_doc = norm_name(str(fa.get("doctor","")).strip())
            fa_col = str(fa.get("column","")).strip().upper()
        except Exception:
            continue
        if fa_doc not in doc_to_idx:
            continue
        # Trova lo slot corrispondente
        target_slots = [s for s in slots_by_day.get(fa_date, [])
                        if fa_col in [str(c).strip().upper() for c in (s.columns or [])]]
        if not target_slots:
            continue
        for ts in target_slots:
            tv = x.get((ts.slot_id, fa_doc))
            if tv is None:
                # Il medico non è nell'allowed originale — aggiunge la variabile
                tv = model.NewBoolVar(f"x_{hash(ts.slot_id)%10**8}_{hash(fa_doc)%10**8}_forced")
                x[(ts.slot_id, fa_doc)] = tv
            model.Add(tv == 1)  # HARD: questo medico deve essere assegnato
            # Forza tutti gli altri a 0 in questo slot
            for d2 in list(ts.allowed) + list(doctors):
                d2n = norm_name(d2)
                if d2n == fa_doc:
                    continue
                v2 = x.get((ts.slot_id, d2n))
                if v2 is not None:
                    model.Add(v2 == 0)

    # ---------------------------------------------------------------
    # DISPONIBILITÀ (soft): il medico VUOLE comparire in una certa fascia.
    # Penalizziamo se NON compare in nessuno slot di quella fascia in quel giorno.
    # ---------------------------------------------------------------
    SHIFT_MAP_AVAIL = {
        "mattina": "Mattina", "morning": "Mattina",
        "pomeriggio": "Pomeriggio", "afternoon": "Pomeriggio",
        "notte": "Notte", "night": "Notte",
        "diurno": "Diurno", "day": "Diurno",
    }
    # Penalità per disponibilità non rispettata.
    # Calibrazione rispetto alle penalità strutturali D/F:
    #   missing_pair_doc=12000, df_share=8000, prefer_df_share=15000
    #   → stack peggiore ~35000
    # "media" (40000) supera ogni singola penalità strutturale e lo stack comune.
    # "alta"  (80000) supera anche gli stack peggiori.
    # "bassa" (15000) può essere superata da singole penalità strutturali (intenzionale).
    AVAIL_PENALTY_BASE = 40_000
    AVAIL_PRIORITY_MULT = {"alta": 2.0, "alta priority": 2.0, "media": 1.0, "bassa": 0.375}
    pref_skipped_log: list = []
    for ap in (availability_preferences or []):
        try:
            ap_date = dt.date.fromisoformat(str(ap.get("date","")).strip())
            ap_doc = norm_name(str(ap.get("doctor","")).strip())
            ap_shift_raw = str(ap.get("shift","")).strip().lower()
            ap_shift = SHIFT_MAP_AVAIL.get(ap_shift_raw, ap_shift_raw.capitalize())
            ap_priority = str(ap.get("priority", "media")).strip().lower()
        except Exception:
            continue
        if ap_doc not in doc_to_idx:
            continue
        mult = AVAIL_PRIORITY_MULT.get(ap_priority, 1.0)
        penalty = int(AVAIL_PENALTY_BASE * mult)
        # Raccoglie tutti i vars dello slot per quella fascia in quel giorno
        ap_vars = []
        for s in slots_by_day.get(ap_date, []):
            s_shift = getattr(s, "shift", "") or ""
            if ap_shift.lower() in s_shift.lower() or s_shift.lower() in ap_shift.lower():
                v = x.get((s.slot_id, ap_doc))
                if v is not None:
                    ap_vars.append(v)
        if ap_vars:
            # b = 1 se il medico compare in almeno uno slot di quella fascia
            b_avail = model.NewBoolVar(f"avail_{ap_date}_{hash(ap_doc)%10**6}_{ap_shift}")
            model.AddMaxEquality(b_avail, ap_vars)
            not_avail = model.NewBoolVar(f"notavail_{ap_date}_{hash(ap_doc)%10**6}_{ap_shift}")
            model.Add(not_avail + b_avail == 1)
            extra_obj.append(penalty * not_avail)
        else:
            pref_skipped_log.append(
                f"  AVAIL SKIP {ap_doc} {ap_date} {ap_shift}: non in nessuno slot ammesso"
            )
    if pref_skipped_log:
        print("WARNING: Preferenze disponibilità non applicabili (medico fuori pool):\n" +
              "\n".join(pref_skipped_log))


    # Uniqueness per day: one doctor max 1 slot/day (exceptions already handled by merged columns)
    gc = cfg.get("global_constraints") or {}
    relief = gc.get("relief_valves") or {}
    enable_kt_share = bool(relief.get("enable_kt_share", False))
    kt_share_penalty = int(relief.get("kt_share_penalty", 5000))

    # D/F share valve (allows the same doctor to cover both D and F ONLY if needed)
    rules_map = cfg.get("rules", {}) or {}
    r_df = rules_map.get("D_F", {}) if isinstance(rules_map.get("D_F", {}), dict) else {}
    enable_df_share = bool(r_df.get("enable_df_share", True))
    df_share_penalty = int(r_df.get("df_share_penalty", 8000))
    prefer_df_share_penalty = int(r_df.get("prefer_df_share_penalty", 15000))

    daily_exempt_cols = {str(c).strip().upper() for c in (gc.get('daily_uniqueness_exempt_columns') or [])}
    def _slot_is_exempt_daily(s: Slot) -> bool:
        return any(str(c).strip().upper() in daily_exempt_cols for c in (s.columns or []))

    for day in days:
        day_slots = slots_by_day.get(day.date, [])
        uniq_slots = [s for s in day_slots if not _slot_is_exempt_daily(s)]

        # Slack vars per doctor (can allow +1 assignment for a specific "share" valve)
        slack_vars_by_doc: Dict[str, List] = defaultdict(list)

        # ---- K/T emergency share
        kt_y_by_doc = None
        slotK = None
        slotT = None
        if enable_kt_share:
            for s in day_slots:
                if s.columns == ["K"]:
                    slotK = s
                elif s.columns == ["T"]:
                    slotT = s
        if enable_kt_share and slotK is not None and slotT is not None:
            kt_y_by_doc = {}
            for d in doctors:
                if (slotK.slot_id, d) in x and (slotT.slot_id, d) in x:
                    y = model.NewBoolVar(f"kt_same_{day.date.isoformat()}_{hash(d)%10**6}")
                    model.Add(y <= x[(slotK.slot_id, d)])
                    model.Add(y <= x[(slotT.slot_id, d)])
                    model.Add(y >= x[(slotK.slot_id, d)] + x[(slotT.slot_id, d)] - 1)
                    kt_y_by_doc[d] = y
            if kt_y_by_doc:
                model.Add(sum(kt_y_by_doc.values()) <= 1)
                extra_obj.append(sum(kt_y_by_doc.values()) * kt_share_penalty)
                for d, y in kt_y_by_doc.items():
                    slack_vars_by_doc[d].append(y)

        # ---- D/F emergency share (used only when the D/F pair is infeasible without it)
        df_y_by_doc = None
        slotD = None
        slotF = None
        if enable_df_share:
            for s in day_slots:
                if s.columns == ["D"]:
                    slotD = s
                elif s.columns == ["F"]:
                    slotF = s
        if enable_df_share and slotD is not None and slotF is not None:
            df_y_by_doc = {}
            for d in doctors:
                if (slotD.slot_id, d) in x and (slotF.slot_id, d) in x:
                    y = model.NewBoolVar(f"df_same_{day.date.isoformat()}_{hash(d)%10**6}")
                    model.Add(y <= x[(slotD.slot_id, d)])
                    model.Add(y <= x[(slotF.slot_id, d)])
                    model.Add(y >= x[(slotD.slot_id, d)] + x[(slotF.slot_id, d)] - 1)
                    df_y_by_doc[d] = y
            if df_y_by_doc:
                model.Add(sum(df_y_by_doc.values()) <= 1)

                # If D/F is in an "emergency" mode (both Grimaldi/Calabrò unavailable, or H pool empty with only one available),
                # we PREFER using the same doctor (D=F) to avoid blanks/infeasible. In that case we do NOT penalize df_share,
                # and we penalize NOT sharing.
                prefer_share = bool(getattr(slotD, "force_same_doctor", False)) and bool(getattr(slotF, "force_same_doctor", False))
                share_pen = 0 if prefer_share else df_share_penalty
                if share_pen:
                    extra_obj.append(sum(df_y_by_doc.values()) * share_pen)

                if prefer_share:
                    share_any = model.NewBoolVar(f"df_share_any_{day.date.isoformat()}")
                    model.AddMaxEquality(share_any, list(df_y_by_doc.values()))
                    notshare = model.NewBoolVar(f"df_notshare_{day.date.isoformat()}")
                    model.Add(notshare + share_any == 1)
                    extra_obj.append(notshare * prefer_df_share_penalty)

                for d, y in df_y_by_doc.items():
                    slack_vars_by_doc[d].append(y)

        # Prevent a single doctor from stacking multiple share valves on the same day (safety)
        if kt_y_by_doc and df_y_by_doc:
            for d in doctors:
                y1 = kt_y_by_doc.get(d)
                y2 = df_y_by_doc.get(d)
                if y1 is not None and y2 is not None:
                    model.Add(y1 + y2 <= 1)

        # Daily uniqueness constraint with optional +1 slack per share valve
        for d in doctors:
            vars_ = []
            for s in uniq_slots:
                if (s.slot_id, d) in x:
                    vars_.append(x[(s.slot_id, d)])
            if not vars_:
                continue
            slack_terms = slack_vars_by_doc.get(d, [])
            if slack_terms:
                model.Add(sum(vars_) <= 1 + sum(slack_terms))
            else:
                model.Add(sum(vars_) <= 1)
# ---- H/I: divieto di stesso medico in giorni consecutivi (H e I indipendenti)
    # Vale anche sui Festivi, dove esiste lo slot unico HI (colonne H+I).
    try:
        slot_ids_by_date_col: Dict[Tuple[dt.date, str], List[str]] = defaultdict(list)
        for s in slots:
            for c in (s.columns or []):
                cc = str(c).strip().upper()
                if cc in {"H", "I"}:
                    slot_ids_by_date_col[(s.day.date, cc)].append(s.slot_id)
        days_sorted = sorted(days, key=lambda d: d.date)
        for i in range(len(days_sorted) - 1):
            d1 = days_sorted[i].date
            d2 = days_sorted[i + 1].date
            for col in ("H", "I"):
                sids1 = slot_ids_by_date_col.get((d1, col), [])
                sids2 = slot_ids_by_date_col.get((d2, col), [])
                if not sids1 and not sids2:
                    continue
                for doc in doctors:
                    if doc == "Recupero":
                        continue
                    v1 = [x.get((sid, doc)) for sid in sids1 if (sid, doc) in x]
                    v2 = [x.get((sid, doc)) for sid in sids2 if (sid, doc) in x]
                    v1 = [v for v in v1 if v is not None]
                    v2 = [v for v in v2 if v is not None]
                    if v1 or v2:
                        model.Add(sum(v1) + sum(v2) <= 1)
    except Exception:
        pass

    # ---- D/F pattern 3+3 for Grimaldi/Calabrò (conditional hard)
    # If configured in rules.D_F (pattern_3_3 + pattern_conditional_hard),
    # enforce:
    #   Mon–Wed: D=doc1, F=doc2
    #   Thu–Sat: D=doc2, F=doc1
    # Pattern can be violated ONLY on days where (H is doc1/doc2) OR (J is doc2).
    rules_map = cfg.get("rules", {}) or {}
    r_df = rules_map.get("D_F", {}) if isinstance(rules_map.get("D_F", {}), dict) else {}
    if r_df.get("pattern_3_3") and r_df.get("pattern_conditional_hard"):
        doc1 = norm_name(r_df.get("pattern_doc1") or "")
        doc2 = norm_name(r_df.get("pattern_doc2") or "")
        if doc1 in doctors and doc2 in doctors:
            _day_idx_map = {d.date: i for i, d in enumerate(days)}
            for day in days:
                if day.dow not in ["Mon","Tue","Wed","Thu","Fri","Sat"]:
                    continue
                dslot = f"{day.date}-D"
                fslot = f"{day.date}-F"
                # if D/F slots don't exist for this day, skip
                v_d1 = x.get((dslot, doc1))
                v_d2 = x.get((dslot, doc2))
                v_f1 = x.get((fslot, doc1))
                v_f2 = x.get((fslot, doc2))
                if (v_d1 is None and v_d2 is None) or (v_f1 is None and v_f2 is None):
                    continue
                # Exception = (H is doc1/doc2) OR (J is doc1/doc2 oggi) OR (J is doc1/doc2 ieri)
                # Serve anche "J ieri": se doc2 ha fatto notte ieri, night_off next_day
                # gli vieta D oggi → il pattern non va imposto (altrimenti INFEASIBLE).
                conds = []
                hslot = f"{day.date}-H"
                jslot = f"{day.date}-J"
                for v in [x.get((hslot, doc1)), x.get((hslot, doc2)),
                          x.get((jslot, doc1)), x.get((jslot, doc2))]:
                    if v is not None:
                        conds.append(v)
                # J del giorno precedente
                _i_today = _day_idx_map.get(day.date)
                if _i_today is not None and _i_today > 0:
                    _prev_date = days[_i_today - 1].date
                    _prev_jslot = f"{_prev_date}-J"
                    for v in [x.get((_prev_jslot, doc1)), x.get((_prev_jslot, doc2))]:
                        if v is not None:
                            conds.append(v)
                if conds:
                    exc = model.NewBoolVar(f"exc_DF_{day.date}")
                    model.AddMaxEquality(exc, conds)  # exc = OR(conds)
                else:
                    exc = None
                if day.dow in ["Mon","Tue","Wed"]:
                    # Only enforce the 3-day pattern if BOTH required assignments are actually possible
                    # (i.e., corresponding variables exist after unavailability / domain filtering).
                    # Otherwise, enforcing only one side (e.g. F=doc2 while D=doc1 cannot happen)
                    # can make the whole month INFEASIBLE due to daily uniqueness constraints.
                    if v_d1 is None or v_f2 is None:
                        continue
                    # enforce D=doc1, F=doc2 when not exception
                    if v_d1 is not None:
                        if exc is None:
                            model.Add(v_d1 == 1)
                        else:
                            model.Add(v_d1 == 1).OnlyEnforceIf(exc.Not())
                    if v_f2 is not None:
                        if exc is None:
                            model.Add(v_f2 == 1)
                        else:
                            model.Add(v_f2 == 1).OnlyEnforceIf(exc.Not())
                else:
                    if v_d2 is None or v_f1 is None:
                        continue
                    # Thu/Fri/Sat: D=doc2, F=doc1 when not exception
                    if v_d2 is not None:
                        if exc is None:
                            model.Add(v_d2 == 1)
                        else:
                            model.Add(v_d2 == 1).OnlyEnforceIf(exc.Not())
                    if v_f1 is not None:
                        if exc is None:
                            model.Add(v_f1 == 1)
                        else:
                            model.Add(v_f1 == 1).OnlyEnforceIf(exc.Not())
    # Night off next day
    gc = cfg.get("global_constraints", {}) or {}
    night_off = (gc.get("night_off") or {})
    night_same = bool(night_off.get("same_day", True))
    night_next = bool(night_off.get("next_day", True))
    # Identify night slots
    night_slot_ids = [s.slot_id for s in slots if "J" in s.columns]
    # Night-off same day: if doctor works night(d) → no other slot(d) [except C which is exempt]
    if night_same:
        for drow in days:
            night_vars_day = []
            for sid in night_slot_ids:
                if sid.startswith(str(drow.date)):
                    for doc in doctors:
                        if (sid, doc) in x:
                            pass  # handled per-doc below
            for doc in doctors:
                night_vars = [x[(sid, doc)] for sid in night_slot_ids
                              if sid.startswith(str(drow.date)) and (sid, doc) in x]
                if not night_vars:
                    continue
                night_var = night_vars[0]
                # Every non-night, non-C slot same day → 0 if night=1
                for s2 in slots_by_day.get(drow.date, []):
                    if "J" in s2.columns:
                        continue  # skip the night slot itself
                    if any(c in {"C"} for c in (s2.columns or [])):
                        continue  # C is exempt
                    if (s2.slot_id, doc) in x:
                        model.Add(x[(s2.slot_id, doc)] == 0).OnlyEnforceIf(night_var)
    # Identify night slots
    night_slot_ids = [s.slot_id for s in slots if "J" in s.columns]
    if night_next:
        # for each day (except last), if doctor works night(d) then no slot(d+1)
        day_index = {d.date:i for i,d in enumerate(days)}
        for drow in days:
            i = day_index[drow.date]
            if i+1 >= len(days):
                continue
            next_day = days[i+1].date
            for doc in doctors:
                night_vars = []
                for sid in night_slot_ids:
                    if sid.startswith(str(drow.date)) and (sid, doc) in x:
                        night_vars.append(x[(sid, doc)])
                if not night_vars:
                    continue
                night_var = night_vars[0]  # only one night slot per day in our model
                # For every slot on next day, doc cannot be assigned if night_var=1
                for s2 in slots_by_day.get(next_day, []):
                    if (s2.slot_id, doc) in x:
                        model.Add(x[(s2.slot_id, doc)] == 0).OnlyEnforceIf(night_var)
    # Night spacing min
    min_gap = int(gc.get("night_spacing_days_min", 5))
    # Build night vars by (day,doc)
    night_var_by_day_doc = {}
    for s in slots:
        if "J" in s.columns:
            for doc in doctors:
                if (s.slot_id, doc) in x:
                    night_var_by_day_doc[(s.day.date, doc)] = x[(s.slot_id, doc)]
    # For each doc, prevent nights too close
    for doc in doctors:
        for i, drow in enumerate(days):
            for k in range(1, min_gap):
                j = i + k
                if j >= len(days):
                    break
                d1, d2 = drow.date, days[j].date
                v1 = night_var_by_day_doc.get((d1, doc))
                v2 = night_var_by_day_doc.get((d2, doc))
                if v1 is not None and v2 is not None:
                    model.Add(v1 + v2 <= 1)
    # Reperibilità constraints: not night same day + next 2 days
    if "rules" in cfg and "C_reperibilita" in cfg["rules"]:
        rC = cfg["rules"]["C_reperibilita"]
        constraints = set(rC.get("constraints") or [])
        if any(c in constraints for c in ("not_night_same_day", "not_night_next_2_days", "not_night_prev_2_days")):
            pos = {d.date: i for i, d in enumerate(days)}
            for day in days:
                c_slot = next((s for s in slots_by_day[day.date] if s.columns == ["C"]), None)
                if not c_slot:
                    continue
                for doc in doctors:
                    if (c_slot.slot_id, doc) not in x:
                        continue
                    cvar = x[(c_slot.slot_id, doc)]
                    i = pos[day.date]
                    # enforce no-night within ±2 days of a Reperibilità assignment
                    for off in [-2, -1, 0, 1, 2]:
                        j = i + off
                        if j < 0 or j >= len(days):
                            continue
                        nvar = night_var_by_day_doc.get((days[j].date, doc))
                        if nvar is not None:
                            model.Add(nvar == 0).OnlyEnforceIf(cvar)
    # K no consecutive days same doctor
    if "rules" in cfg and "K" in cfg["rules"] and cfg["rules"]["K"].get("no_consecutive_days_same_doctor", False):
        for i in range(len(days)-1):
            d1, d2 = days[i].date, days[i+1].date
            k1 = next((s for s in slots_by_day[d1] if "K" in s.columns), None)
            k2 = next((s for s in slots_by_day[d2] if "K" in s.columns), None)
            if not k1 or not k2:
                continue
            for doc in doctors:
                v1 = x.get((k1.slot_id, doc))
                v2 = x.get((k2.slot_id, doc))
                if v1 is not None and v2 is not None:
                    model.Add(v1 + v2 <= 1)
    
# D/F weekly behavior (Mon–Sat)
    # Goal (robust, NEVER infeasible):
    #   - Prefer Grimaldi + Calabrò on D/F when both are available
    #   - If only one of the pair is available: prefer D=that doctor; prefer F from H.pool_mon_fri
    #   - If none are available: prefer D and F from H.pool_mon_fri and prefer D=F (share)
    #
    # IMPORTANT: These are implemented as SOFT constraints with high penalties, so the model
    # can still remain FEASIBLE in "tight" days (lots of unavailability / other hard rules).
    if "rules" in cfg and "D_F" in cfg["rules"] and isinstance(cfg["rules"]["D_F"], dict):
        rDF = cfg["rules"]["D_F"]
        if rDF.get("pattern_3_3", False):
            doc1 = norm_name(rDF.get("pattern_doc1") or "Grimaldi")
            doc2 = norm_name(rDF.get("pattern_doc2") or "Calabrò")
            pair = {doc1, doc2}

            # Penalties (tunable in YAML)
            pen_pattern = int(rDF.get("pattern_penalty", 80) or 80)
            pen_outside_pair = int(rDF.get("outside_pair_penalty", 6000) or 6000)
            pen_missing_pair_doc = int(rDF.get("missing_pair_doc_penalty", 12000) or 12000)
            pen_d_not_available = int(rDF.get("d_not_available_penalty", 15000) or 15000)
            pen_outside_hpool = int(rDF.get("outside_hpool_penalty", 5000) or 5000)

            hpool = set(norm_name(d) for d in ((cfg.get("rules", {}).get("H", {}) or {}).get("pool_mon_fri") or []))

            for day in days:
                if day.dow not in ["Mon","Tue","Wed","Thu","Fri","Sat"]:
                    continue
                sD = next((s for s in slots_by_day[day.date] if s.columns == ["D"]), None)
                sF = next((s for s in slots_by_day[day.date] if s.columns == ["F"]), None)
                if not sD or not sF:
                    continue

                vD1 = x.get((sD.slot_id, doc1)); vF1 = x.get((sF.slot_id, doc1))
                vD2 = x.get((sD.slot_id, doc2)); vF2 = x.get((sF.slot_id, doc2))
                avail_pair = []
                if vD1 is not None and vF1 is not None:
                    avail_pair.append(doc1)
                if vD2 is not None and vF2 is not None:
                    avail_pair.append(doc2)

                # Helper: penalize choosing a doctor outside a set only if the set is actually feasible in-domain
                def penalize_outside(slot: Slot, allowed_set: set, penalty: int):
                    if not any((doc in allowed_set) and ((slot.slot_id, doc) in x) for doc in slot.allowed):
                        return
                    for doc in slot.allowed:
                        v = x.get((slot.slot_id, doc))
                        if v is not None and doc not in allowed_set:
                            extra_obj.append(penalty * v)

                # CASE: both pair doctors are available in-domain
                if len(avail_pair) == 2:
                    # Strongly prefer staying within the pair on BOTH slots
                    penalize_outside(sD, pair, pen_outside_pair)
                    penalize_outside(sF, pair, pen_outside_pair)

                    # Prefer that BOTH doctors appear across {D,F}
                    # (soft: if impossible due to other hard rules, solver may pay the penalty)
                    used1 = model.NewBoolVar(f"df_used_{day.date}_1")
                    used2 = model.NewBoolVar(f"df_used_{day.date}_2")
                    model.AddMaxEquality(used1, [v for v in [vD1, vF1] if v is not None])
                    model.AddMaxEquality(used2, [v for v in [vD2, vF2] if v is not None])
                    n1 = model.NewBoolVar(f"df_notused_{day.date}_1")
                    n2 = model.NewBoolVar(f"df_notused_{day.date}_2")
                    model.Add(n1 + used1 == 1)
                    model.Add(n2 + used2 == 1)
                    extra_obj.append(n1 * pen_missing_pair_doc)
                    extra_obj.append(n2 * pen_missing_pair_doc)

                    # Weekly pattern preference (only among the pair)
                    prefD, prefF = (doc1, doc2) if day.dow in ["Mon","Tue","Wed"] else (doc2, doc1)
                    for doc in pair:
                        v = x.get((sD.slot_id, doc))
                        if v is not None and doc != prefD:
                            extra_obj.append(pen_pattern * v)
                    for doc in pair:
                        v = x.get((sF.slot_id, doc))
                        if v is not None and doc != prefF:
                            extra_obj.append(pen_pattern * v)

                # CASE: only one of the pair is available in-domain
                elif len(avail_pair) == 1:
                    only_doc = avail_pair[0]
                    vD_only = x.get((sD.slot_id, only_doc))
                    vF_only = x.get((sF.slot_id, only_doc))

                    # HARD: il solo medico disponibile del pair DEVE stare in D
                    if vD_only is not None:
                        model.Add(vD_only == 1)

                    # HARD: il solo medico disponibile del pair DEVE stare anche in F (D=F share)
                    # Questo è il comportamento richiesto: se Grimaldi è indisponibile,
                    # Calabrò deve coprire sia D che F.
                    if vF_only is not None:
                        model.Add(vF_only == 1)

                    # NON penalizzare altri medici in F (l'unico del pair li occupa entrambi)

                # CASE: none of the pair is available in-domain
                else:
                    # HARD: D e F devono essere assegnati allo stesso medico dell'H-pool.
                    # Usiamo il vincolo df_share già costruito sopra (df_y_by_doc) per forzare la share.
                    # Penalizziamo solo dottori fuori dall'H-pool per orientare la scelta.
                    penalize_outside(sD, hpool, pen_outside_hpool)
                    penalize_outside(sF, hpool, pen_outside_hpool)
                    # Forza D=F tramite il meccanismo df_share già presente
                    if df_y_by_doc:
                        share_any = model.NewBoolVar(f"df_share_none_{day.date.isoformat()}")
                        model.AddMaxEquality(share_any, list(df_y_by_doc.values()))
                        model.Add(share_any == 1)  # HARD: deve esserci un medico che copre sia D che F
                    else:
                        pre_solve_warnings.append(
                            f"D/F il {day.date}: nessun medico disponibile per coprire entrambe le colonne (D e F lasciate scoperte)"
                        )
# E/G weekly blocks (Mon-Sat) if block_days=6
    if "rules" in cfg and "E_G" in cfg["rules"]:
        block_days = int(cfg["rules"]["E_G"].get("block_days", 0) or 0)
        # Find all EG slots by date
        eg_by_date = {}
        for s in slots:
            if s.columns == ["E","G"]:
                eg_by_date[s.day.date] = s
        if block_days == 6:
            # For each Monday, enforce same doctor Mon..Sat (if all present)
            for day in days:
                if day.dow != "Mon":
                    continue
                seq = [day.date + dt.timedelta(days=k) for k in range(0,6)]
                if not all(d in eg_by_date for d in seq):
                    continue
                for doc in doctors:
                    v0 = x.get((eg_by_date[seq[0]].slot_id, doc))
                    if v0 is None:
                        continue
                    for d in seq[1:]:
                        v = x.get((eg_by_date[d].slot_id, doc))
                        if v is not None:
                            model.Add(v == v0)
        elif block_days == 3:
            # Split week into 2 blocks: Mon-Wed and Thu-Sat (if all present).
            for day in days:
                if day.dow != "Mon":
                    continue
                seq1 = [day.date + dt.timedelta(days=k) for k in range(0,3)]
                seq2 = [day.date + dt.timedelta(days=k) for k in range(3,6)]
                if not all(d in eg_by_date for d in (seq1 + seq2)):
                    continue
                for doc in doctors:
                    v0 = x.get((eg_by_date[seq1[0]].slot_id, doc))
                    if v0 is not None:
                        for d in seq1[1:]:
                            v = x.get((eg_by_date[d].slot_id, doc))
                            if v is not None:
                                model.Add(v == v0)
                for doc in doctors:
                    v3 = x.get((eg_by_date[seq2[0]].slot_id, doc))
                    if v3 is not None:
                        for d in seq2[1:]:
                            v = x.get((eg_by_date[d].slot_id, doc))
                            if v is not None:
                                model.Add(v == v3)
                # Soft preference: use two different doctors between the two 3-day blocks.
                # Penalize if the SAME doctor is chosen on Mon (block1 start) and Thu (block2 start).
                pen_split = int((cfg["rules"]["E_G"].get("block_split_penalty", 10) or 10))
                mon_slot = eg_by_date[seq1[0]]
                thu_slot = eg_by_date[seq2[0]]
                same_terms = []
                for doc in doctors:
                    vmon = x.get((mon_slot.slot_id, doc))
                    vthu = x.get((thu_slot.slot_id, doc))
                    if vmon is None or vthu is None:
                        continue
                    b = model.NewBoolVar(f"eg_same_{hash(doc)%10**6}_{day.date}")
                    model.AddBoolAnd([vmon, vthu]).OnlyEnforceIf(b)
                    model.AddBoolOr([vmon.Not(), vthu.Not()]).OnlyEnforceIf(b.Not())
                    same_terms.append(b)
                if same_terms:
                    extra_obj.append(pen_split * sum(same_terms))

        # E/G: bilanciamento hard — ogni medico può fare al massimo ceil(slots/pool_attivo)+1 blocchi
        eg_slots_all = list(eg_by_date.values())
        if eg_slots_all:
            rEG = cfg["rules"]["E_G"]
            eg_pool = [norm_name(d) for d in (rEG.get("allowed") or []) if norm_name(d) in doctors]
            if eg_pool:
                n_eg = len(eg_slots_all)
                import math
                # Conta solo i medici con almeno una variabile disponibile (tiene conto delle indisponibilità)
                eg_active_docs = [
                    d for d in eg_pool
                    if any(x.get((s.slot_id, d)) is not None for s in eg_slots_all)
                ]
                n_pool_active = max(len(eg_active_docs), 1)
                eg_max_hard = math.ceil(n_eg / n_pool_active) + 1  # basato su pool attivo
                eg_cnt_vars = []
                for doc in eg_pool:
                    vars_ = [x.get((s.slot_id, doc)) for s in eg_slots_all if x.get((s.slot_id, doc)) is not None]
                    if vars_:
                        eg_cnt = model.NewIntVar(0, n_eg, f"eg_cnt_{hash(doc)%10**6}")
                        model.Add(eg_cnt == sum(vars_))
                        model.Add(eg_cnt <= eg_max_hard)  # HARD cap
                        eg_cnt_vars.append(eg_cnt)
                if eg_cnt_vars:
                    eg_max_v = model.NewIntVar(0, n_eg, "eg_max_load")
                    model.AddMaxEquality(eg_max_v, eg_cnt_vars)
                    extra_obj.append(300 * eg_max_v)  # forte penalità per minimizzare il massimo
    # Monthly quotas (hard) — J
    # I fixed_assignments in J vengono applicati PRIMA (vedi sopra) e contano
    # già come variabili x=1 nel solver. Quindi la quota rimane SEMPRE == q:
    # se Calabrò ha quota=2 e 1 notte fissa, il solver trovera' esattamente 1
    # altra notte libera per arrivare a 2 totali.
    if "rules" in cfg and "J" in cfg["rules"]:
        mq = cfg["rules"]["J"].get("monthly_quotas") or {}
        for doc_raw, q in mq.items():
            doc = norm_name(doc_raw)
            if doc not in doctors:
                continue
            vars_ = [night_var_by_day_doc.get((d.date, doc)) for d in days]
            vars_ = [v for v in vars_ if v is not None]
            if not vars_:
                continue
            q_int = int(q)
            n_avail = len(vars_)
            if n_avail < q_int:
                # Quota irraggiungibile per indisponibilità — vincolo rilassato per evitare INFEASIBLE
                pre_solve_warnings.append(
                    f"J quota {doc}: richieste {q_int} notti ma solo {n_avail} disponibili "
                    f"(indisponibilità). Quota ridotta a {n_avail}."
                )
                model.Add(sum(vars_) == n_avail)
            else:
                model.Add(sum(vars_) == q_int)
    # Pool quota overrides max/min — da pool_config (tutti i tipi e colonne)
    qov = cfg.get("pool_quota_overrides") or {}
    for (doc_n, col), spec in qov.items():
        col_slots = [s for s in slots if col in (s.columns or [])]
        vars_ = [x.get((s.slot_id, doc_n)) for s in col_slots]
        vars_ = [v for v in vars_ if v is not None]
        if not vars_:
            continue
        sv = sum(vars_)
        qt = spec.get("type", "max")
        val = int(spec.get("value", 0))
        n_avail = len(vars_)
        if qt == "max":
            model.Add(sv <= val)
        elif qt == "min":
            effective_min = min(val, n_avail)
            if effective_min > 0:
                model.Add(sv >= effective_min)
            if effective_min < val:
                pre_solve_warnings.append(
                    f"Quota min {doc_n}/{col}: richiesto min {val} ma solo {n_avail} slot disponibili."
                )
        elif qt == "fixed":
            effective_val = min(val, n_avail)
            model.Add(sv == effective_val)
            if effective_val < val:
                pre_solve_warnings.append(
                    f"Quota fixed {doc_n}/{col}: richiesto {val} ma solo {n_avail} slot disponibili, ridotto a {effective_val}."
                )

    # Monthly quotas (hard) — Festivi DE+HI
    if "rules" in cfg and "Festivi" in cfg["rules"]:
        rFest = cfg["rules"]["Festivi"]
        fest_quotas = {norm_name(k): int(v) for k, v in (rFest.get("quotas") or {}).items()}
        if fest_quotas:
            festivo_slots = [s for s in slots if s.rule_tag in ("Festivo_DE", "Festivo_HI")]
            for doc, q in fest_quotas.items():
                if doc not in doctors:
                    continue
                vars_ = [x.get((s.slot_id, doc)) for s in festivo_slots]
                vars_ = [v for v in vars_ if v is not None]
                if vars_:
                    q_eff = min(q, len(vars_))
                    model.Add(sum(vars_) == q_eff)
                    if q_eff < q:
                        pre_solve_warnings.append(
                            f"Festivi quota {doc}: richiesti {q} ma solo {len(vars_)} slot disponibili."
                        )
    # Soft balance festivi — minimizza il massimo carico tra i medici del pool
    # senza quota fissa (evita Crea=3, Trio=0 ecc.)
    if "rules" in cfg and "Festivi" in cfg["rules"]:
        try:
            rFest = cfg["rules"]["Festivi"]
            fest_pool_raw = [norm_name(d) for d in (rFest.get("pool") or [])
                             if norm_name(d) in doctors and norm_name(d) != "Recupero"]
            fest_fixed = {norm_name(k) for k in (rFest.get("quotas") or {}).keys()}
            balance_pool = [d for d in fest_pool_raw if d not in fest_fixed]
            festivo_slots_all = [s for s in slots if s.rule_tag in ("Festivo_DE", "Festivo_HI")]
            fest_bal_w = int(rFest.get("balance_weight") or 500)
            if balance_pool and festivo_slots_all:
                max_fest = model.NewIntVar(0, len(festivo_slots_all), "max_fest_load")
                for d in balance_pool:
                    vars_d = [x[(s.slot_id, d)] for s in festivo_slots_all
                              if (s.slot_id, d) in x]
                    if not vars_d:
                        continue
                    load_d = model.NewIntVar(0, len(festivo_slots_all),
                                            f"fest_load_{hash(d) % 10**6}")
                    model.Add(load_d == sum(vars_d))
                    model.Add(load_d <= max_fest)
                extra_obj.append(fest_bal_w * max_fest)
        except Exception:
            pass
    # Soft: alcuni medici devono preferibilmente avere almeno N notti weekend (sab/dom)
    if "rules" in cfg and "J" in cfg["rules"]:
        rJ_wn = cfg["rules"]["J"]
        wn_min_soft = rJ_wn.get("weekend_night_min_soft") or {}
        wn_pen = int(rJ_wn.get("weekend_night_min_soft_penalty") or 3000)
        for doc_raw, min_we in wn_min_soft.items():
            doc = norm_name(doc_raw)
            if doc not in doctors:
                continue
            we_vars = [night_var_by_day_doc.get((d.date, doc))
                       for d in days if d.dow in ("Sat", "Sun")]
            we_vars = [v for v in we_vars if v is not None]
            if not we_vars:
                continue  # Zito indisponibile tutti i weekend: vincolo ignorato
            min_we = int(min_we)
            no_we = model.NewBoolVar(f"no_we_night_{hash(doc)%10**6}")
            we_sum = model.NewIntVar(0, len(we_vars), f"we_sum_{hash(doc)%10**6}")
            model.Add(we_sum == sum(we_vars))
            model.Add(we_sum < min_we).OnlyEnforceIf(no_we)
            model.Add(we_sum >= min_we).OnlyEnforceIf(no_we.Not())
            extra_obj.append(wn_pen * no_we)
    # Night distribution (HARD min/max per dottore + soft balance weekend)
    # Logica: total_nights = giorni del mese - giovedì (thursday_blank).
    # pool_available_nights esclude slot J pre-assegnate a medici fuori pool (es. festivi fissi).
    # Quota fissa (monthly_quotas YAML) sottratta → free_total diviso equamente tra free_docs.
    # Regola generale: ogni free doctor fa MIN floor(free_total/n), MAX floor+1 notti.
    if "rules" in cfg and "J" in cfg["rules"]:
        rJ = cfg["rules"]["J"]
        night_pool = set(norm_name(d) for d in (rJ.get("pool_other") or []))
        night_pool |= set(norm_name(d) for d in (rJ.get("monthly_quotas") or {}).keys())
        night_pool = {d for d in night_pool if d in doctors and d != "Recupero"}
        mq_fixed = {norm_name(k): int(v) for k,v in (rJ.get("monthly_quotas") or {}).items()
                    if norm_name(k) in doctors}
        total_nights = sum(1 for s in slots if s.columns == ["J"])
        # Notti disponibili per i pool doctors (esclude slot pre-assegnate a medici fuori pool)
        pool_available_nights = sum(
            1 for s in slots
            if s.columns == ["J"] and any(d in night_pool for d in s.allowed)
        )

        if night_pool and total_nights > 0:
            # Medici con quota fissa: già vincolati con == sopra.
            # Medici senza quota fissa: imponiamo min=2, max=3 hard.
            free_docs = [d for d in sorted(night_pool) if d not in mq_fixed]
            fixed_total = sum(mq_fixed.values())
            free_total = max(0, pool_available_nights - fixed_total)

            # Calcola min/max bilanciati per i medici liberi
            if free_docs:
                n_free = len(free_docs)
                # free_total / n_free → es. 21/9 = 2.33 → min=2, max=3
                min_per = free_total // n_free  # minimo garantito
                remainder = free_total - min_per * n_free
                # max_per = min_per se il resto è 0, altrimenti min_per+1
                max_per = min_per + (1 if remainder > 0 else 0)
                max_per = max(max_per, 0)  # sicurezza: mai negativo

                for doc in free_docs:
                    vars_ = [night_var_by_day_doc.get((d.date, doc)) for d in days
                             if night_var_by_day_doc.get((d.date, doc)) is not None]
                    if vars_:
                        model.Add(sum(vars_) >= min_per)   # HARD: minimo
                        model.Add(sum(vars_) <= max_per)   # HARD: massimo

                # Soft balance: minimizza la differenza max-min tra i medici liberi
                # per distribuire equamente il "resto"
                if remainder > 0 and len(free_docs) > 1:
                    cnt_vars = []
                    for doc in free_docs:
                        vars_ = [night_var_by_day_doc.get((d.date, doc)) for d in days
                                 if night_var_by_day_doc.get((d.date, doc)) is not None]
                        if vars_:
                            cnt = model.NewIntVar(0, total_nights, f"nightcnt_{hash(doc)%10**6}")
                            model.Add(cnt == sum(vars_))
                            cnt_vars.append(cnt)
                    if cnt_vars:
                        max_cnt = model.NewIntVar(0, total_nights, "night_max_free")
                        min_cnt = model.NewIntVar(0, total_nights, "night_min_free")
                        model.AddMaxEquality(max_cnt, cnt_vars)
                        model.AddMinEquality(min_cnt, cnt_vars)
                        diff_cnt = model.NewIntVar(0, total_nights, "night_diff_free")
                        model.Add(diff_cnt == max_cnt - min_cnt)
                        extra_obj.append(200 * diff_cnt)

        # Weekend nights: ogni dottore al massimo 2 weekend nights (Sat+Sun)
        # e distribuzione equa (minimizza massimo)
        _j_wex = {norm_name(d) for d in (rJ.get("weekend_excluded_doctors") or ["Calabrò"])}
        weekend_docs = night_pool - _j_wex
        # Calcola il cap minimo feasible: ceil(notti_weekend_totali / medici_disponibili)
        import math as _math
        total_we_nights = sum(
            1 for day in days if day.dow in ["Sat", "Sun"]
            if any(night_var_by_day_doc.get((day.date, doc)) is not None for doc in weekend_docs)
        )
        n_we_docs = len(weekend_docs)
        min_feasible_cap = _math.ceil(total_we_nights / n_we_docs) if n_we_docs > 0 else 2
        we_hard_cap = max(2, min_feasible_cap)
        we_cnt_vars = []
        for doc in sorted(weekend_docs):
            we_vars = []
            for day in days:
                if day.dow in ["Sat", "Sun"]:
                    v = night_var_by_day_doc.get((day.date, doc))
                    if v is not None:
                        we_vars.append(v)
            if we_vars:
                we_cnt = model.NewIntVar(0, len(we_vars), f"we_night_{hash(doc)%10**6}")
                model.Add(we_cnt == sum(we_vars))
                model.Add(we_cnt <= we_hard_cap)  # cap dinamico: max(2, ceil(totale/pool))
                we_cnt_vars.append(we_cnt)
        if we_cnt_vars:
            we_max = model.NewIntVar(0, 10, "we_night_max")
            model.AddMaxEquality(we_max, we_cnt_vars)
            extra_obj.append(500 * we_max)  # minimizza il massimo fortemente
    # H monthly quotas Mon-Fri
    # MODIFICA 1: Grimaldi e Calabrò sono esclusi da H; ignora eventuali quote riferite a loro
    _h_df_pair = {norm_name("Grimaldi"), norm_name("Calabrò")}
    if "rules" in cfg and "H" in cfg["rules"]:
        mqH = cfg["rules"]["H"].get("monthly_quotas") or {}
        for key, q in mqH.items():
            # keys could be 'Grimaldi_mon_fri'
            m = re.match(r"(.+)_mon_fri", str(key).strip(), flags=re.I)
            if not m:
                continue
            doc = norm_name(m.group(1))
            if doc not in doctors:
                continue
            if doc in _h_df_pair:
                continue   # Grimaldi/Calabrò non vanno mai in H
            vars_ = []
            for day in days:
                if day.dow in ["Mon","Tue","Wed","Thu","Fri"]:
                    sH = next((s for s in slots_by_day[day.date] if s.columns == ["H"]), None)
                    if sH and (sH.slot_id, doc) in x:
                        vars_.append(x[(sH.slot_id, doc)])
            if vars_:
                model.Add(sum(vars_) == int(q))
        # cap per doctor for pool_mon_fri
        cap = cfg["rules"]["H"].get("cap_mon_fri_per_doctor")
        if cap is not None:
            cap = int(cap)
            pool_cap = [norm_name(d) for d in (cfg["rules"]["H"].get("pool_mon_fri") or [])]
            for doc in pool_cap:
                if doc not in doctors:
                    continue
                vars_ = []
                for day in days:
                    if day.dow in ["Mon","Tue","Wed","Thu","Fri"]:
                        sH = next((s for s in slots_by_day[day.date] if s.columns == ["H"]), None)
                        if sH and (sH.slot_id, doc) in x:
                            vars_.append(x[(sH.slot_id, doc)])
                if vars_:
                    model.Add(sum(vars_) <= cap)
    # L quota Recupero
    if "rules" in cfg and "L" in cfg["rules"]:
        qrec = cfg["rules"]["L"].get("quota_recupero_per_month")
        if qrec is not None and "Recupero" in doctors:
            vars_=[]
            for s in slots:
                if s.columns == ["L"] and (s.slot_id, "Recupero") in x:
                    vars_.append(x[(s.slot_id, "Recupero")])
            if vars_:
                # Hard cap (<=) and soft preference to reach the target.
                qrec_int = int(qrec)
                model.Add(sum(vars_) <= qrec_int)
                # soft: minimize shortfall (qrec_int - sum(vars_))
                short = model.NewIntVar(0, qrec_int, f"L_rec_short")
                model.Add(sum(vars_) + short == qrec_int)
                extra_obj.append(5 * short)

    
    # ------------------------------------------------------------
    # Vincoli richiesti aggiuntivi (Roberto)
    # ------------------------------------------------------------
    # Y (Ambulatori) – due lunedì/mese: Recupero deve risultare come affiancamento,
    # ma SENZA creare uno slot extra (per evitare di "consumare" un medico in più).
    # Implementazione: imponiamo che in esattamente 2 lunedì/mese T=Recupero;
    # in output, quando T=Recupero di lunedì, aggiungiamo "Recupero" in Y come seconda riga.
    if "rules" in cfg and "Y" in cfg["rules"] and "T" in cfg["rules"]:
        rY = cfg["rules"]["Y"] or {}
        if (
            rY.get("recupero_two_mondays_per_month", False)
            and rY.get("recupero_affianca_in_T", False)
            and "Recupero" in doctors
        ):
            # Fixed choice: first 2 Mondays of the month
            monday_days = [d for d in days if d.dow == "Mon"]
            monday_days.sort(key=lambda d: d.date)
            target_mondays = monday_days[:2]
            other_mondays = monday_days[2:]

            # Save for diagnostics/logging
            MonT_need_rec = len(target_mondays)
            MonT_target_dates = [d.date for d in target_mondays]

            # Force T=Recupero on target Mondays
            for d in target_mondays:
                sT = next((s for s in slots_by_day[d.date] if s.columns == ["T"]), None)
                if sT and (sT.slot_id, "Recupero") in x:
                    model.Add(x[(sT.slot_id, "Recupero")] == 1)
                else:
                    pre_solve_warnings.append(
                        f"Recupero non disponibile per T il {d.date} (lunedì target): vincolo ignorato"
                    )

            # For all other Mondays, forbid T=Recupero (so it happens on exactly 2 Mondays)
            for d in other_mondays:
                sT = next((s for s in slots_by_day[d.date] if s.columns == ["T"]), None)
                if sT and (sT.slot_id, "Recupero") in x:
                    model.Add(x[(sT.slot_id, "Recupero")] == 0)

    # U (Contr.PM) – Cimino esattamente N volte/mese (default: 2)
    if "rules" in cfg and "U" in cfg["rules"] and "Cimino" in doctors:
        rU = cfg["rules"]["U"] or {}
        exact = int(rU.get("cimino_exact_per_month", 0) or 0)
        if exact > 0:
            vars_ = []
            for s in slots:
                if s.columns == ["U"] and (s.slot_id, "Cimino") in x:
                    vars_.append(x[(s.slot_id, "Cimino")])
            if vars_:
                effective_exact = min(exact, len(vars_))
                model.Add(sum(vars_) == effective_exact)
                if effective_exact < exact:
                    pre_solve_warnings.append(
                        f"Cimino U: richiesti esattamente {exact} ma solo {len(vars_)} slot disponibili, ridotto a {effective_exact}."
                    )
            else:
                pre_solve_warnings.append(
                    f"Cimino non ha slot U disponibili questo mese: vincolo cimino_exact_per_month ignorato"
                )

    # MODIFICA 3: se lunedì V=Allegra allora U deve essere Crea o Dattilo (hard constraint nel solver)
    if "rules" in cfg and "U" in cfg["rules"] and "V" in cfg["rules"]:
        rU_c = cfg["rules"]["U"] or {}
        if rU_c.get("v_allegra_monday_constraint", False):
            _allegra = norm_name("Allegra")
            _crea = norm_name("Crea")
            _dattilo = norm_name("Dattilo")
            _forbidden_in_u_if_v_allegra = [d for d in doctors if norm_name(d) not in {_crea, _dattilo, _allegra}]
            for day in [d for d in days if d.dow == "Mon"]:
                sV = next((s for s in slots_by_day[day.date] if s.columns == ["V"]), None)
                sU = next((s for s in slots_by_day[day.date] if s.columns == ["U"]), None)
                if sV is None or sU is None:
                    continue
                v_allegra = x.get((sV.slot_id, _allegra))
                if v_allegra is None:
                    continue
                # Se V=Allegra il lunedì → U deve essere Crea o Dattilo
                for forb in _forbidden_in_u_if_v_allegra:
                    u_forb = x.get((sU.slot_id, forb))
                    if u_forb is not None:
                        # u_forb=0 when v_allegra=1
                        model.Add(u_forb == 0).OnlyEnforceIf(v_allegra)

    # I (Cardiologia pomeriggio) – De Gregorio max N nei feriali (i festivi sono HI, quindi esclusi)
    if "rules" in cfg and "I" in cfg["rules"] and "De Gregorio" in doctors:
        rI = cfg["rules"]["I"] or {}
        max_i = int(rI.get("degregorio_max_weekdays", 0) or 0)
        if max_i > 0:
            vars_ = []
            for s in slots:
                # Count only real I slots on Mon-Sat. Festivi use the unified HI slot.
                if (
                    s.columns == ["I"]
                    and getattr(s.day, "dow", "") in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
                    and (s.slot_id, "De Gregorio") in x
                ):
                    vars_.append(x[(s.slot_id, "De Gregorio")])
            if vars_:
                model.Add(sum(vars_) <= max_i)
# ------------------------------------------------------------
    # Vincoli specifici per 'Recupero' (richiesti):
    # - T (Interni): almeno N/mese (hard) + target soft (es. 5)
    # - Q (ECO base): massimo N/mese
    # ------------------------------------------------------------
    if "rules" in cfg and "T" in cfg["rules"]:
        rT = cfg["rules"]["T"] or {}
        if "Recupero" in doctors:
            min_rec = int(rT.get("recupero_min_per_month", 0) or 0)
            target_rec = int(rT.get("recupero_target_per_month", 0) or 0)
            target_pen = int(rT.get("recupero_target_penalty", 2000) or 2000)
            vars_ = []
            for s in slots:
                if s.columns == ["T"] and (s.slot_id, "Recupero") in x:
                    vars_.append(x[(s.slot_id, "Recupero")])
            if vars_:
                cnt = model.NewIntVar(0, len(vars_), "T_rec_cnt")
                model.Add(cnt == sum(vars_))
                if min_rec > 0:
                    effective_min_rec = min(min_rec, len(vars_))
                    model.Add(cnt >= effective_min_rec)
                    if effective_min_rec < min_rec:
                        pre_solve_warnings.append(
                            f"Recupero T min: richiesto min {min_rec} ma solo {len(vars_)} slot T disponibili, ridotto a {effective_min_rec}."
                        )
                if target_rec > 0:
                    short = model.NewIntVar(0, target_rec, "T_rec_target_short")
                    model.Add(cnt + short >= target_rec)
                    extra_obj.append(target_pen * short)

    if "rules" in cfg and "Q" in cfg["rules"]:
        rQ = cfg["rules"]["Q"] or {}
        if "Recupero" in doctors:
            max_rec = int(rQ.get("recupero_max_per_month", 0) or 0)
            if max_rec > 0:
                vars_ = []
                for s in slots:
                    if s.columns == ["Q"] and (s.slot_id, "Recupero") in x:
                        vars_.append(x[(s.slot_id, "Recupero")])
                if vars_:
                    model.Add(sum(vars_) <= max_rec)
        # Q: hard cap per ogni medico del pool per evitare dominanza
        # Ideale: round(n_Q_slots / pool_size) + 1
        q_pool_raw = [norm_name(d) for d in (rQ.get("pool") or []) if norm_name(d) in doctors]
        q_slots = [s for s in slots if s.columns == ["Q"]]
        if q_pool_raw and q_slots:
            import math as _math
            q_cap = _math.ceil(len(q_slots) / len(q_pool_raw)) + 1
            for doc in q_pool_raw:
                vars_ = [x[(s.slot_id, doc)] for s in q_slots if (s.slot_id, doc) in x]
                if vars_:
                    model.Add(sum(vars_) <= q_cap)

    # T: hard cap per ogni medico del pool per evitare dominanza (Cimino/D'Angelo/Recupero a 5)
    if "rules" in cfg and "T" in cfg["rules"]:
        rT2 = cfg["rules"]["T"] or {}
        t_pool_raw = [norm_name(d) for d in (rT2.get("pool") or []) if norm_name(d) in doctors]
        t_slots = [s for s in slots if s.columns == ["T"]]
        if t_pool_raw and t_slots:
            import math as _math
            t_cap = _math.ceil(len(t_slots) / len(t_pool_raw)) + 1
            for doc in t_pool_raw:
                vars_ = [x[(s.slot_id, doc)] for s in t_slots if (s.slot_id, doc) in x]
                if vars_:
                    model.Add(sum(vars_) <= t_cap)

    # W: cap Recupero per evitare che domini il turno
    if "rules" in cfg and "W" in cfg["rules"] and "Recupero" in doctors:
        rW = cfg["rules"]["W"] or {}
        w_rec_max = int(rW.get("recupero_max_per_month", 0) or 0)
        if w_rec_max > 0:
            vars_ = []
            for s in slots:
                if s.columns == ["W"] and (s.slot_id, "Recupero") in x:
                    vars_.append(x[(s.slot_id, "Recupero")])
            if vars_:
                model.Add(sum(vars_) <= w_rec_max)

    # ---------------------------------------------------------------
    # VINCOLI UNIVERSITARI — calcolati PRIMA delle assegnazioni generali
    # Zito, Dattilo, De Gregorio: monte ore = 60% degli ospedalieri.
    #
    #   working_days = giorni lun-sab del mese (es. marzo 2026 = 26)
    #   target = round(working_days × 0.6)  → es. 16
    #
    #   Pesi:
    #   - J (notte 12h) = 2 turni per Zito e Dattilo
    #   - Ogni altro turno operativo = 1 (incluso V)
    #   - C (Reperibilità) = NON conta (turno di guardia extra, non monte ore)
    #
    #   Vincoli:
    #   - CAP HARD (≤ target+1): non si può sforare il contratto
    #   - FLOOR SOFT (penalità se < target): il solver cerca di avvicinarsi
    #     senza rendere infeasible se il pool è ristretto (es. Zito solo Q/R/S/J)
    # ---------------------------------------------------------------
    UNIV_EXCLUDE_COLS = {"C"}   # solo Reperibilità esclusa
    gc_uni = gc.get("university_doctors") or {}
    uni_ratio = float(gc.get("university_ratio", 0.6))
    if gc_uni and uni_ratio > 0:
        working_days = sum(1 for d in days if d.dow in ["Mon","Tue","Wed","Thu","Fri","Sat"] and not is_festivo(d, cfg))
        target = round(working_days * uni_ratio)
        for doc_raw, doc_cfg in gc_uni.items():
            doc = norm_name(doc_raw)
            if doc not in doctors:
                continue
            night_double = bool((doc_cfg or {}).get("night_counts_double", False))
            counts_as_map = cfg.get("pool_counts_as") or {}
            weighted_terms = []
            for s in slots:
                v = x.get((s.slot_id, doc))
                if v is None:
                    continue
                col_letters = s.columns or []
                if counts_as_map:
                    # counts_as da pool_config: C=0 (non conta), J=2 (vale doppio), altri=1
                    weight = max((counts_as_map.get(c, 1) for c in col_letters), default=1)
                else:
                    # Retrocompatibilità: escludi C, usa night_counts_double per J
                    if any(c in UNIV_EXCLUDE_COLS for c in col_letters):
                        weight = 0
                    else:
                        is_night = "J" in col_letters
                        weight = 2 if (night_double and is_night) else 1
                if weight == 0:
                    continue
                weighted_terms.append(weight * v)
            if not weighted_terms:
                continue
            max_possible = len(weighted_terms) * 2
            uni_cnt = model.NewIntVar(0, max_possible, f"uni_cnt_{hash(doc)%10**6}")
            model.Add(uni_cnt == sum(weighted_terms))
            # CAP HARD: mai più di target+1 (contratto)
            model.Add(uni_cnt <= target + 1)
            # FLOOR SOFT: penalizza se sotto target
            under = model.NewIntVar(0, target, f"uni_under_{hash(doc)%10**6}")
            model.Add(under >= target - uni_cnt)
            model.Add(under >= 0)
            extra_obj.append(500 * under)

    # AB: MODIFICA 5 — giovedì BILANCIATO (nessuna preferenza per Crea), sabati HARD con Crea
    if "rules" in cfg and "AB" in cfg["rules"]:
        rAB = cfg["rules"]["AB"]
        # Giovedì: soft balance tra tutti i medici del pool AB
        ab_thu_slots = [s for s in slots if s.rule_tag == "AB"]
        ab_thu_pool = []
        _seen = set()
        for _d in (rAB.get("fallback_pool") or []):
            _dn = norm_name(_d)
            if _dn not in _seen and _dn in doctors:
                ab_thu_pool.append(_dn)
                _seen.add(_dn)
        if ab_thu_slots and ab_thu_pool:
            n_thu = len(ab_thu_slots)
            ab_max_per_doc = model.NewIntVar(0, n_thu, "AB_thu_max")
            for doc in ab_thu_pool:
                vars_ = [x[(s.slot_id, doc)] for s in ab_thu_slots if (s.slot_id, doc) in x]
                if vars_:
                    cnt_doc = model.NewIntVar(0, n_thu, f"AB_thu_cnt_{hash(doc)%10**6}")
                    model.Add(cnt_doc == sum(vars_))
                    model.Add(ab_max_per_doc >= cnt_doc)
            extra_obj.append(50 * ab_max_per_doc)  # minimizza il massimo → bilanciamento

        # Sabati: HARD con Crea (2 sabati/mese)
        sat_n = int(rAB.get("saturday_per_month", 0) or 0)
        sat_doc = norm_name(rAB.get("saturday_only_doctor") or "Crea")
        if sat_n > 0 and sat_doc in doctors:
            vars_=[]
            for s in slots:
                if s.rule_tag == "AB_SAT" and (s.slot_id, sat_doc) in x:
                    vars_.append(x[(s.slot_id, sat_doc)])
            if vars_:
                if bool(rAB.get("saturday_soft", False)):
                    # soft (fallback per evitare infeasible)
                    model.Add(sum(vars_) <= sat_n)
                    short = model.NewIntVar(0, sat_n, "AB_sat_short")
                    model.Add(sum(vars_) + short == sat_n)
                    extra_obj.append(int(rAB.get("saturday_shortfall_penalty", 10000)) * short)
                else:
                    # HARD: i 2 sabati devono essere SOLO con Crea
                    model.Add(sum(vars_) == sat_n)

    # Weekend full off: at least N Sat+Sun "full weekends off" per doctor.
    # By default this is a HARD constraint. If it makes the month infeasible,
    # you can set global_constraints.weekend_off_soft: true to make it a SOFT constraint
    # (the solver will minimize the number of missing weekends-off).
    min_weekends = int(gc.get("min_full_weekends_off_per_month", 0) or 0)
    weekend_exempt = set(norm_name(x) for x in (gc.get("weekend_off_exempt") or []))
    # NOTA: 'Recupero' è trattato come medico reale: rientra nel conteggio dei weekend-off.
    weekend_soft = bool(gc.get("weekend_off_soft", False))
    weekend_penalty = int(gc.get("weekend_off_penalty", 50) or 50)  # penalty per missing full weekend off
    weekend_shortfalls: List = []
    if min_weekends > 0:
        # find weekend pairs within available days
        date_set = {d.date for d in days}
        weekend_pairs = []
        for drow in days:
            if drow.dow == "Sat":
                sun = drow.date + dt.timedelta(days=1)
                # count only complete Sat+Sun pairs present in the template for this month
                if sun in date_set and DOW_MAP[sun.weekday()] == "Sun":
                    weekend_pairs.append((drow.date, sun))
        # For each doctor, create weekend_off[w,doc] bool
        weekend_off: Dict[Tuple[int, str], "cp_model.IntVar"] = {}
        for wi, (sat, sun) in enumerate(weekend_pairs):
            for doc in doctors:
                if doc in weekend_exempt:
                    continue
                b = model.NewBoolVar(f"wkoff_{wi}_{hash(doc)%10**6}")
                weekend_off[(wi, doc)] = b
                # If weekend_off=1 then doc has no assignment on sat and sun
                # NOTE: C (Reperibilità) is excluded — it's passive on-call,
                # not an active shift, so it doesn't consume a weekend off.
                for s in slots_by_day.get(sat, []):
                    if s.columns == ["C"]:
                        continue
                    if (s.slot_id, doc) in x:
                        model.Add(x[(s.slot_id, doc)] == 0).OnlyEnforceIf(b)
                for s in slots_by_day.get(sun, []):
                    if s.columns == ["C"]:
                        continue
                    if (s.slot_id, doc) in x:
                        model.Add(x[(s.slot_id, doc)] == 0).OnlyEnforceIf(b)
                # Reverse implication: if any assignment on sat or sun then b=0
                any_vars = []
                for s in slots_by_day.get(sat, []):
                    if s.columns == ["C"]:
                        continue
                    v = x.get((s.slot_id, doc))
                    if v is not None:
                        any_vars.append(v)
                for s in slots_by_day.get(sun, []):
                    if s.columns == ["C"]:
                        continue
                    v = x.get((s.slot_id, doc))
                    if v is not None:
                        any_vars.append(v)
                if any_vars:
                    # sum(any_vars)==0 -> b=1
                    model.Add(sum(any_vars) == 0).OnlyEnforceIf(b)
                    # if any assignment then b=0
                    for av in any_vars:
                        model.Add(b + av <= 1)
        for doc in doctors:
            if doc in weekend_exempt:
                continue
            vars_ = [weekend_off[(wi, doc)] for wi, _ in enumerate(weekend_pairs) if (wi, doc) in weekend_off]
            if vars_:
                if weekend_soft:
                    short = model.NewIntVar(0, len(weekend_pairs), f"wkshort_{hash(doc)%10**6}")
                    model.Add(sum(vars_) + short >= min_weekends)
                    weekend_shortfalls.append(short)
                else:
                    model.Add(sum(vars_) >= min_weekends)
    # Objectives (soft): fairness + maximize S dedicated + minimize weekend night concentration + prefer night spacing >=7
    objective_terms = []
    # Weekend-off soft penalties
    if weekend_shortfalls:
        objective_terms.append(weekend_penalty * sum(weekend_shortfalls))
    # fairness: minimize max assignments per doctor
    real_doctors = list(doctors)
    max_load = model.NewIntVar(0, 999, "max_load")
    for doc in real_doctors:
        vars_ = []
        for s in slots:
            v = x.get((s.slot_id, doc))
            if v is not None:
                vars_.append(v)
        if vars_:
            load = model.NewIntVar(0, 999, f"load_{hash(doc)%10**6}")
            model.Add(load == sum(vars_))
            model.Add(load <= max_load)
    objective_terms.append(max_load * 10)
    # Column-specific balancing (and optional soft caps) for specific columns.
    # If a rule has `balance: true`, or defines a `distribution_pool`, we try to balance assignments within that column.
    # If a rule defines `max_per_doctor: N`, we penalize assignments above N (soft cap).
    # NOTE: a hard cap is not enforced here because it can easily make the model infeasible when the pool is small;
    # the soft cap is still reported via resulting counts (and can be made hard by widening the pool).
    try:
        for col_key, rcol in (rules_map or {}).items():
            if not isinstance(rcol, dict):
                continue
            if not (rcol.get('balance') or rcol.get('distribution_pool') or rcol.get('max_per_doctor')):
                continue
            # consider only slots whose rule_tag matches this column key (avoid Festivo_HI etc.)
            slot_ids = [s.slot_id for s in slots if (getattr(s, 'rule_tag', '') == col_key)]
            if not slot_ids:
                continue
            # prefer explicit pools; otherwise infer from variables
            pool_raw = (rcol.get('pool') or rcol.get('distribution_pool') or rcol.get('allowed') or [])
            pool = []
            if isinstance(pool_raw, list):
                for p in pool_raw:
                    pn = norm_name(p)
                    if pn and pn in doc_to_idx :
                        pool.append(pn)
            if not pool:
                # fallback: any doctor that actually appears in some var for these slots
                pool = []
                for d in real_doctors:
                    for sid in slot_ids:
                        if x.get((sid, d)) is not None:
                            pool.append(d)
                            break
            if not pool:
                continue
            bal_w = int(rcol.get('balance_weight') or 40)
            cap = int(rcol.get('max_per_doctor') or 0)
            cap_pen = int(rcol.get('max_per_doctor_penalty') or 800)
            max_col = model.NewIntVar(0, len(slot_ids), f"max_{col_key}_load")
            for d in pool:
                vars_d = []
                for sid in slot_ids:
                    v = x.get((sid, d))
                    if v is not None:
                        vars_d.append(v)
                if not vars_d:
                    continue
                load_d = model.NewIntVar(0, len(slot_ids), f"load_{col_key}_{hash(d)%10**6}")
                model.Add(load_d == sum(vars_d))
                model.Add(load_d <= max_col)
                if cap > 0:
                    over = model.NewIntVar(0, len(slot_ids), f"over_{col_key}_{hash(d)%10**6}")
                    # over >= load_d - cap, over >= 0
                    model.Add(load_d - cap <= over)
                    model.Add(over >= 0)
                    objective_terms.append(over * cap_pen)
            objective_terms.append(max_col * bal_w)
    except Exception:
        # never fail scheduling due to a balance/cap config issue
        pass
    # Maximize number of Wednesdays with dedicated S assignment (if optional)
    s_slots = [s for s in slots if s.columns == ["S"]]
    s_dedicated = []
    for s in s_slots:
        vars_ = [x.get((s.slot_id, d)) for d in doctors if (s.slot_id, d) in x]
        vars_ = [v for v in vars_ if v is not None]
        if vars_:
            b = model.NewBoolVar(f"sfilled_{hash(s.slot_id)%10**6}")
            model.Add(sum(vars_) == 1).OnlyEnforceIf(b)
            model.Add(sum(vars_) == 0).OnlyEnforceIf(b.Not())
            s_dedicated.append(b)
    if s_dedicated:
        # subtract to maximize (minimize negative)
        objective_terms.append(-5 * sum(s_dedicated))
    # Prefer night spacing >=7: penalize gaps of 5 or 6
    preferred = int(gc.get("night_spacing_days_preferred", 7))
    if preferred > min_gap:
        penalties=[]
        for doc in real_doctors:
            for i, drow in enumerate(days):
                for k in range(min_gap, preferred):
                    j=i+k
                    if j>=len(days): break
                    v1=night_var_by_day_doc.get((drow.date, doc))
                    v2=night_var_by_day_doc.get((days[j].date, doc))
                    if v1 is not None and v2 is not None:
                        p = model.NewBoolVar(f"ngap_{hash(doc)%10**6}_{i}_{k}")
                        # p=1 if both nights, else 0
                        model.Add(v1 + v2 == 2).OnlyEnforceIf(p)
                        model.Add(v1 + v2 <= 1).OnlyEnforceIf(p.Not())
                        penalties.append(p)
        if penalties:
            objective_terms.append(3 * sum(penalties))
    # Weekend night concentration: già gestito nel blocco J sopra con hard max=2 e strong soft

    # ── Historical balance: soft constraints from previous months ────────────
    _hist = historical_stats or {}
    if _hist:
        HIST_NIGHT_PENALTY = 150
        HIST_FEST_PENALTY = 200

        # Notti J: penalizza medici con più notti storiche
        _rJ_cfg = (cfg.get("rules") or {}).get("J") or {}
        _mq_fixed_hist = set()
        for _k in (_rJ_cfg.get("monthly_quotas") or {}).keys():
            _mq_fixed_hist.add(norm_name(_k))
        try:
            _night_pool_hist = night_pool  # defined inside "J" block
        except NameError:
            _night_pool_hist = set()
        _night_pool_free_hist = [d for d in sorted(_night_pool_hist) if d not in _mq_fixed_hist]

        for _doc in _night_pool_free_hist:
            _j_data = _hist.get(_doc, {}).get("J", {})
            _hist_j_total = _j_data.get("total", 0) if isinstance(_j_data, dict) else 0
            if _hist_j_total <= 0:
                continue
            _vars = [night_var_by_day_doc.get((d.date, _doc)) for d in days
                     if night_var_by_day_doc.get((d.date, _doc)) is not None]
            if _vars:
                _cnt = model.NewIntVar(0, len(days), f"hist_j_{abs(hash(_doc)) % 10 ** 6}")
                model.Add(_cnt == sum(_vars))
                extra_obj.append(HIST_NIGHT_PENALTY * _hist_j_total * _cnt)

        # Domeniche J: penalizza notti domenicali storiche
        for _doc in sorted(_night_pool_hist):
            _j_data2 = _hist.get(_doc, {}).get("J", {})
            _hist_dom_j = _j_data2.get("domeniche", 0) if isinstance(_j_data2, dict) else 0
            if _hist_dom_j <= 0:
                continue
            _sun_vars = [night_var_by_day_doc.get((d.date, _doc))
                         for d in days if d.dow == "Sun"
                         if night_var_by_day_doc.get((d.date, _doc)) is not None]
            if _sun_vars:
                _sun_cnt = model.NewIntVar(0, len(_sun_vars), f"hist_sunj_{abs(hash(_doc)) % 10 ** 6}")
                model.Add(_sun_cnt == sum(_sun_vars))
                extra_obj.append(HIST_FEST_PENALTY * _hist_dom_j * _sun_cnt)

        # Festivi DEHI: penalizza medici con più festivi storici (solo pool senza quota fissa)
        HIST_DEHI_PENALTY = 200
        _rFest_cfg = (cfg.get("rules") or {}).get("Festivi") or {}
        _fest_fixed_hist = {norm_name(k) for k in (_rFest_cfg.get("quotas") or {}).keys()}
        _fest_pool_hist = [norm_name(d) for d in (_rFest_cfg.get("pool") or [])
                          if norm_name(d) in doctors and norm_name(d) != "Recupero"]
        _fest_pool_free_hist = [d for d in _fest_pool_hist if d not in _fest_fixed_hist]
        _festivo_slots_hist = [s for s in slots if s.rule_tag in ("Festivo_DE", "Festivo_HI")]

        if _fest_pool_free_hist and _festivo_slots_hist:
            for _doc in _fest_pool_free_hist:
                _hist_dehi = _hist.get(_doc, {}).get("_festivi_DE_HI", 0)
                if not isinstance(_hist_dehi, int):
                    _hist_dehi = 0
                if _hist_dehi <= 0:
                    continue
                _fvars = [x.get((s.slot_id, _doc)) for s in _festivo_slots_hist
                          if (s.slot_id, _doc) in x]
                if _fvars:
                    _fcnt = model.NewIntVar(0, len(_festivo_slots_hist),
                                            f"hist_dehi_{abs(hash(_doc)) % 10 ** 6}")
                    model.Add(_fcnt == sum(_fvars))
                    extra_obj.append(HIST_DEHI_PENALTY * _hist_dehi * _fcnt)

    model.Minimize(sum(objective_terms + extra_obj))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        # ── Diagnostica: retry senza vincoli di quota hard per identificare la causa ──
        _diag_hints: List[str] = []
        try:
            _model2 = cp_model.CpModel()
            _extra2: List = []
            _BLANK_PEN2 = 5_000_000
            _BLANK_CRIT2 = 50_000_000
            _CRIT2 = {"H", "I", "J", "Festivo_HI"}
            _x2: Dict = {}
            for s2 in slots:
                for d2 in (s2.allowed or []):
                    if norm_name(d2) in doc_to_idx:
                        _x2[(s2.slot_id, norm_name(d2))] = _model2.NewBoolVar(
                            f"x2_{hash(s2.slot_id)%10**7}_{hash(d2)%10**7}"
                        )
            for s2 in slots:
                _v2 = [_x2[(s2.slot_id, norm_name(d2))] for d2 in (s2.allowed or [])
                       if (s2.slot_id, norm_name(d2)) in _x2]
                if not _v2:
                    continue
                if s2.required:
                    _bb2 = _model2.NewBoolVar(f"bb2_{hash(s2.slot_id)%10**7}")
                    _model2.Add(sum(_v2) + _bb2 == 1)
                    _pen2 = _BLANK_CRIT2 if s2.rule_tag in _CRIT2 else _BLANK_PEN2
                    _extra2.append(_pen2 * _bb2)
                else:
                    _model2.Add(sum(_v2) <= 1)
            # one_per_day (semplificato, senza share slack)
            for _d2 in days:
                _ds2 = slots_by_day.get(_d2.date, [])
                _uslots2 = [s2 for s2 in _ds2 if not _slot_is_exempt_daily(s2)]
                for _doc2 in doctors:
                    _dv2 = [_x2[(s2.slot_id, _doc2)] for s2 in _uslots2 if (_x2.get((s2.slot_id, _doc2)))]
                    if _dv2:
                        _model2.Add(sum(_dv2) <= 1)
            # fixed assignments (hard)
            for _fa2 in (fixed_assignments or []):
                try:
                    _fc2 = str(_fa2.get("column","")).strip().upper()
                    _fd2 = date.fromisoformat(str(_fa2.get("date","")).strip())
                    _fdc2 = norm_name(str(_fa2.get("doctor","")).strip())
                    if _fc2 == "J":
                        continue
                    _sf2 = next((s2 for s2 in slots_by_day.get(_fd2, []) if _fc2 in (s2.columns or [])), None)
                    if _sf2 and (_x2.get((_sf2.slot_id, _fdc2))):
                        _model2.Add(_x2[(_sf2.slot_id, _fdc2)] == 1)
                except Exception:
                    pass
            _model2.Minimize(sum(_extra2) if _extra2 else _model2.NewIntVar(0, 0, "z2"))
            _solver2 = cp_model.CpSolver()
            _solver2.parameters.max_time_in_seconds = 15.0
            _solver2.parameters.num_search_workers = 4
            _status2 = _solver2.Solve(_model2)
            if _status2 in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                _diag_hints.append(
                    "Il modello BASE (solo slot + one_per_day + fixed) è FEASIBLE: "
                    "i vincoli di quota/spacing/university causano l'infeasibility."
                )
            elif _status2 == cp_model.INFEASIBLE:
                _diag_hints.append(
                    "INFEASIBLE anche senza quote/spacing: il problema è nella copertura base degli slot. "
                    "Controlla pool vuoti o fixed_assignments impossibili."
                )
            else:
                _diag_hints.append(f"Retry diagnostico: status={_status2} (timeout o unknown).")
        except Exception as _de:
            _diag_hints.append(f"Retry diagnostico fallito: {_de}")
        # Slot summary
        _slots_by_day_d: dict = {}
        for _s in slots:
            _slots_by_day_d.setdefault(_s.day.date.isoformat(), []).append(_s)
        _tight_days = []
        for _day_str, _day_slots in sorted(_slots_by_day_d.items()):
            _req = [_s for _s in _day_slots if _s.required]
            _empty = [_s for _s in _req if not _s.allowed]
            _very_small = [f"{'+'.join(_s.columns)}({len(_s.allowed)})" for _s in _req if 0 < len(_s.allowed) <= 2]
            if _empty or _very_small:
                _tight_days.append(f"{_day_str}: vuoti={['+'.join(_s.columns) for _s in _empty]} ristretti={_very_small}")
        _slots_diag = "; ".join(_tight_days) if _tight_days else "nessun slot vuoto"
        raise RuntimeError(
            f"{'; '.join(_diag_hints)} | Slot critici: {_slots_diag}"
        )
    # Identify required slots left blank (b_blank == 1) for diagnostics
    forced_blank_slots: List[str] = []
    for sid, bv in blank_required_vars.items():
        try:
            if solver.Value(bv) == 1:
                forced_blank_slots.append(sid)
        except Exception:
            pass

    assignment: Dict[str, Optional[str]] = {s.slot_id: None for s in slots}
    for s in slots:
        chosen = None
        for d in s.allowed:
            v = x.get((s.slot_id, d))
            if v is not None and solver.Value(v) == 1:
                chosen = d
                break
        assignment[s.slot_id] = chosen
    has_forced_blanks = bool(forced_blank_slots)
    stats = {
        "status": "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE",
        "objective": solver.ObjectiveValue(),
        "warnings": pre_solve_warnings,
    }
    if has_forced_blanks:
        stats["status"] = "PARTIAL"
        stats["forced_blank_slots"] = sorted(forced_blank_slots)
        stats.setdefault("warnings", []).append(
            f"{len(forced_blank_slots)} slot obbligatori lasciati vuoti (infeasible): "
            + ", ".join(sorted(forced_blank_slots))
        )

    # Diagnostics for the "2 lunedì" rule (do not affect feasibility).
    MonT_used_rec = 0
    if MonT_need_rec and MonT_target_dates:
        for d in MonT_target_dates:
            if assignment.get(f"{d}-T") == "Recupero":
                MonT_used_rec += 1
        stats["MonT_recupero"] = {
            "need": int(MonT_need_rec),
            "used": int(MonT_used_rec),
            "dates": [dd.isoformat() for dd in MonT_target_dates],
        }
        if MonT_used_rec != MonT_need_rec:
            stats.setdefault("warnings", []).append(
                f"T=Recupero su lunedì: attesi {MonT_need_rec}, ottenuti {MonT_used_rec}"
            )

    # Final layer: assign/reassign Reperibilità (C) using definitive rules
    assignment, cdiag = assign_reperibilita_C(cfg, days, slots, assignment)
    if isinstance(cdiag, dict):
        stats.update(cdiag)
    return assignment, stats
# -------------------------
# Fallback greedy (simple)
# -------------------------
def solve_greedy(cfg: dict, days: List[DayRow], slots: List[Slot]) -> Tuple[Dict[str, Optional[str]], Dict]:
    """
    Greedy with rule-aware priorities (fallback when OR-Tools is missing or after autorelax still infeasible).
    Goals:
    - always cover required slots when possible
    - enforce key HARD quotas (e.g., Recupero on L <= 3, Recupero on Y = 2 Mondays) and fixed-day rules
    - avoid obviously bad imbalances (especially nights)
    """
    # Rank shifts: Night first helps satisfy spacing/off constraints early
    shift_rank = {"Notte": 0, "Pomeriggio": 1, "Mattina": 2, "Any": 3}
    gc = cfg.get('global_constraints', {}) or {}
    daily_exempt_cols = {str(c).strip().upper() for c in (gc.get('daily_uniqueness_exempt_columns') or [])}
    def _slot_is_exempt_daily(s: Slot) -> bool:
        return any(str(c).strip().upper() in daily_exempt_cols for c in (s.columns or []))
    # Assign C (Reperibilità) as a final layer (so it never blocks core shifts)
    slots_non_c = [s for s in slots if s.columns != ['C']]
    
    # Sort: required first, then smallest candidate pool, then shift priority
    slots_sorted = sorted(
        slots_non_c,
        key=lambda s: (not s.required, len(s.allowed), shift_rank.get(s.shift, 9), s.rule_tag or "", s.slot_id),
    )
    assignment: Dict[str, Optional[str]] = {s.slot_id: None for s in slots}
    used_per_day: Dict[dt.date, Set[str]] = defaultdict(set)
    nights_by_doc: Dict[str, List[dt.date]] = defaultdict(list)
    load_total = Counter()
    load_by_tag: Dict[str, Counter] = defaultdict(Counter)
    gc = cfg.get("global_constraints", {}) or {}
    min_gap = int(gc.get("night_spacing_days_min", 5) or 5)
    night_off_next = bool((gc.get("night_off") or {}).get("next_day", True))
    # --- Hard quotas / caps from rules
    rules = cfg.get("rules", {}) or {}
    # L: Recupero cap (interpret as MAX, not exact)
    L_cap_rec = None
    if isinstance(rules.get("L"), dict):
        L_cap_rec = rules["L"].get("quota_recupero_per_month")
        L_cap_rec = int(L_cap_rec) if L_cap_rec is not None else None
    L_rec_used = 0
    # Y: Monday specialist clinics
    # - Y_MAIN: always 1 doctor among the main pool (rotation)
    # - Y_REC: optional second line (Recupero) on exactly 2 Mondays/month
    Y_need_rec = 0
    if isinstance(rules.get("Y"), dict) and rules["Y"].get("recupero_two_mondays_per_month", False):
        Y_need_rec = 2
    Y_rec_used = 0
    Y_rec_slots_total = sum(1 for s in slots if getattr(s, "rule_tag", "") == "Y_REC")
    Y_rec_slots_done = 0
    Y_pool_counts = Counter()
    # Y affiancamento Recupero su T: i primi 2 lunedì del mese (fallback greedy)
    MonT_need_rec = 0
    MonT_target_dates: Set[dt.date] = set()
    if (
        isinstance(rules.get("Y"), dict)
        and rules["Y"].get("recupero_two_mondays_per_month", False)
        and rules["Y"].get("recupero_affianca_in_T", False)
    ):
        MonT_need_rec = 2
        monday_dates = sorted([d.date for d in days if d.dow == "Mon"])
        MonT_target_dates = set(monday_dates[:2])
    MonT_used_rec = 0

    # Night equalization target (if divisible)
    night_pool = set()
    if isinstance(rules.get("J"), dict):
        rJ = rules["J"]
        night_pool |= {norm_name(d) for d in (rJ.get("pool_other") or [])}
        night_pool |= {norm_name(d) for d in (rJ.get("monthly_quotas") or {}).keys()}
        night_pool.discard("Recupero")
    night_slots_total = sum(1 for s in slots if s.columns == ["J"])
    night_target = None
    if night_pool and night_slots_total > 0 and night_slots_total % len(night_pool) == 0:
        night_target = night_slots_total // len(night_pool)
    def can_assign(s: Slot, doc: str) -> bool:
        # per-day uniqueness (ignore placeholder 'Recupero')
        if (not _slot_is_exempt_daily(s)) and doc in used_per_day[s.day.date]:
            return False
        # L cap for Recupero
        if s.columns == ["L"] and doc == "Recupero" and L_cap_rec is not None and L_rec_used >= L_cap_rec:
            return False
        # Y_REC quota for Recupero (avoid exceeding)
        if (getattr(s, "rule_tag", "") == "Y_REC") and doc == "Recupero" and Y_need_rec and Y_rec_used >= Y_need_rec:
            return False
        # Night spacing
        if s.columns == ["J"]:
            for prev in nights_by_doc[doc]:
                if abs((s.day.date - prev).days) < min_gap:
                    return False
        # Night off next day (if doc did night previous day)
        if night_off_next:
            prev_day = s.day.date - dt.timedelta(days=1)
            if prev_day in nights_by_doc[doc]:
                return False
        # K no consecutive days (if enabled in rules)
        if "K" in s.columns and isinstance(rules.get("K"), dict) and rules["K"].get("no_consecutive_days_same_doctor", False):
            prev_day = s.day.date - dt.timedelta(days=1)
            if assignment.get(f"{prev_day}-K") == doc:
                return False
        # D/F different is handled by per-day uniqueness + separate slots; ok.
        return True
    def score_candidate(s: Slot, doc: str) -> Tuple:
        """
        Lower is better.
        Prioritize:
        - Night balance (nights first, then total load)
        - Column-specific balance (e.g., Y rotation)
        - Total load balance
        """
        tag = s.rule_tag or ""
        if s.columns == ["J"]:
            night_cnt = len(nights_by_doc[doc])
            # keep closer to target if known
            tgt_pen = abs(night_cnt - (night_target or 0)) if night_target is not None else night_cnt
            return (tgt_pen, night_cnt, load_total[doc], doc.lower())
        if (s.rule_tag or "") == "Y_REC":
            return (0, load_total[doc], doc.lower())
        if (s.rule_tag or "") == "Y_MAIN":
            return (Y_pool_counts[doc], load_total[doc], doc.lower())
        # default: balance by tag then total
        return (load_by_tag[tag][doc], load_total[doc], doc.lower())
    def pick(s: Slot) -> Optional[str]:
        candidates = [d for d in s.allowed if can_assign(s, d)]
        if not candidates:
            return None
        # Forza/nega Recupero su T (lunedi) in base ai 2 lunedi target (fallback greedy)
        nonlocal MonT_used_rec
        if s.columns == ["T"] and getattr(s.day, "dow", "") == "Mon" and MonT_need_rec:
            if s.day.date in MonT_target_dates:
                if "Recupero" in candidates:
                    return "Recupero"
            else:
                # fuori dai 2 lunedi target: evita Recupero se ci sono alternative
                if "Recupero" in candidates and len(candidates) > 1:
                    candidates = [d for d in candidates if d != "Recupero"]

        # Force Recupero on Y_REC when needed and we're running out of Monday slots
        nonlocal Y_rec_slots_done, Y_rec_used
        if (getattr(s, "rule_tag", "") == "Y_REC") and Y_need_rec:
            remaining_after_this = (Y_rec_slots_total - (Y_rec_slots_done + 1))
            remaining_need = (Y_need_rec - Y_rec_used)
            if remaining_need <= 0:
                return None  # leave blank
            if remaining_need > remaining_after_this:
                # must use Recupero now (if possible)
                if "Recupero" in candidates:
                    return "Recupero"
                return None
            # otherwise, leave blank to keep flexibility
            return None
        candidates.sort(key=lambda d: score_candidate(s, d))
        return candidates[0]
    conflicts = []
    for s in slots_sorted:
        if (s.rule_tag or "") == "Y_REC":
            Y_rec_slots_done += 1
        chosen = pick(s)
        if chosen is None:
            if s.required:
                conflicts.append(f"UNFILLED required slot {s.slot_id} ({s.columns})")
            continue
        assignment[s.slot_id] = chosen
        if not _slot_is_exempt_daily(s):
            used_per_day[s.day.date].add(chosen)
        load_total[chosen] += 1
        load_by_tag[s.rule_tag or ""][chosen] += 1
        if s.columns == ["J"]:
            nights_by_doc[chosen].append(s.day.date)
        if s.columns == ["L"] and chosen == "Recupero":
            L_rec_used += 1
        if s.columns == ["T"] and getattr(s.day, "dow", "") == "Mon" and chosen == "Recupero":
            MonT_used_rec += 1
        if (s.rule_tag or "") == "Y_REC":
            if chosen == "Recupero":
                Y_rec_used += 1
        if (s.rule_tag or "") == "Y_MAIN":
            Y_pool_counts[chosen] += 1
    # Final layer: assign/reassign Reperibilità (C) using definitive rules
    assignment, cdiag = assign_reperibilita_C(cfg, days, slots, assignment)
    # restore C slots as required, even if greedy did not assign them earlier
    stats = {
        "status": "GREEDY",
        "conflicts": conflicts,
        "loads": dict(load_total),
        "nights_per_doc": {k: len(v) for k, v in nights_by_doc.items()},
        "L_recupero_used": L_rec_used,
        "Y_recupero_used": Y_rec_used,
        "T_recupero_mondays_used": MonT_used_rec,
    }
    if isinstance(cdiag, dict):
        stats.update(cdiag)
    return assignment, stats
# -------------------------
# Write output Excel
# -------------------------
def write_output(
    wb: openpyxl.Workbook,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    days: List[DayRow],
    slots: List[Slot],
    assignment: Dict[str, Optional[str]],
    out_path: Path,
    cfg: Optional[dict] = None,
    unav_map: Optional[Dict[str, Dict[dt.date, Set[str]]]] = None,
):
    # Clear target columns (only those managed)
    managed_cols=set()
    for s in slots:
        managed_cols |= set(s.columns)
    # do not wipe A,B headers; wipe from row 2
    if cfg and isinstance(cfg.get("rules", {}), dict) and "AA" in (cfg.get("rules") or {}):
        managed_cols.add("AA")
    for drow in days:
        for col in managed_cols:
            ws[f"{col}{drow.row_idx}"].value = None
    # Fill (support multiple assignments on the same cell, e.g. Y_MAIN + Y_REC)
    slot_by_id = {s.slot_id: s for s in slots}
    assigned_by_day: Dict[dt.date, Set[str]] = defaultdict(set)
    cell_values: Dict[Tuple[int, str], List[str]] = defaultdict(list)
    # Iterate slots in creation order for stable writing
    for s in slots:
        doc = assignment.get(s.slot_id)
        if not doc:
            continue
        for col in s.columns:
            cell_values[(s.day.row_idx, col)].append(doc)
        assigned_by_day[s.day.date].add(doc)
    
    # Post-process: affiancamento Recupero in Y sui 2 lunedì fissi (i primi 2 del mese) – vedi vincolo su T.
    if cfg and isinstance(cfg.get("rules", {}), dict):
        rY = (cfg.get("rules") or {}).get("Y") or {}
        if rY.get("recupero_two_mondays_per_month", False) and rY.get("recupero_affianca_in_T", False):
            monday_dates = [drow.date for drow in days if getattr(drow, "dow", "") == "Mon"]
            monday_dates = sorted(monday_dates)[:2]
            target = set(monday_dates)
            for drow in days:
                if drow.date in target and assignment.get(f"{drow.date}-T") == "Recupero":
                    cell_values[(drow.row_idx, "Y")].append("Recupero")

    for (row_idx, col), docs in cell_values.items():
        # de-duplicate while preserving order
        seen = set()
        uniq = [d for d in docs if not (d in seen or seen.add(d))]
        # Y può legittimamente avere Recupero+medico (affiancamento): usa \n
        # V il venerdì ha legittimamente 2 medici (Crea + Dattilo/Allegra): usa \n
        # Tutte le altre colonne: deve esserci un solo medico; se ce ne sono due è un bug → primo
        if col in ("Y", "V"):
            ws[f"{col}{row_idx}"].value = "\n".join(uniq)
        else:
            ws[f"{col}{row_idx}"].value = uniq[0] if uniq else None

    # Ensure headers exist for new/optional columns (Z, AA) even on older templates.
    for _col, _label in [
        ("Z", ((cfg or {}).get("columns") or {}).get("Z", "Vascolare")),
        ("AA", ((cfg or {}).get("columns") or {}).get("AA", "SPOC")),
    ]:
        try:
            h = ws[f"{_col}1"]
            if h.value is None or str(h.value).strip() == "":
                h.value = _label
        except Exception:
            pass

    # Backward-compat: older templates may only have AD/AE (or no headers at all).
    # Ensure headers exist for the "medici liberi" block.
    for _col, _label in [
        ("AD", "Medici liberi 1"),
        ("AE", "Medici liberi 2"),
        ("AF", "Medici liberi 3"),
        ("AG", "Medici liberi 4"),
    ]:
        try:
            h = ws[f"{_col}1"]
            if h.value is None or str(h.value).strip() == "":
                h.value = _label
        except Exception:
            pass
    # Fill medici liberi 1/2/3/4 (AD/AE/AF/AG)
    # Nota: le colonne AD/AE possono rimanere vuote. Se però inseriamo un nome,
    # deve essere un medico *disponibile* quel giorno (nessuna indisponibilità registrata).
    # Base roster: unione dei medici noti da YAML (pools/fixed/unavailability),
    # NON dipende dal dominio dei singoli slot (che può diventare vuoto per indisponibilità).
    all_docs = set(collect_doctors(cfg))
    # 'Recupero' è un medico a tutti gli effetti: può comparire anche tra i liberi.
    unav_map = unav_map or {}
    # Escludi d'ufficio lo SMONTANTE NOTTE: chi ha fatto la NOTTE (colonna J) il giorno prima
    # non può essere considerato "libero" il giorno successivo (anche se non assegnato a nessuna colonna).
    night_by_date: Dict[dt.date, str] = {}
    for s in slots:
        if s.columns == ["J"]:
            nd = assignment.get(s.slot_id)
            if nd:
                night_by_date[s.day.date] = nd

    # Pre-calcola, per ogni giorno, quali fasce (M/P/N) ogni medico può effettivamente
    # coprire in base ai pool degli slot — esclude chi non è mai in un pool di notte, ecc.
    _SHIFT_TO_KEY = {"Mattina": "M", "Pomeriggio": "P", "Notte": "N",
                     "Diurno": None}  # Diurno → M e P
    doc_eligible_by_date: Dict[dt.date, Dict[str, Set[str]]] = {}
    for s in slots:
        s_date = s.day.date
        s_shift = getattr(s, "shift", "") or ""
        if s_shift == "Any":
            continue  # C reperibilità — ignora per il label
        keys: List[str] = []
        if s_shift == "Diurno":
            keys = ["M", "P"]
        elif s_shift in _SHIFT_TO_KEY and _SHIFT_TO_KEY[s_shift]:
            keys = [_SHIFT_TO_KEY[s_shift]]
        else:
            continue
        if s_date not in doc_eligible_by_date:
            doc_eligible_by_date[s_date] = {}
        for doc in (s.allowed or []):
            if doc not in doc_eligible_by_date[s_date]:
                doc_eligible_by_date[s_date][doc] = set()
            doc_eligible_by_date[s_date][doc].update(keys)

    # Per i giorni festivi con assegnazione forzata (sorteggio), s.allowed viene ridotto
    # a [solo_quel_medico], quindi gli altri medici del pool festivi spariscono da
    # doc_eligible_by_date → non appaiono in Medici liberi. Aggiungiamo il pool completo.
    if cfg and isinstance(cfg.get("rules"), dict):
        rFest = (cfg["rules"].get("Festivi") or {})
        _fp_m = [norm_name(d) for d in (rFest.get("pool_mattina") or rFest.get("pool") or []) if d]
        _fp_p = [norm_name(d) for d in (rFest.get("pool_pomeriggio") or rFest.get("pool") or []) if d]
        for drow in days:
            if not is_festivo(drow, cfg):
                continue
            fdate = drow.date
            if fdate not in doc_eligible_by_date:
                doc_eligible_by_date[fdate] = {}
            for doc in _fp_m:
                doc_eligible_by_date[fdate].setdefault(doc, set()).add("M")
            for doc in _fp_p:
                doc_eligible_by_date[fdate].setdefault(doc, set()).add("P")

    def _free_label(doc: str, unav_shifts: Set[str], eligible: Set[str]) -> Optional[str]:
        """Restituisce la stringa da scrivere in Medici liberi, o None se non disponibile.
        - eligible: fasce (M/P/N) per cui il medico è in almeno un pool in quel giorno
        - Nessuna indisponibilità e tutto eligible → solo il nome
        - Indisponibilità totale (Any / Tutto il giorno) → None
        - Indisponibilità parziale o pool parziale → "Nome (fasce_libere)"
        """
        if "Any" in unav_shifts or "Tutto il giorno" in unav_shifts:
            return None
        # Espandi fasce indisponibili
        unav_exp: Set[str] = set()
        for sh in unav_shifts:
            if sh == "Mattina":
                unav_exp.add("M")
            elif sh == "Pomeriggio":
                unav_exp.add("P")
            elif sh == "Notte":
                unav_exp.add("N")
            elif sh == "Diurno":
                unav_exp.update({"M", "P"})
        # Fasce effettivamente disponibili = eligible - indisponibili
        avail = [k for k in ["M", "P", "N"] if k in eligible and k not in unav_exp]
        if not avail:
            return None
        if avail == [k for k in ["M", "P", "N"] if k in eligible]:
            # Tutte le fasce eligibili sono disponibili
            if not unav_shifts:
                return doc  # nessuna indisponibilità → no suffisso
            # Ha indisponibilità ma non nelle fasce del suo pool → mostra comunque
            if eligible == {"M", "P", "N"}:
                return doc
        abbr_str = ", ".join(avail)
        # Mostra suffisso solo se non ha tutte e 3 le fasce o ha indisponibilità
        all_three = {"M", "P", "N"}
        if eligible >= all_three and not unav_shifts:
            return doc
        if set(avail) == eligible and not unav_shifts:
            return doc
        return f"{doc} ({abbr_str})"

    for drow in days:
        assigned_today = assigned_by_day.get(drow.date, set())
        smontante = night_by_date.get(drow.date - dt.timedelta(days=1))
        smontanti = {smontante} if smontante else set()
        _eligible_today = doc_eligible_by_date.get(drow.date, {})
        free_full: List[str] = []   # completamente disponibili
        free_partial: List[str] = []  # parzialmente disponibili (con suffisso)
        for doc in sorted(all_docs, key=lambda s: s.lower()):
            if doc in assigned_today or doc in smontanti:
                continue
            unav_shifts = unav_map.get(doc, {}).get(drow.date, set())
            eligible = _eligible_today.get(doc, set())
            if not eligible:
                continue  # medico senza pool attivo quel giorno → non compare
            label = _free_label(doc, unav_shifts, eligible)
            if label is None:
                continue
            if "(" in label:
                free_partial.append(label)
            else:
                free_full.append(label)
        free = free_full + free_partial
        ws[f"AD{drow.row_idx}"].value = free[0] if len(free) > 0 else None
        ws[f"AE{drow.row_idx}"].value = free[1] if len(free) > 1 else None
        ws[f"AF{drow.row_idx}"].value = free[2] if len(free) > 2 else None
        ws[f"AG{drow.row_idx}"].value = free[3] if len(free) > 3 else None

    # Fill AA (SPOC): solo Lun/Mer, copiando il medico di K o T. Il bilanciamento è fatto
    # in post-process scegliendo (quando K != T) il candidato meno usato nel mese.
    if cfg and "rules" in cfg and "AA" in (cfg.get("rules") or {}):
        rAA = (cfg.get("rules") or {}).get("AA") or {}
        copy_from = [str(c).strip().upper() for c in (rAA.get("copy_from") or ["K", "T"])]
        counts_by_month: Dict[Tuple[int, int], Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for drow in days:
            if not dayspec_contains(drow.dow, rAA.get("days")):
                continue
            candidates: List[str] = []
            for src in copy_from:
                try:
                    v = ws[f"{src}{drow.row_idx}"].value
                except Exception:
                    v = None
                if v is None:
                    continue
                name = str(v).splitlines()[0].strip()
                if name:
                    candidates.append(norm_name(name))
            # uniq preserve order
            seen=set()
            candidates=[c for c in candidates if c and not (c in seen or seen.add(c))]
            candidates=[c for c in candidates if c != "Recupero"]
            if not candidates:
                continue
            if len(candidates) == 1:
                chosen = candidates[0]
            else:
                mkey = (drow.date.year, drow.date.month)
                chosen = min(candidates, key=lambda c: (counts_by_month[mkey].get(c, 0), c.lower()))
            ws[f"AA{drow.row_idx}"].value = chosen
            mkey = (drow.date.year, drow.date.month)
            counts_by_month[mkey][chosen] += 1

    # AB: the template may pre-color AB cells on Saturdays. Since AB is only required
    # on Thursdays and on exactly N Saturdays (filled via the optional AB_SAT slot),
    # clear any pre-existing fill on Saturdays when AB is intentionally blank.
    try:
        no_fill = PatternFill()
        for drow in days:
            if getattr(drow, "dow", "") != "Sat":
                continue
            cell = ws[f"AB{drow.row_idx}"]
            v = cell.value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                cell.fill = no_fill
    except Exception:
        pass

    # Highlight "relief" blanks in yellow:
    # - slots that ended up with an empty allowed domain after applying unavailability
    # - penalized optional slots left blank by the solver (blank_penalty > 0)
    yellow_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    for s in slots:
        if assignment.get(s.slot_id) is not None:
            continue
        bp = int(getattr(s, "blank_penalty", 0) or 0)
        if bp <= 0:
            continue
        for col in (s.columns or []):
            cell = ws[f"{col}{s.day.row_idx}"]
            v = cell.value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                cell.fill = yellow_fill
    _write_riepilogo_sheet(wb, ws, days, slots, assignment, cfg)
    wb.save(out_path)

def _write_riepilogo_sheet(
    wb: openpyxl.Workbook,
    ws_main: openpyxl.worksheet.worksheet.Worksheet,
    days: List[DayRow],
    slots: List[Slot],
    assignment: Dict[str, Optional[str]],
    cfg: Optional[dict] = None,
) -> None:
    """Aggiunge/aggiorna il foglio 'Riepilogo' con conteggi turni per medico.

    I conteggi sono formule Excel (COUNTIFS) che si aggiornano automaticamente
    quando l'admin modifica nomi nel foglio principale e salva.

    Colonna helper nascosta: contiene flag festivo (1) / feriale (0) per ogni
    riga-giorno — statica (calcolata dalle date), non cambia con i nomi.

    Peso per giorno: J (notte) = 2 | C (reperibilità) non conta | altro = 1.
    Obiettivo universitari usa lo stesso peso ma con J=1 se night_counts_double=False.
    """
    cfg = cfg or {}
    col_map: Dict[str, str] = cfg.get("columns") or {}
    SKIP_COLS = {"AD", "AE", "AF", "AG"}
    op_cols = [c for c in col_map if c not in SKIP_COLS]

    # ── Festivi ────────────────────────────────────────────────────────────
    year = days[0].date.year if days else dt.date.today().year
    holidays = italy_public_holidays(year)
    extra_hol: Set[dt.date] = set()
    for _x in (cfg.get("festivi_extra") or []):
        try:
            extra_hol.add(parse_date(_x))
        except Exception:
            pass
    day_info = {d.date: d for d in days}

    def _is_fes(date_: dt.date) -> bool:
        d = day_info.get(date_)
        return d is not None and (d.dow == "Sun" or date_ in holidays or date_ in extra_hol)

    # ── University doctors config ──────────────────────────────────────────
    gc = cfg.get("global_constraints") or {}
    gc_uni = gc.get("university_doctors") or {}
    uni_ratio = float(gc.get("university_ratio", 0.6))
    pct = int(uni_ratio * 100)
    working_days_n = sum(1 for d in days if d.dow in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat") and not is_festivo(d, cfg))
    uni_target = round(working_days_n * uni_ratio)
    uni_docs = {norm_name(k) for k in gc_uni}
    uni_night_double = {
        norm_name(k): bool((v or {}).get("night_counts_double", False))
        for k, v in gc_uni.items()
    }

    # ── All assigned doctors (list determines Riepilogo rows) ─────────────
    all_docs: Set[str] = set()
    for s in slots:
        doc = assignment.get(s.slot_id)
        if doc:
            all_docs.add(doc)

    # ── Styles ────────────────────────────────────────────────────────────
    bold = openpyxl.styles.Font(bold=True)
    title_font = openpyxl.styles.Font(bold=True, size=13)
    hdr_fill = PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2")
    uni_fill = PatternFill(fill_type="solid", start_color="FFFFE2CC", end_color="FFFFE2CC")
    center = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

    mese_label = days[0].date.strftime("%B %Y") if days else ""
    n_hdr_cols = len(op_cols) + 5  # Medico + op_cols + Feriali + Festivi + Totale + Obiettivo

    # ── Create Riepilogo as second sheet (right after main sheet) ─────────
    sname = "Riepilogo"
    if sname in wb.sheetnames:
        del wb[sname]
    main_idx = wb.sheetnames.index(ws_main.title)
    ws = wb.create_sheet(sname, main_idx + 1)

    # ── Formula building blocks ────────────────────────────────────────────
    first_row = days[0].row_idx   # main sheet row of first day (= 2)
    last_row = days[-1].row_idx

    # Escape sheet name for cross-sheet formula reference
    ms_name = ws_main.title
    ms_ref = f"'{ms_name}'" if any(c in ms_name for c in (" ", "'", "!", "[", "]")) else ms_name

    # Hidden helper column: festivo flags at the same row indices as the main sheet.
    # Riepilogo!$AH$2:$AH$32  ←→  main sheet rows 2..32 (one per day).
    helper_col_idx = n_hdr_cols + 3
    hlp_letter = get_column_letter(helper_col_idx)
    hlp_range = f"Riepilogo!${hlp_letter}${first_row}:${hlp_letter}${last_row}"

    def _main_range(col: str) -> str:
        return f"{ms_ref}!${col}${first_row}:${col}${last_row}"

    def _cifs(col: str, dc: str, flag: int) -> str:
        """COUNTIFS: doctor in col filtered by feriale(0)/festivo(1).
        Wildcard (*name*) handles multi-doctor cells (e.g. V on Friday)."""
        return f'COUNTIFS({_main_range(col)},"*"&{dc}&"*",{hlp_range},{flag})'

    def _cif(col: str, dc: str) -> str:
        """COUNTIF (no flag): total occurrences of doctor in col."""
        return f'COUNTIF({_main_range(col)},"*"&{dc}&"*")'

    def col_display_formula(col: str, r: int) -> str:
        """Returns cell formula producing 'N+Mf' | 'N' | 'Mf' | ''."""
        dc = f"$A{r}"
        fer = _cifs(col, dc, 0)
        fes = _cifs(col, dc, 1)
        return (
            f'=IF(AND({fer}=0,{fes}=0),"",'
            f'IF({fes}=0,{fer},'
            f'IF({fer}=0,{fes}&"f",{fer}&"+"&{fes}&"f")))'
        )

    def weighted_formula(r: int, j_coeff: int, fes_flag: Optional[int]) -> str:
        """Numeric weighted-turni formula. C always excluded.
        Usa SUMPRODUCT+ISNUMBER(SEARCH()) per riga per evitare il doppio conteggio
        dei medici che coprono più colonne nello stesso giorno (D+F share, slot EG,
        DE/HI festivi, K+AA copia). Ogni giornata lavorativa conta 1, J conta j_coeff.
        j_coeff: coefficient for J (2 = standard, 1 = university without night_double).
        fes_flag: 0=feriali only, 1=festivi only, None=all."""
        dc = f"$A{r}"
        non_j_cols = [c for c in op_cols if c != "C" and c != "J"]
        has_j = "J" in op_cols

        # Filtro festivo/feriale (la colonna helper contiene 0=feriale, 1=festivo)
        if fes_flag is None:
            filt = ""
        else:
            filt = f"*({hlp_range}={fes_flag})"

        parts = []

        # Colonne non-J: presenza per riga → 1 per giornata lavorativa
        # (deduplica slot multi-colonna: D+F share, EG, DE festivo, HI festivo, K+AA, ecc.)
        if non_j_cols:
            search_terms = "+".join(
                f"ISNUMBER(SEARCH({dc},{_main_range(c)}))"
                for c in non_j_cols
            )
            parts.append(f"SUMPRODUCT((({search_terms})>0)*1{filt})")

        # Colonna J: presenza per riga × j_coeff
        if has_j:
            parts.append(f"SUMPRODUCT(ISNUMBER(SEARCH({dc},{_main_range('J')}))*{j_coeff}{filt})")

        return ("=" + "+".join(parts)) if parts else "=0"

    # ── Row 1: title ───────────────────────────────────────────────────────
    write_row = 1
    ws.cell(write_row, 1, f"Riepilogo Turni – {mese_label}").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_hdr_cols)

    # ── Row 2: headers ─────────────────────────────────────────────────────
    write_row = 2
    hdr = (
        ["Medico"]
        + [col_map.get(c, c) for c in op_cols]
        + ["Feriali (peso)", "Festivi (peso)", "Totale (peso)", "Obiettivo *"]
    )
    for ci, val in enumerate(hdr, 1):
        cell = ws.cell(write_row, ci, val)
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = center

    # Column indices for the summary columns
    fer_col_idx = len(op_cols) + 2
    fes_col_idx = fer_col_idx + 1
    tot_col_idx = fes_col_idx + 1
    obj_col_idx = tot_col_idx + 1
    fer_letter = get_column_letter(fer_col_idx)
    fes_letter = get_column_letter(fes_col_idx)

    # ── Rows 3+: one per doctor ────────────────────────────────────────────
    for doc in sorted(all_docs):
        write_row += 1
        r = write_row
        doc_n = norm_name(doc)
        is_uni = doc_n in uni_docs
        night_double = uni_night_double.get(doc_n, False)

        # A: doctor name — static, used by all formulas in this row via $A{r}
        ws.cell(r, 1, doc)

        # Per-column display: text formula "N+Mf" (updates live)
        for ci, col in enumerate(op_cols, 2):
            cell = ws.cell(r, ci)
            cell.value = col_display_formula(col, r)
            cell.alignment = center

        # Feriali (peso), Festivi (peso): numeric COUNTIFS formulas
        ws.cell(r, fer_col_idx).value = weighted_formula(r, j_coeff=2, fes_flag=0)
        ws.cell(r, fes_col_idx).value = weighted_formula(r, j_coeff=2, fes_flag=1)
        # Totale = Feriali + Festivi
        ws.cell(r, tot_col_idx).value = f"={fer_letter}{r}+{fes_letter}{r}"

        # Obiettivo: only for university doctors
        if is_uni:
            j_uni = 2 if night_double else 1
            uni_body = weighted_formula(r, j_coeff=j_uni, fes_flag=None)[1:]  # strip leading "="
            ws.cell(r, obj_col_idx).value = f'=TEXT({uni_body},"0")&"/{uni_target} ({pct}%)"'
            for ci in range(1, n_hdr_cols + 1):
                ws.cell(r, ci).fill = uni_fill

    # ── Notes ─────────────────────────────────────────────────────────────
    write_row += 2
    ws.cell(write_row, 1, "Note:").font = bold
    notes = [
        "  • Peso per giorno: J (notte) = 2 | C (reperibilità) non conta | altro = 1 per giornata lavorativa",
        "  • Turni multipli nella stessa giornata (es. D+F share, slot EG, festivi DE/HI, K+AA) contano 1",
        "  • Formato celle colonne: N = feriali  |  N+Mf = N feriali + M festivi  |  Mf = solo festivo",
        "  • Festivi = domeniche + festivi nazionali italiani",
        f"  • Giorni lavorativi lun-sab del mese: {working_days_n}",
        f"  • Obiettivo universitari (arancione): /{uni_target} = {working_days_n}×{pct}%"
        f"  (J={'2' if any(uni_night_double.values()) else '1'} se night_counts_double, C esclusa)",
        f"  • Universitari: {', '.join(sorted(uni_docs))}",
        "  ⚠ I conteggi si aggiornano automaticamente modificando i nomi nel foglio principale.",
    ]
    for note in notes:
        write_row += 1
        ws.cell(write_row, 1, note)

    # ── Helper column: festivo flags ───────────────────────────────────────
    # Written last to avoid interfering with write_row tracking.
    # Row alignment: day.row_idx == row index in this helper column
    # (both start at 2), so COUNTIFS ranges align correctly.
    for day in days:
        ws.cell(day.row_idx, helper_col_idx, 1 if _is_fes(day.date) else 0)
    ws.column_dimensions[hlp_letter].hidden = True

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    for i in range(2, n_hdr_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13


def write_solver_log(out_path: Path, stats: Dict) -> Optional[Path]:
    """
    Writes a human-readable solver log next to the output Excel.
    Includes: per-month status, objective, relief valves used, and day-level bottlenecks if any.
    """
    try:
        log_path = out_path.with_name(out_path.stem + "_solverlog.txt")
        lines: List[str] = []
        lines.append(f"Output: {out_path}")
        lines.append(f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}")
        lines.append(f"Overall status: {stats.get('status')}")
        months = stats.get("months") or {}
        for mk in sorted(months.keys()):
            sm = months.get(mk) or {}
            lines.append("")
            lines.append(f"== {mk} ==")
            lines.append(f"status: {sm.get('status')}")
            if "objective" in sm:
                lines.append(f"objective: {sm.get('objective')}")
            if sm.get("autorelax"):
                lines.append(f"autorelax: {sm.get('autorelax')}")
            if sm.get("solver_error"):
                lines.append(f"solver_error: {sm.get('solver_error')}")
            # Slot obbligatori lasciati bianchi (PARTIAL)
            fbs = sm.get("forced_blank_slots") or []
            if fbs:
                lines.append(f"ATTENZIONE: {len(fbs)} slot obbligatori NON coperti (infeasible parziale):")
                for sid in fbs:
                    lines.append(f"  - {sid}")
            # Relief used
            ru = sm.get("relief_used") or {}
            if ru.get("kt_share_days") or ru.get("blank_columns"):
                if ru.get("kt_share_days"):
                    lines.append("relief: K=T same doctor days: " + ", ".join(ru.get("kt_share_days")))
                bc = ru.get("blank_columns") or {}
                for c in sorted(bc.keys()):
                    lines.append(f"relief: blank {c} on: " + ", ".join(bc[c]))
            # Day-level bottlenecks (if OR-Tools failed at least once)
            bl = sm.get("day_level_bottlenecks") or []
            if bl:
                lines.append("")
                lines.append("Day-level bottlenecks (ignores cross-day constraints):")
                for item in bl[:10]:
                    lines.append(f"- {item.get('date')} ({item.get('dow')}): required_slots={item.get('required_slots')}, union_doctors={item.get('union_doctors')}")
                    for us in (item.get("unmatched_slots") or [])[:3]:
                        lines.append(f"    * {us.get('slot_id')} cols={us.get('columns')} allowed_n={us.get('allowed_n')}")
            # Day-by-day diagnostic
            dd = sm.get("daily_diagnostic") or []
            if dd:
                lines.append("")
                lines.append("== Diagnostica giorno per giorno ==")
                for item in dd:
                    tag = " [FESTIVO]" if item.get("festivo") else ""
                    issues_str = "; ".join(item.get("issues", []))
                    lines.append(f"{item.get('date')} ({item.get('dow')}){tag}: {issues_str}")
        log_path.write_text("\n".join(lines), encoding="utf-8")
        return log_path
    except Exception:
        return None
def solve_across_months(
    cfg: dict,
    days: List[DayRow],
    unav_map: Dict[str, Dict[dt.date, Set[str]]],
    carryover_by_month: Optional[dict] = None,
    fixed_assignments: Optional[List[dict]] = None,
    availability_preferences: Optional[List[dict]] = None,
    v_double_overrides: Optional[List[str]] = None,
    j_blank_week_overrides: Optional[Dict[str, Optional[str]]] = None,
    historical_stats: Optional[dict] = None,
) -> Tuple[List[Slot], Dict[str, Optional[str]], Dict]:
    """Solve schedules month-by-month and merge.

    `carryover_by_month` lets the caller inject cross-month constraints/history.
    `fixed_assignments`: list of {"doctor": str, "date": "YYYY-MM-DD", "column": str}
        Il medico DEVE comparire in quella colonna quel giorno (vincolo hard).
    `availability_preferences`: list of {"doctor": str, "date": "YYYY-MM-DD", "shift": str}
        Il software PROVA (soft) a far comparire il medico nella fascia indicata.
    """
    month_keys = sorted({(d.date.year, d.date.month) for d in days})
    slots_all: List[Slot] = []
    assignment_all: Dict[str, Optional[str]] = {}
    stats_all: Dict = {"status": "OK", "months": {}}

    gc = (cfg.get("global_constraints") or {})
    night_gap = int(gc.get("night_spacing_days_min", 5) or 5)
    night_off = (gc.get("night_off") or {}) if isinstance(gc.get("night_off"), dict) else {}
    night_off_next_day = bool(night_off.get("next_day", True))

    def _norm_key(yy: int, mm: int) -> str:
        return f"{yy}-{mm:02d}"

    def _parse_iso_date(s: str) -> Optional[dt.date]:
        try:
            return dt.date.fromisoformat(str(s).strip())
        except Exception:
            return None

    def _add_unav(local: Dict[str, Dict[dt.date, Set[str]]], doc: str, day: dt.date, shifts: Set[str]) -> None:
        docn = norm_name(doc)
        if not docn:
            return
        if docn not in local:
            local[docn] = {}
        if day not in local[docn]:
            local[docn][day] = set()
        local[docn][day].update(shifts)

    for (yy, mm) in month_keys:
        days_m = [d for d in days if (d.date.year, d.date.month) == (yy, mm)]
        mk = _norm_key(yy, mm)

        # Local copy of unavailability so we can inject carryover constraints without mutating input
        local_unav: Dict[str, Dict[dt.date, Set[str]]] = {k: {dk: set(sv) for dk, sv in v.items()} for k, v in (unav_map or {}).items()}

        carry = (carryover_by_month or {}).get(mk) if isinstance(carryover_by_month, dict) else None
        if carry and days_m:
            # Block day 1 entirely for doctors coming from previous-month last night
            for dname in (carry.get("blocked_day1_doctors") or []):
                _add_unav(local_unav, dname, days_m[0].date, {"Any"})

            # Enforce night spacing at the beginning of the month
            recent = carry.get("recent_nights_by_doc") or {}
            if isinstance(recent, dict):
                for dname, date_list in recent.items():
                    for ds in (date_list or []):
                        prev = _parse_iso_date(ds)
                        if not prev:
                            continue
                        for drow in days_m:
                            delta = (drow.date - prev).days
                            if 0 <= delta < night_gap:
                                _add_unav(local_unav, dname, drow.date, {"Notte"})
                            if night_off_next_day and delta == 1:
                                _add_unav(local_unav, dname, drow.date, {"Mattina", "Pomeriggio"})

        # Filtra fixed_assignments e availability_preferences per questo mese
        fixed_m = []
        for f in (fixed_assignments or []):
            d = _parse_iso_date(str(f.get("date", "")))
            if d is not None and d.year == yy and d.month == mm:
                fixed_m.append(f)

        slots_m = slots_for_month(cfg, days_m, local_unav, fixed_assignments=fixed_m, v_double_overrides=v_double_overrides, j_blank_week_overrides=j_blank_week_overrides)
        avail_m = []
        for a in (availability_preferences or []):
            d = _parse_iso_date(str(a.get("date", "")))
            if d is not None and d.year == yy and d.month == mm:
                avail_m.append(a)

        try:
            assignment_m, stats_m = solve_with_ortools(
                cfg, days_m, slots_m,
                fixed_assignments=fixed_m,
                availability_preferences=avail_m,
                unav_map=local_unav,
                historical_stats=historical_stats,
            )
        except Exception as e:
            # MODIFICA 2: niente Greedy. Se OR-Tools fallisce, propaga l'errore
            # e lascia le celle vuote (gestite dalla logica blank_penalty già esistente).
            stats_m = {
                "status": "INFEASIBLE",
                "solver_error": str(e),
                "note": "Greedy disabilitato: turni non coperti rimarranno vuoti (celle gialle).",
            }
            assignment_m = {}


        # Track where relief valves / blanks were used (useful for logs)
        try:
            stats_m = dict(stats_m or {})
            stats_m["relief_used"] = build_relief_log(days_m, slots_m, assignment_m)
        except Exception:
            pass
        # Day-by-day diagnostic
        try:
            stats_m["daily_diagnostic"] = build_daily_diagnostic(days_m, slots_m, assignment_m, cfg)
        except Exception:
            pass

        # Merge
        slots_all.extend(slots_m)
        assignment_all.update(assignment_m)
        stats_all["months"][mk] = stats_m
        st = str(stats_m.get("status", "")).upper()
        if "INFEAS" in st:
            stats_all["status"] = "INFEASIBLE"
        elif "PARTIAL" in st:
            stats_all["status"] = "PARTIAL"
        elif stats_all.get("status") == "OK" and "FEAS" in st:
            stats_all["status"] = "FEASIBLE"

    return slots_all, assignment_all, stats_all


def extract_carryover_from_output_xlsx(
    output_xlsx: "Path | str",
    sheet_name: Optional[str] = None,
    night_col_letter: str = "J",
    min_gap: int = 5,
) -> dict:
    """Extract carryover information from a previously generated output Excel.

    Returns:
      {
        "source_last_date": "YYYY-MM-DD",
        "night_last_day_doctor": "Name" or None,
        "blocked_day1_doctors": ["Name"] (0/1 element),
        "recent_nights_by_doc": {"Name": ["YYYY-MM-DD", ...]}
      }
    """
    p = Path(output_xlsx)
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Assume Date in column A with header in row 1
    dates = []
    night_docs = []

    col_night = openpyxl.utils.column_index_from_string(night_col_letter)

    for r in range(2, ws.max_row + 1):
        dv = ws.cell(r, 1).value
        if dv is None:
            continue
        # Convert to date
        if isinstance(dv, dt.datetime):
            d = dv.date()
        elif isinstance(dv, dt.date):
            d = dv
        else:
            try:
                # strings like 2026-02-01
                d = dt.date.fromisoformat(str(dv)[:10])
            except Exception:
                continue
        nd = ws.cell(r, col_night).value
        nd = norm_name(nd) if nd else None
        dates.append(d)
        night_docs.append(nd)

    if not dates:
        return {
            "source_last_date": None,
            "night_last_day_doctor": None,
            "blocked_day1_doctors": [],
            "recent_nights_by_doc": {},
        }

    # sort by date just in case
    combined = sorted(zip(dates, night_docs), key=lambda x: x[0])
    dates, night_docs = zip(*combined)
    last_date = dates[-1]
    last_night_doc = night_docs[-1]

    # recent window: last (min_gap-1) days (inclusive of last day)
    recent_n = max(int(min_gap) - 1, 0)
    window = list(zip(dates[-recent_n:] if recent_n else [], night_docs[-recent_n:] if recent_n else []))

    recent_by_doc: Dict[str, List[str]] = {}
    for d, doc in window:
        if not doc:
            continue
        recent_by_doc.setdefault(doc, []).append(d.isoformat())

    return {
        "source_last_date": last_date.isoformat(),
        "night_last_day_doctor": last_night_doc,
        "blocked_day1_doctors": [last_night_doc] if last_night_doc else [],
        "recent_nights_by_doc": recent_by_doc,
    }


def generate_schedule(
    template_xlsx: "Path | str",
    rules_yml: "Path | str",
    out_xlsx: "Path | str",
    unavailability_path: "Path | str | None" = None,
    sheet_name: "str | None" = None,
    carryover_by_month: Optional[dict] = None,
    fixed_assignments: Optional[List[dict]] = None,
    availability_preferences: Optional[List[dict]] = None,
    v_double_overrides: Optional[List[str]] = None,
    j_blank_week_overrides: Optional[Dict[str, Optional[str]]] = None,
    historical_stats: Optional[dict] = None,
    pool_config: Optional[dict] = None,
):
    """Generate schedules without Tkinter.

    This is the function used by Streamlit (and can be used programmatically).
    fixed_assignments: [{"doctor":str,"date":"YYYY-MM-DD","column":str}, ...]
    availability_preferences: [{"doctor":str,"date":"YYYY-MM-DD","shift":str}, ...]
    pool_config: dict caricato da pool_config_store (overlay JSON su YAML).
    """
    template = Path(template_xlsx)
    rules = Path(rules_yml)
    outp = Path(out_xlsx)
    unav = Path(unavailability_path) if unavailability_path else None

    cfg = load_rules(rules)
    if pool_config:
        cfg = apply_pool_config(cfg, pool_config)

    # Merge pre-assigned holiday shifts from data/turni_festivi.yml
    tf = load_turni_festivi()
    if tf["festivi_extra"]:
        existing_fe = list(cfg.get("festivi_extra") or [])
        existing_fe_set = set(str(x).strip() for x in existing_fe)
        cfg["festivi_extra"] = existing_fe + [x for x in tf["festivi_extra"] if x not in existing_fe_set]
    if tf["fixed_assignments"]:
        fixed_assignments = list(fixed_assignments or []) + tf["fixed_assignments"]

    wb, ws, days = load_template_days(template, sheet_name=sheet_name)
    unav_map = load_unavailability(unav)
    # Strip unavailability entries that conflict with pre-assigned festivo sorteggio shifts
    if tf["fixed_assignments"]:
        _strip_festivi_unavailability(unav_map, tf["fixed_assignments"])
    slots, assignment, stats = solve_across_months(
        cfg, days, unav_map,
        carryover_by_month=carryover_by_month,
        v_double_overrides=v_double_overrides,
        j_blank_week_overrides=j_blank_week_overrides,
        fixed_assignments=fixed_assignments,
        availability_preferences=availability_preferences,
        historical_stats=historical_stats,
    )
    # REMOVED: terza chiamata ridondante a assign_reperibilita_C (sovrascriveva C già ottimizzata)
    write_output(wb, ws, days, slots, assignment, cfg=cfg, out_path=outp, unav_map=unav_map)
    logp = write_solver_log(outp, stats)
    return stats, str(logp) if logp else None

# GUI (tkinter)
# -------------------------
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox
    root = tk.Tk()
    root.title("Turni Autogenerator (prototype)")
    template_var = tk.StringVar()
    rules_var = tk.StringVar()
    unav_var = tk.StringVar()
    out_var = tk.StringVar()
    def pick_file(var: tk.StringVar, types):
        p = filedialog.askopenfilename(filetypes=types)
        if p:
            var.set(p)
    def pick_save(var: tk.StringVar):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            var.set(p)
    def go():
        try:
            template = Path(template_var.get())
            rules = Path(rules_var.get())
            unav = Path(unav_var.get()) if unav_var.get().strip() else None
            outp = Path(out_var.get())
            if not template.exists() or not rules.exists() or not outp:
                raise ValueError("Seleziona template, regole e output.")
            cfg = load_rules(rules)
            wb, ws, days = load_template_days(template)
            unav_map = load_unavailability(unav)
            slots, assignment, stats = solve_across_months(cfg, days, unav_map)
            # Avviso se qualche mese è andato in fallback greedy
            greedy_months = [k for k,v in (stats.get('months') or {}).items() if isinstance(v, dict) and v.get('solver_error')]
            if greedy_months:
                messagebox.showwarning(
                    "Solver",
                    "OR-Tools non disponibile o schedule infeasible per: " + ", ".join(greedy_months) + "\nUso greedy per quei mesi."
                )
            write_output(wb, ws, days, slots, assignment, cfg=cfg, out_path=outp, unav_map=unav_map)
            logp = write_solver_log(outp, stats)
            msg = f"Creato: {outp}\nSolver: {stats.get('status')}"
            if logp:
                msg += f"\nLog: {logp}"
            messagebox.showinfo("OK", msg)
        except Exception as e:
            messagebox.showerror("Errore", str(e))
    frm = tk.Frame(root, padx=10, pady=10)
    frm.pack(fill="both", expand=True)
    def row(label, var, btn_text, cmd, r):
        tk.Label(frm, text=label, anchor="w").grid(row=r, column=0, sticky="w")
        tk.Entry(frm, textvariable=var, width=60).grid(row=r, column=1, padx=5)
        tk.Button(frm, text=btn_text, command=cmd).grid(row=r, column=2)
    row("Template turni (.xlsx)", template_var, "Scegli...", lambda: pick_file(template_var,[("Excel","*.xlsx")]), 0)
    row("Regole (.yml)", rules_var, "Scegli...", lambda: pick_file(rules_var,[("YAML","*.yml;*.yaml")]), 1)
    row("Indisponibilità (opz.)", unav_var, "Scegli...", lambda: pick_file(unav_var,[("Excel/CSV","*.xlsx;*.xls;*.csv;*.tsv")]), 2)
    row("Output (.xlsx)", out_var, "Salva come...", lambda: pick_save(out_var), 3)
    tk.Button(frm, text="Genera turni", command=go, height=2).grid(row=4, column=0, columnspan=3, pady=10, sticky="ew")
    root.mainloop()
# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser(description="Turni Autogenerator (UTIC/Cardiologia)")
    ap.add_argument("--template", type=str, help="Template Excel .xlsx")
    ap.add_argument("--rules", type=str, help="Regole YAML .yml/.yaml")
    ap.add_argument("--unavailability", type=str, default="", help="Indisponibilità mensili .xlsx/.csv (opzionale)")
    ap.add_argument("--out", type=str, help="Output Excel .xlsx")
    ap.add_argument("--sheet", type=str, default="", help="Nome foglio (opzionale)")
    ap.add_argument("--gui", action="store_true", help="Avvia interfaccia grafica (Tkinter)")
    args = ap.parse_args()
    if args.gui or (not args.template and not args.rules):
        run_gui()
        return
    if not args.template or not args.rules or not args.out:
        ap.error("In modalità CLI devi specificare: --template, --rules, --out")
    template = Path(args.template)
    rules = Path(args.rules)
    outp = Path(args.out)
    unav = Path(args.unavailability) if args.unavailability.strip() else None
    cfg = load_rules(rules)
    wb, ws, days = load_template_days(template, sheet_name=args.sheet if args.sheet else None)
    unav_map = load_unavailability(unav)
    slots, assignment, stats = solve_across_months(cfg, days, unav_map)
    # If any month fell back to greedy, print it
    greedy_months = [k for k,v in (stats.get('months') or {}).items() if isinstance(v, dict) and v.get('solver_error')]
    if greedy_months:
        print("[WARN] OR-Tools non disponibile o infeasible per:", ", ".join(greedy_months), file=sys.stderr)
    write_output(wb, ws, days, slots, assignment, cfg=cfg, out_path=outp, unav_map=unav_map)
    logp = write_solver_log(outp, stats)
    if logp:
        print(f"OK: creato {outp} | solver={stats.get('status')} | log={logp}")
    else:
        print(f"OK: creato {outp} | solver={stats.get('status')}")
if __name__ == "__main__":
    main()
