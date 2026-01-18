import io
import json
import platform
import tempfile
import time
import traceback
import uuid
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from turni_generator import generate_schedule


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _get_session_id() -> str:
    # Stable within a single browser session (reruns included)
    if "_session_id" not in st.session_state:
        st.session_state["_session_id"] = str(uuid.uuid4())
    return st.session_state["_session_id"]


def _summarize_stats(stats: dict | None) -> dict:
    """Return a compact summary suitable for audit logs."""
    if not isinstance(stats, dict):
        return {"status": "UNKNOWN"}

    months = stats.get("months") or {}
    month_summary = {}
    greedy_months = []
    infeasible_months = []

    for k, v in months.items():
        if not isinstance(v, dict):
            month_summary[k] = {"status": str(v)}
            continue
        st_m = str(v.get("status", "")).upper()
        se = v.get("solver_error")
        if se:
            greedy_months.append(k)
        if "INFEAS" in st_m:
            infeasible_months.append(k)
        month_summary[k] = {
            "status": v.get("status"),
            "solver_error": (str(se)[:400] if se else None),
            "autorelax": v.get("autorelax"),
        }

    return {
        "status": stats.get("status"),
        "greedy_months": greedy_months,
        "infeasible_months": infeasible_months,
        "months": month_summary,
    }


def _github_audit_log(event: dict) -> tuple[bool, str]:
    """Append an audit event as a comment to a GitHub Issue.

    Configure in Streamlit Secrets (Manage app → Settings → Secrets):

    [github]
    token = "..."
    repo  = "owner/repo"
    issue = 1

    Token needs permission to write Issue comments.
    """
    try:
        gh = st.secrets.get("github", {})
        token = gh.get("token") or st.secrets.get("GITHUB_TOKEN")
        repo = gh.get("repo")
        issue = gh.get("issue")
        if not (token and repo and issue):
            return False, "github audit log not configured"

        url = f"https://api.github.com/repos/{repo}/issues/{int(issue)}/comments"
        body = {
            "body": "```json\n" + json.dumps(event, ensure_ascii=False) + "\n```"
        }
        data = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(
            url,
            data=data,
            method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "User-Agent": "turni-autogen-streamlit",
                "Content-Type": "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return True, f"ok ({resp.status})"
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode("utf-8", errors="ignore")[:500]
        except Exception:
            detail = ""
        return False, f"HTTPError {getattr(e, 'code', '')} {detail}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"

st.set_page_config(page_title="Turni Autogenerator", layout="wide")

st.title("Turni Autogenerator – versione web (Streamlit)")

st.markdown(
    """
Questa versione gira **senza Tkinter** (Streamlit Cloud non supporta GUI desktop).

**Workflow:** carichi *Template Excel*, *Regole YAML* (o usi quelle del repo) e, opzionalmente, *Indisponibilità* → generi un **.xlsx** scaricabile.
"""
)

# --- Inputs ---
col1, col2 = st.columns(2)

with col1:
    template_up = st.file_uploader("Template turni (.xlsx)", type=["xlsx"], accept_multiple_files=False)
    # Evita input libero del nome foglio: è la causa più comune di KeyError su Streamlit Cloud.
    # Leggiamo i nomi reali dei fogli dal template e facciamo scegliere da dropdown.
    sheet_name = None
    if template_up is not None:
        try:
            wb_tmp = load_workbook(io.BytesIO(template_up.getvalue()), read_only=True, data_only=True)
            sheets = wb_tmp.sheetnames
        except Exception:
            sheets = []

        if sheets:
            st.caption("Fogli trovati nel template: " + ", ".join(sheets))
            opt = st.selectbox(
                "Seleziona foglio",
                options=["(foglio attivo / primo foglio)"] + sheets,
                index=0,
            )
            sheet_name = None if opt.startswith("(") else opt
        else:
            st.warning(
                "Non riesco a leggere i fogli del template: verrà usato il foglio attivo (primo foglio)."
            )
            sheet_name = None

with col2:
    use_repo_rules = st.checkbox("Usa Regole_Turni.yml del repo", value=True)
    rules_up = None
    if not use_repo_rules:
        rules_up = st.file_uploader("Regole (.yml/.yaml)", type=["yml", "yaml"], accept_multiple_files=False)

unav_up = st.file_uploader(
    "Indisponibilità (opzionale: .xlsx/.csv/.tsv)",
    type=["xlsx", "xls", "csv", "tsv"],
    accept_multiple_files=False,
)

st.divider()

# Optional operator tag (helps audit trails; leave blank if not needed)
operator_tag = st.text_input("Operatore (opzionale)", value="", help="Nome/codice di chi sta usando l'app (per audit log).")

run_btn = st.button("Genera turni", type="primary")

if run_btn:
    if template_up is None:
        st.error("Carica prima il **Template turni (.xlsx)**.")
        st.stop()

    if not use_repo_rules and rules_up is None:
        st.error("Hai disattivato 'Usa Regole_Turni.yml del repo': carica un file **Regole (.yml/.yaml)**.")
        st.stop()

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)

        # Save template
        template_path = td / "template.xlsx"
        template_path.write_bytes(template_up.getvalue())

        # Save rules
        if use_repo_rules:
            repo_rules = Path(__file__).with_name("Regole_Turni.yml")
            if not repo_rules.exists():
                st.error("Non trovo 'Regole_Turni.yml' nel repo. Carica un file regole manualmente.")
                st.stop()
            rules_path = td / "Regole_Turni.yml"
            rules_path.write_bytes(repo_rules.read_bytes())
        else:
            rules_path = td / "Regole_Turni.yml"
            rules_path.write_bytes(rules_up.getvalue())

        # Save unavailability (optional)
        unav_path = None
        if unav_up is not None:
            unav_path = td / f"unavailability.{unav_up.name.split('.')[-1]}"
            unav_path.write_bytes(unav_up.getvalue())

        out_path = td / "turni_output.xlsx"

        # --- Audit base ---
        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = _get_session_id()
        base_event = {
            "ts_utc": _utc_now_iso(),
            "run_id": run_id,
            "session_id": session_id,
            "operator": operator_tag.strip() or None,
            "template_filename": getattr(template_up, "name", None),
            "template_bytes": len(template_up.getvalue()) if template_up is not None else None,
            "sheet_name": sheet_name,
            "rules_source": "repo" if use_repo_rules else (getattr(rules_up, "name", None) if rules_up else None),
            "unavailability_filename": getattr(unav_up, "name", None) if unav_up is not None else None,
            "unavailability_bytes": len(unav_up.getvalue()) if unav_up is not None else None,
            "python": platform.python_version(),
        }

        t0 = time.time()
        try:
            with st.spinner("Calcolo in corso…"):
                stats, log_path = generate_schedule(
                    template_xlsx=template_path,
                    rules_yml=rules_path,
                    out_xlsx=out_path,
                    unavailability_path=unav_path,
                    sheet_name=sheet_name,
                )

            duration = round(time.time() - t0, 3)
            event = {
                **base_event,
                "result": "ok",
                "duration_s": duration,
                "stats": _summarize_stats(stats),
            }
            ok, msg = _github_audit_log(event)
            if not ok:
                st.caption(f"Audit log GitHub non scritto: {msg}")
        except Exception as e:
            duration = round(time.time() - t0, 3)
            event = {
                **base_event,
                "result": "error",
                "duration_s": duration,
                "error_type": type(e).__name__,
                "error": str(e),
                "traceback": traceback.format_exc()[:8000],
            }
            _github_audit_log(event)
            st.error("Errore durante la generazione dei turni.")
            st.code(event["traceback"])
            st.stop()

        # --- Results ---
        st.success("Turni generati.")

        # Show solver summary
        status = (stats or {}).get("status", "")
        st.subheader("Esito solver")
        st.write(f"**Status:** {status}")

        months = (stats or {}).get("months") or {}
        greedy_months = [k for k, v in months.items() if isinstance(v, dict) and v.get("solver_error")]
        if greedy_months:
            st.warning(
                "OR-Tools non disponibile o schedule infeasible per: "
                + ", ".join(greedy_months)
                + ". In quei mesi è stato usato il greedy."
            )

        # Download output
        st.download_button(
            label="Scarica Excel generato",
            data=out_path.read_bytes(),
            file_name="turni_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Show log (if any)
        if log_path and Path(log_path).exists():
            st.subheader("Log")
            try:
                st.code(Path(log_path).read_text(encoding="utf-8", errors="ignore"))
            except Exception:
                st.code(Path(log_path).read_text(errors="ignore"))
